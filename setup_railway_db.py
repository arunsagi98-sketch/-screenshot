"""
setup_railway_db.py
====================
Seeds the Railway PostgreSQL DB with:
  1. All tables (via SQLAlchemy create_all)
  2. app_url_reference data from Excel
  3. Default super_admin user (admin / Admin@123)

Run ONCE after creating Railway DB:
    python setup_railway_db.py
"""
import sys
import os
import pathlib
import psycopg2
from psycopg2.extras import execute_values

RAILWAY_URL = "postgresql://postgres:imGiWkhxTvTjtEGvcgyaPUKpJVrrDWjP@zephyr.proxy.rlwy.net:45185/railway"

EXCEL = pathlib.Path(r"C:\Users\HP\Desktop\$Screenshot\Fianl Site for automation (1)-4e173f79.xlsx")

BOXES = [
    (0, 1, "regular"),
    (4, 5, "important"),
    (8, 9, "more_important"),
]
SKIP_NAMES = {"summary"}
BAD_VALUES = {"", "none", "nan", "sites", "url", "n/a"}

# ── Step 1: Create all tables via SQLAlchemy ───────────────────────────────────
print("Step 1: Creating tables via SQLAlchemy...")
os.environ["DATABASE_URL"]     = RAILWAY_URL
os.environ["CRM_DATABASE_URL"] = RAILWAY_URL

sys.path.insert(0, str(pathlib.Path(__file__).parent / "Backend_Screenshot"))

from sqlalchemy import create_engine
from sqlalchemy.orm import declarative_base

engine = create_engine(RAILWAY_URL, pool_pre_ping=True)

# Import all models so metadata is populated
from database.db import Base as ScanBase
from database.crm_db import CrmBase
import models.screenshot       # noqa
import models.user             # noqa
import models.scan_screenshot  # noqa
import models.crm              # noqa

ScanBase.metadata.create_all(bind=engine)
CrmBase.metadata.create_all(bind=engine)
print("  ✅ Tables created")

# ── Step 2: Create app_url_reference table + seed data ────────────────────────
print("\nStep 2: Seeding app_url_reference...")
import openpyxl

conn = psycopg2.connect(RAILWAY_URL)
cur  = conn.cursor()

cur.execute("DROP TABLE IF EXISTS app_url_reference CASCADE;")
cur.execute("""
    CREATE TABLE app_url_reference (
        id          SERIAL PRIMARY KEY,
        sheet_name  VARCHAR(200) NOT NULL,
        url_id      INTEGER,
        url         VARCHAR(1000) NOT NULL,
        priority    VARCHAR(20)  NOT NULL DEFAULT 'regular'
    );
    CREATE INDEX idx_app_url_sheet    ON app_url_reference (sheet_name);
    CREATE INDEX idx_app_url_priority ON app_url_reference (sheet_name, priority);
""")
conn.commit()

if EXCEL.exists():
    wb = openpyxl.load_workbook(EXCEL, read_only=True, data_only=True)
    records = []
    sheet_count = 0

    for sname in wb.sheetnames:
        if sname.strip().lower() in SKIP_NAMES:
            continue
        ws = wb[sname]
        clean_name = sname.strip().upper()
        data_rows  = list(ws.iter_rows(values_only=True))[1:]

        for row in data_rows:
            for id_ci, url_ci, priority in BOXES:
                uid = row[id_ci]  if len(row) > id_ci  else None
                url = row[url_ci] if len(row) > url_ci else None
                if url is None:
                    continue
                url_str = str(url).strip()
                if url_str.lower() in BAD_VALUES:
                    continue
                try:
                    uid_int = int(float(uid)) if uid is not None else None
                except (ValueError, TypeError):
                    uid_int = None
                records.append((clean_name, uid_int, url_str, priority))
        sheet_count += 1

    wb.close()
    execute_values(cur,
        "INSERT INTO app_url_reference (sheet_name, url_id, url, priority) VALUES %s",
        records, page_size=500)
    conn.commit()
    print(f"  ✅ {len(records)} URLs inserted across {sheet_count} sheets")
else:
    print(f"  ⚠️  Excel not found at {EXCEL} — skipping app_url_reference seed")

# ── Step 3: Create default super_admin user ───────────────────────────────────
print("\nStep 3: Creating super_admin user...")
import bcrypt

hashed = bcrypt.hashpw(b"Admin@123", bcrypt.gensalt()).decode()
cur.execute("""
    INSERT INTO users (username, email, hashed_password, role, allowed_pages, is_active)
    VALUES (%s, %s, %s, %s, %s, %s)
    ON CONFLICT (username) DO NOTHING
""", ("admin", None, hashed, "super_admin", None, True))
conn.commit()
print("  ✅ super_admin created — username: admin | password: Admin@123")

cur.close()
conn.close()

print("\n" + "="*55)
print("  Railway DB setup complete!")
print("  Now set these on Render:")
print(f"  DATABASE_URL     = {RAILWAY_URL}")
print(f"  CRM_DATABASE_URL = {RAILWAY_URL}")
print("="*55)
