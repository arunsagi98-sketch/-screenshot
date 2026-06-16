"""
migrate_city_db.py
==================
Reads  Cleaned_Australia_Geo_Master_Manual.xlsx  and inserts every sheet
into a PostgreSQL  city_reference  table.

Usage
-----
    python migrate_city_db.py --excel "path/to/Cleaned_Australia_Geo_Master_Manual (2).xlsx"

Requirements
------------
    pip install openpyxl psycopg2-binary python-dotenv

PostgreSQL connection is read from environment variables (or a .env file):
    PGHOST, PGPORT, PGDATABASE, PGUSER, PGPASSWORD
    -- OR --
    DATABASE_URL=postgresql://user:pass@host:5432/dbname
"""

from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

import openpyxl
import psycopg2
from psycopg2.extras import execute_values

# ---------------------------------------------------------------------------
# PostgreSQL connection
# ---------------------------------------------------------------------------

def _get_conn():
    # Always use direct PostgreSQL credentials — ignore any SQLite DATABASE_URL
    return psycopg2.connect(
        host="localhost",
        port=5432,
        dbname="ctr_db",
        user="postgres",
        password="Arun@123$",
    )


# ---------------------------------------------------------------------------
# Schema
# ---------------------------------------------------------------------------

CREATE_TABLE_SQL = """
CREATE TABLE IF NOT EXISTS city_reference (
    id                   SERIAL PRIMARY KEY,
    sheet_name           VARCHAR(150)   NOT NULL,   -- Excel sheet name (state/region)
    row_number           INTEGER,
    city_name            VARCHAR(255)   NOT NULL,
    country              VARCHAR(10),
    state_region         VARCHAR(150),
    potential_impressions BIGINT,
    unique_cookies       BIGINT
);

CREATE INDEX IF NOT EXISTS idx_city_ref_sheet ON city_reference (sheet_name);
CREATE INDEX IF NOT EXISTS idx_city_ref_name  ON city_reference (LOWER(city_name));
"""


# ---------------------------------------------------------------------------
# Read Excel
# ---------------------------------------------------------------------------

def read_excel(path: str) -> dict[str, list[tuple]]:
    """
    Returns { sheet_name: [(row_no, city, country, state, pot_imp, uniq_cookies), ...] }
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    data: dict[str, list[tuple]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)          # skip header row
        if header is None:
            continue

        records = []
        for row in rows_iter:
            # Expected: (#, City, Country, State/Region, Potential Impressions, Unique Cookies)
            if len(row) < 6:
                continue
            row_no, city, country, state, pot_imp, uniq = row[:6]
            if not city or str(city).strip().lower() in ("", "nan", "none"):
                continue
            records.append((
                int(row_no)   if row_no   is not None else None,
                str(city).strip(),
                str(country).strip() if country else "AU",
                str(state).strip()   if state   else "",
                int(pot_imp)  if pot_imp  is not None else None,
                int(uniq)     if uniq     is not None else None,
            ))
        data[sheet_name] = records
        print(f"  [read]  {sheet_name!r:45s} → {len(records)} rows")

    wb.close()
    return data


# ---------------------------------------------------------------------------
# Insert into PostgreSQL
# ---------------------------------------------------------------------------

def migrate(excel_path: str, drop_first: bool = False) -> None:
    print(f"\n▶  Reading Excel: {excel_path}")
    all_data = read_excel(excel_path)

    print(f"\n▶  Connecting to PostgreSQL …")
    conn = _get_conn()
    cur  = conn.cursor()

    # Create table
    cur.execute(CREATE_TABLE_SQL)

    if drop_first:
        print("  [warn]  Truncating existing city_reference table …")
        cur.execute("TRUNCATE TABLE city_reference RESTART IDENTITY;")

    total_inserted = 0
    for sheet_name, records in all_data.items():
        if not records:
            continue

        # Check if this sheet already has data
        cur.execute("SELECT COUNT(*) FROM city_reference WHERE sheet_name = %s", (sheet_name,))
        existing = cur.fetchone()[0]
        if existing > 0 and not drop_first:
            print(f"  [skip]  {sheet_name!r} already has {existing} rows — skipping (use --drop to reload)")
            continue

        rows = [
            (sheet_name, r[0], r[1], r[2], r[3], r[4], r[5])
            for r in records
        ]
        execute_values(
            cur,
            """
            INSERT INTO city_reference
                (sheet_name, row_number, city_name, country, state_region,
                 potential_impressions, unique_cookies)
            VALUES %s
            """,
            rows,
            page_size=500,
        )
        total_inserted += len(rows)
        print(f"  [insert] {sheet_name!r:45s} → {len(rows)} rows inserted")

    conn.commit()
    cur.close()
    conn.close()

    print(f"\n✅  Done! Total rows inserted: {total_inserted}")
    print("    Table: city_reference")
    print("    Sheets loaded:", list(all_data.keys()))


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Migrate city Excel → PostgreSQL")
    parser.add_argument(
        "--excel",
        default="Cleaned_Australia_Geo_Master_Manual (2).xlsx",
        help="Path to the Excel file",
    )
    parser.add_argument(
        "--drop",
        action="store_true",
        help="Truncate the table before inserting (full reload)",
    )
    args = parser.parse_args()

    excel_path = Path(args.excel)
    if not excel_path.exists():
        print(f"❌  File not found: {excel_path}")
        sys.exit(1)

    migrate(str(excel_path), drop_first=args.drop)
