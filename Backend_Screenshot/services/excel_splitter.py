"""
Final Report — Excel Splitter Service
======================================
Implements the 7-step pipeline from the architecture diagram:

  1. INPUT          — accept .xlsx / .xls bytes
  2. EXCEL READER   — open workbook, iterate sheets
  3. SHEET PROCESS  — send each sheet to box detection
  4. BOX DETECTION  — find data regions (boxes) using empty-row/col gaps
  5. BLOCK COLLECT  — unify all boxes across all sheets into a flat block list
  6. BLOCK PROCESS  — determine output sheet name (B1 → Line Item → Sheet → Default)
  7. OUTPUT         — write one block = one sheet in a new workbook → BytesIO

Public API
----------
    split_excel(file_bytes: bytes, filename: str) -> io.BytesIO
"""
from __future__ import annotations

import io
import re
from typing import List, Optional, Tuple

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Types
# ---------------------------------------------------------------------------
# A Box is a rectangular region (1-indexed, inclusive) inside a sheet
Box = Tuple[int, int, int, int]   # (min_row, min_col, max_row, max_col)


# ---------------------------------------------------------------------------
# Step 4 — Box Detection
# ---------------------------------------------------------------------------

def _sheet_used_range(ws) -> Optional[Box]:
    """Return (min_row, min_col, max_row, max_col) of non-empty cells, or None."""
    min_r = min_c = float("inf")
    max_r = max_c = 0
    found = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and str(cell.value).strip() != "":
                found = True
                min_r = min(min_r, cell.row)
                min_c = min(min_c, cell.column)
                max_r = max(max_r, cell.row)
                max_c = max(max_c, cell.column)
    return (int(min_r), int(min_c), int(max_r), int(max_c)) if found else None


def _row_is_empty(ws, row_idx: int, min_col: int, max_col: int) -> bool:
    for col in range(min_col, max_col + 1):
        cell = ws.cell(row=row_idx, column=col)
        if cell.value is not None and str(cell.value).strip() != "":
            return False
    return True


def _col_is_empty(ws, col_idx: int, min_row: int, max_row: int) -> bool:
    for row in range(min_row, max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        if cell.value is not None and str(cell.value).strip() != "":
            return False
    return True


def detect_boxes(ws) -> List[Box]:
    """
    Detect data boxes in a worksheet using empty-row gaps.

    Algorithm:
      1. Find the overall used range.
      2. Split into horizontal bands separated by fully-empty rows.
      3. For each band, split into vertical segments by fully-empty columns.
      4. Each segment = one box.

    Returns list of (min_row, min_col, max_row, max_col).
    If no boxes detected → returns [] (caller treats sheet as single block).
    """
    used = _sheet_used_range(ws)
    if used is None:
        return []

    min_r, min_c, max_r, max_c = used

    # ── horizontal split: find empty rows ───────────────────────────────────
    row_bands: List[Tuple[int, int]] = []
    band_start = min_r
    for r in range(min_r, max_r + 2):   # +2 to close last band
        is_empty = (r > max_r) or _row_is_empty(ws, r, min_c, max_c)
        if is_empty:
            if band_start < r:
                row_bands.append((band_start, r - 1))
            band_start = r + 1

    if not row_bands:
        return []

    # ── vertical split within each band ─────────────────────────────────────
    boxes: List[Box] = []
    for (band_min_r, band_max_r) in row_bands:
        col_start = min_c
        for c in range(min_c, max_c + 2):
            is_empty = (c > max_c) or _col_is_empty(ws, c, band_min_r, band_max_r)
            if is_empty:
                if col_start < c:
                    boxes.append((band_min_r, col_start, band_max_r, c - 1))
                col_start = c + 1

    return boxes


# ---------------------------------------------------------------------------
# Step 5 — Block Collection
# ---------------------------------------------------------------------------

def collect_blocks(wb: Workbook) -> List[dict]:
    """
    Walk all sheets, detect boxes, return a flat list of block dicts:
      {
        sheet_name: str,
        box_index:  int,          # 0 = whole sheet
        is_whole:   bool,
        min_row, min_col, max_row, max_col: int,
        ws: worksheet reference,
      }
    """
    blocks: List[dict] = []
    for sheet_name in wb.sheetnames:
        ws    = wb[sheet_name]
        boxes = detect_boxes(ws)

        if not boxes:
            # No boxes → whole sheet is one block
            used = _sheet_used_range(ws)
            if used is None:
                continue   # completely empty sheet — skip
            min_r, min_c, max_r, max_c = used
            blocks.append({
                "sheet_name": sheet_name,
                "box_index":  0,
                "is_whole":   True,
                "min_row": min_r, "min_col": min_c,
                "max_row": max_r, "max_col": max_c,
                "ws": ws,
            })
        else:
            for idx, (min_r, min_c, max_r, max_c) in enumerate(boxes, start=1):
                blocks.append({
                    "sheet_name": sheet_name,
                    "box_index":  idx,
                    "is_whole":   False,
                    "min_row": min_r, "min_col": min_c,
                    "max_row": max_r, "max_col": max_c,
                    "ws": ws,
                })
    return blocks


# ---------------------------------------------------------------------------
# Step 6 — Block Processing: sheet name determination
# ---------------------------------------------------------------------------

_INVALID_CHARS = re.compile(r'[\\/*?\[\]:]')

def _clean_sheet_name(name: str) -> str:
    """Strip invalid Excel sheet-name characters, trim, limit to 31 chars."""
    name = _INVALID_CHARS.sub("", str(name)).strip()
    return name[:31] if name else ""


def _cell_value(ws, row: int, col: int) -> str:
    val = ws.cell(row=row, column=col).value
    return str(val).strip() if val is not None else ""


def _find_line_item_value(ws, min_row: int, min_col: int,
                           max_row: int, max_col: int) -> str:
    """
    Search for a 'Line Item' column header in the FIRST ROW of the block only.
    Returns the first data value found below that header.
    Limiting to the header row prevents data cells from accidentally matching.
    """
    for c in range(min_col, max_col + 1):
        val = _cell_value(ws, min_row, c)
        if "line item" in val.lower():
            # Value in the next row = first data row under this header
            below = _cell_value(ws, min_row + 1, c)
            if below:
                return below
    return ""


def determine_sheet_name(block: dict) -> tuple:
    """
    Two-format detection. Returns (clean_name, raw_name) where:
      clean_name — Excel-safe, max 31 chars (used as sheet tab name)
      raw_name   — full untruncated source value (used for Video/Banner detection)

    Format A — Line Item column format (flat table with 'Line Item' header):
      → Find 'Line Item' in the first row of the block, use the value below it.

    Format B — Box type format (multiple labelled boxes per sheet):
      → Use B1 of each box (first row, second column = the box label).

    Fallback: original sheet name → Sheet_N default.
    """
    ws    = block["ws"]
    min_r = block["min_row"]
    min_c = block["min_col"]
    max_r = block["max_row"]
    max_c = block["max_col"]

    # Format A: 'Line Item' header in first row → raw = full line-item value
    li_val = _find_line_item_value(ws, min_r, min_c, max_r, max_c)
    if li_val:
        return _clean_sheet_name(li_val), li_val

    # Format B: box sub-region → raw = full B1 value (untruncated)
    if not block["is_whole"]:
        b1 = _cell_value(ws, min_r, min_c + 1) if min_c + 1 <= max_c else ""
        if b1:
            return _clean_sheet_name(b1), b1

    # Fallback: original worksheet name
    orig = _clean_sheet_name(block["sheet_name"])
    return (orig or ""), block["sheet_name"]


def _unique_name(name: str, used: set) -> str:
    """Return name (or name_1, name_2…) that isn't in `used`."""
    if name not in used:
        return name
    i = 1
    while f"{name}_{i}" in used:
        i += 1
    return f"{name}_{i}"


# ---------------------------------------------------------------------------
# Step 7 — Output workbook
# ---------------------------------------------------------------------------

def _copy_block_to_sheet(src_ws, out_ws,
                          min_row: int, min_col: int,
                          max_row: int, max_col: int) -> None:
    """Copy cell values (and basic number formats) from block region to out_ws."""
    for r_idx, r in enumerate(range(min_row, max_row + 1), start=1):
        for c_idx, c in enumerate(range(min_col, max_col + 1), start=1):
            src_cell = src_ws.cell(row=r, column=c)
            dst_cell = out_ws.cell(row=r_idx, column=c_idx)
            dst_cell.value = src_cell.value
            if src_cell.number_format:
                dst_cell.number_format = src_cell.number_format

    # Approximate column widths
    for c_idx, c in enumerate(range(min_col, max_col + 1), start=1):
        col_letter = get_column_letter(c_idx)
        try:
            src_dim = src_ws.column_dimensions.get(get_column_letter(c))
            if src_dim and src_dim.width:
                out_ws.column_dimensions[col_letter].width = src_dim.width
        except Exception:
            pass


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def _prepare_blocks(file_bytes: bytes, filename: str) -> list:
    """Shared setup: load workbook, collect blocks, assign output names."""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext not in ("xlsx", "xls"):
        raise ValueError(f"Unsupported file type: .{ext}. Only .xlsx and .xls accepted.")

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=False)
    blocks = collect_blocks(wb)
    if not blocks:
        raise ValueError("No data found in the uploaded workbook.")

    used_names: set = set()
    default_counter = 1
    for block in blocks:
        name, raw_name = determine_sheet_name(block)
        if not name:
            name = f"Sheet_{default_counter}"
            default_counter += 1
        name = _unique_name(name, used_names)
        used_names.add(name)
        block["output_sheet_name"] = name
        block["raw_name"] = raw_name  # full untruncated name for type detection

    return blocks


def split_excel_to_files(file_bytes: bytes, filename: str) -> list:
    """
    Full pipeline → list of dicts:
      [{"sheet_name": str, "buf": BytesIO}, ...]

    One entry per detected block. Each buf is a single-sheet .xlsx file.
    """
    blocks = _prepare_blocks(file_bytes, filename)
    results = []
    for block in blocks:
        single_wb = Workbook()
        single_wb.remove(single_wb.active)
        out_ws = single_wb.create_sheet(title=block["output_sheet_name"])
        _copy_block_to_sheet(
            block["ws"], out_ws,
            block["min_row"], block["min_col"],
            block["max_row"], block["max_col"],
        )
        buf = io.BytesIO()
        single_wb.save(buf)
        buf.seek(0)
        results.append({
            "sheet_name": block["output_sheet_name"],
            "raw_name":   block.get("raw_name", block["output_sheet_name"]),
            "buf":        buf,
        })
    return results


def split_excel(file_bytes: bytes, filename: str) -> io.BytesIO:
    """
    Full pipeline:  bytes → combined workbook (one sheet per block) → BytesIO (.xlsx)
    """
    blocks = _prepare_blocks(file_bytes, filename)

    out_wb = Workbook()
    out_wb.remove(out_wb.active)
    for block in blocks:
        out_ws = out_wb.create_sheet(title=block["output_sheet_name"])
        _copy_block_to_sheet(
            block["ws"], out_ws,
            block["min_row"], block["min_col"],
            block["max_row"], block["max_col"],
        )

    buf = io.BytesIO()
    out_wb.save(buf)
    buf.seek(0)
    return buf
