"""
CRM Excel writer — builds the output .xlsx workbook from processed rows.

Public API:
    build_excel(output_rows) → io.BytesIO

Column type sets control how each value is written into the cell:
  - PERCENT_COLS  → divide by 100, apply "0.00%" number format
  - INTEGER_COLS  → cast to int
  - FLOAT_COLS    → cast to float
  - everything else → string
"""
from __future__ import annotations

import io
import math
from typing import Any, List

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from services.crm_processor import OUTPUT_COLUMNS

INTEGER_COLS = {
    "Advertiser ID",
    "Insertion Order ID",
    "Impressions",
    "Billable Impressions",
    "Clicks",
    "Start Views",
    "1st Quartile Views",
    "Midpoint Views",
    "3rd Quartile Views",
    "Complete Views",
    "Viewable Impressions",
    "Measurable Impressions",
    "For Checking (Measurable-Impression)",
    "Start Views-Impression",
}

FLOAT_COLS = {
    "Revenue (Adv Currency)",
    "Media Cost (Advertiser Currency)",
}

PERCENT_COLS = {
    "Click Rate (CTR)",
    "Video Completion Rate",
    "Viewability",
}


def _is_null(v: Any) -> bool:
    if v is None or v == "":
        return True
    if isinstance(v, float):
        try:
            return math.isnan(v)
        except Exception:
            return False
    return False


def build_excel(output_rows: List[dict]) -> io.BytesIO:
    """Convert processed row dicts into a styled .xlsx BytesIO buffer."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Processed"

    # Header row
    for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Data rows
    for row_idx, row in enumerate(output_rows, start=2):
        for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
            raw = row.get(col_name)

            if _is_null(raw):
                ws.cell(row=row_idx, column=col_idx, value=None)

            elif col_name in PERCENT_COLS:
                try:
                    num = float(str(raw).replace("%", "").strip()) / 100
                    cell = ws.cell(row=row_idx, column=col_idx, value=num)
                    cell.number_format = "0.00%"
                except Exception:
                    ws.cell(row=row_idx, column=col_idx, value=None)

            elif col_name in INTEGER_COLS:
                try:
                    ws.cell(row=row_idx, column=col_idx, value=int(float(str(raw))))
                except Exception:
                    ws.cell(row=row_idx, column=col_idx, value=None)

            elif col_name in FLOAT_COLS:
                try:
                    ws.cell(row=row_idx, column=col_idx, value=float(str(raw)))
                except Exception:
                    ws.cell(row=row_idx, column=col_idx, value=None)

            else:
                val = str(raw).strip() if raw is not None else None
                ws.cell(row=row_idx, column=col_idx, value=val or None)

    # Suppress "number stored as text" Excel warnings on the data range
    try:
        last_col = get_column_letter(len(OUTPUT_COLUMNS))
        last_row = max(len(output_rows) + 1, 2)
        ws.ignored_errors.append(
            {"sqref": f"A1:{last_col}{last_row}", "numberStoredAsText": True}
        )
    except Exception:
        pass

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
