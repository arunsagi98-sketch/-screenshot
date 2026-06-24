# -*- coding: utf-8 -*-
"""
report_generator.py
===================
Ported from Flask App.py → FastAPI service.
Generates a multi-sheet Excel report from campaign data.

Sheets produced:
  REACH / DATE / APP URL / TIME OF DAY / EXCHANGE / DEVICE /
  CREATIVE / CITY / AGE / GENDER
"""

from __future__ import annotations

import io
import math
import random
import re
from datetime import datetime, timezone
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.errors import IgnoredError

from database.crm_db import crm_engine

# ---------------------------------------------------------------------------
# DB connection (ctr_db — credentials come from .env via crm_engine)
# ---------------------------------------------------------------------------

def _get_conn():
    """Return a raw DBAPI connection from the shared crm_engine (no hardcoded creds)."""
    return crm_engine.raw_connection()


# ---------------------------------------------------------------------------
# DB helpers
# ---------------------------------------------------------------------------

def get_urls_from_sheets(sheet_names: list[str]) -> list[str]:
    """Return distinct cleaned URLs from app_url_reference for given sheets."""
    if not sheet_names:
        return []
    try:
        conn = _get_conn()
        cur  = conn.cursor()
        placeholders = ",".join(["%s"] * len(sheet_names))
        cur.execute(
            f"SELECT DISTINCT url FROM app_url_reference WHERE sheet_name IN ({placeholders}) ORDER BY url",
            sheet_names,
        )
        urls = [row[0] for row in cur.fetchall()]
        cur.close(); conn.close()
        return urls
    except Exception as e:
        print(f"[DB] get_urls_from_sheets failed: {e}")
        return []


def load_city_db_sheet(sheet_names: list[str]) -> list[dict]:
    """Return (city_name, weight) dicts from city_reference for given sheets."""
    if not sheet_names:
        return []
    try:
        conn = _get_conn()
        cur  = conn.cursor()
        placeholders = ",".join(["%s"] * len(sheet_names))
        cur.execute(
            f"""
            SELECT city_name, COALESCE(potential_impressions, unique_cookies, 1) AS weight
            FROM   city_reference
            WHERE  sheet_name IN ({placeholders})
            ORDER  BY weight DESC NULLS LAST
            """,
            sheet_names,
        )
        rows = cur.fetchall()
        cur.close(); conn.close()
        seen: set[str] = set()
        results = []
        for city_name, weight in rows:
            if city_name is None:
                continue
            name_str = str(city_name).strip()
            if not name_str or name_str.lower() in ("nan", "none"):
                continue
            key = name_str.lower()
            if key in seen:
                continue
            seen.add(key)
            results.append({"name": name_str, "weight": float(weight or 1.0)})
        return results
    except Exception as e:
        print(f"[DB] load_city_db_sheet failed: {e}")
        return []


# ---------------------------------------------------------------------------
# Math / formatting helpers
# ---------------------------------------------------------------------------

def pct(num: float, den: float, decimals: int = 2) -> str:
    if den == 0:
        return "0.00%"
    return f"{round((num / den) * 100, decimals):.{decimals}f}%"

def rand_float(mn: float, mx: float) -> float:
    return mn + random.random() * (mx - mn)

def rand_int(mn: int, mx: int) -> int:
    return random.randint(mn, mx)

def safe_float(val, default: float = 0.0) -> float:
    if val is None:
        return default
    try:
        if pd.isna(val):
            return default
    except Exception:
        pass
    s = str(val).strip().replace(",", "").replace(" ", "")
    if s.lower() in {"", "nan", "none", "<na>", "na", "null"}:
        return default
    if s.endswith("%"):
        try:
            out = float(s[:-1]) / 100.0
            return out if math.isfinite(out) else default
        except ValueError:
            return default
    try:
        out = float(s)
        return out if math.isfinite(out) else default
    except ValueError:
        return default

def safe_int(val, default: int = 0) -> int:
    try:
        out = safe_float(val, default)
        return int(out) if math.isfinite(out) else int(default)
    except (TypeError, ValueError, OverflowError):
        return int(default)

def serial_to_date(serial) -> str:
    try:
        serial = float(serial)
        utc_days = int(serial - 25569)
        import datetime as _dt
        dt = datetime(1970, 1, 1, tzinfo=timezone.utc) + _dt.timedelta(days=utc_days)
        return dt.strftime("%d %B, %Y")
    except Exception:
        return str(serial)

def _clean_url(u: str) -> str:
    u = str(u).lower().strip()
    for prefix in ("https://", "http://", "www."):
        if u.startswith(prefix):
            u = u[len(prefix):]
    return u.rstrip("/")

def _largest_remainder(weights: list[float], total_weight: float, total_count: int) -> list[int]:
    if not weights or total_count <= 0:
        return [0] * len(weights)
    if total_weight <= 0:
        total_weight = sum(weights) or 1.0
    fractions = [(w / total_weight) * total_count for w in weights]
    integers  = [int(f) for f in fractions]
    remainders = [(f - i, idx) for idx, (f, i) in enumerate(zip(fractions, integers))]
    remainders.sort(key=lambda x: x[0], reverse=True)
    gap = total_count - sum(integers)
    for i in range(int(gap)):
        integers[remainders[i][1]] += 1
    return integers

def _rescale_to_total(values: list[int], target: int) -> list[int]:
    """
    Hamilton largest-remainder rescale.
    Takes a list of non-negative ints and returns a new list that sums
    exactly to `target`, preserving proportions as closely as possible.
    """
    n = len(values)
    if n == 0:
        return []
    current = sum(values)
    if current == 0:
        base, rem = divmod(target, max(n, 1))
        return [base + (1 if i < rem else 0) for i in range(n)]
    ratio  = target / current
    scaled = [v * ratio for v in values]
    floors = [int(s) for s in scaled]
    order  = sorted(range(n), key=lambda i: scaled[i] - floors[i], reverse=True)
    diff   = target - sum(floors)
    for i in range(diff):
        floors[order[i]] += 1
    return floors


def deduplicate_preserving_sum(values: list[int], gap: int = 1) -> list[int]:
    _n = len(values)
    if _n < 2:
        return values
    items = [[i, v] for i, v in enumerate(values)]
    items.sort(key=lambda x: x[1], reverse=True)
    used: set[int] = set()
    for k in range(_n):
        val = items[k][1]
        if val not in used and all(abs(val - u) >= gap for u in used):
            used.add(val); continue
        found = False
        for _gap in sorted([gap, 5, 3, 2, 1], reverse=True):
            if _gap > gap: continue
            for _ in range(50):
                off  = random.randint(_gap, max(_gap + 5, _n))
                sign = random.choice([-1, 1])
                cand = val + sign * off
                if cand >= 1 and cand not in used and all(abs(cand - u) >= _gap for u in used):
                    items[k][1] = cand; used.add(cand); found = True; break
            if found: break
        if not found:
            off = 1
            while True:
                for cand in [val + off, val - off]:
                    if cand >= 1 and cand not in used:
                        items[k][1] = cand; used.add(cand); found = True; break
                if found: break
                off += 1
    drift = sum(values) - sum(x[1] for x in items)
    if drift:
        items.sort(key=lambda x: x[1], reverse=True)
        for k in range(_n):
            if not drift: break
            step = 1 if drift > 0 else -1
            cand = items[k][1] + step
            if cand >= 1 and cand not in used:
                items[k][1] = cand; used.add(cand); drift -= step
        if drift:
            items[0][1] += drift
    items.sort(key=lambda x: x[0])
    return [x[1] for x in items]


# ---------------------------------------------------------------------------
# Generation-time sanitizer  (runs BEFORE writing to Excel)
# ---------------------------------------------------------------------------

_SANITIZE_MAX_CTR = 0.009   # 0.90 % hard ceiling per row

def _sanitize_sheet_data(
    rows: list[dict],
    total_dict: dict,
    total_imp: int,
    total_clk: int,
    ctr_reach: str,
    imp_key: str = "Impressions",
    clk_key:  str = "Clicks",
    ctr_key:  str = "Click Rate (CTR)",
) -> tuple[list[dict], dict]:
    """
    Three-step clean-up applied to every sheet's row data immediately after
    the builder returns and BEFORE anything is written to the workbook.

    Step 1 — Impression sum fix
        Proportionally rescale row impressions so they sum exactly to
        total_imp (eliminates rounding drift from _largest_remainder /
        deduplicate_preserving_sum).

    Step 2 — Click sum fix
        Same proportional rescale for clicks → exact total_clk.

    Step 3 — Per-row CTR cap (0.90 %)
        If any row's CTR exceeds 0.90 %, clip its clicks and redistribute
        the excess to rows that still have headroom (same algorithm already
        used in build_sheet10_apps).  Recalculates per-row CTR string after.

    Step 4 — Grand Total dict override
        Always write total_imp / total_clk / ctr_reach into the GT dict so
        the Excel Grand Total row is always the exact reference value —
        never a =SUM() formula that openpyxl can't read back.

    Returns the mutated (rows, total_dict) tuple.
    """
    n = len(rows)
    if n == 0:
        # Empty sheet (e.g. CREATIVE with no data) — just fix GT dict
        if total_dict is not None:
            total_dict[imp_key] = total_imp
            total_dict[clk_key] = total_clk
            total_dict[ctr_key] = ctr_reach
        return rows, total_dict

    # ── Step 1: Fix impression sum ───────────────────────────────────────────
    imps = [int(r.get(imp_key, 0) or 0) for r in rows]
    imp_sum = sum(imps)
    if imp_sum > 0 and imp_sum != total_imp:
        ratio   = total_imp / imp_sum
        scaled  = [v * ratio for v in imps]
        floors  = [int(s) for s in scaled]
        order   = sorted(range(n), key=lambda i: scaled[i] - floors[i], reverse=True)
        diff    = total_imp - sum(floors)
        for i in range(diff):
            floors[order[i]] += 1
        imps = floors
        for i, r in enumerate(rows):
            r[imp_key] = imps[i]

    # ── Step 2: Fix click sum ────────────────────────────────────────────────
    clks = [int(r.get(clk_key, 0) or 0) for r in rows]
    clk_sum = sum(clks)
    if clk_sum > 0 and clk_sum != total_clk:
        ratio   = total_clk / clk_sum
        scaled  = [v * ratio for v in clks]
        floors  = [int(s) for s in scaled]
        order   = sorted(range(n), key=lambda i: scaled[i] - floors[i], reverse=True)
        diff    = total_clk - sum(floors)
        for i in range(diff):
            floors[order[i]] += 1
        clks = floors
        for i, r in enumerate(rows):
            r[clk_key] = clks[i]

    # ── Step 3: Cap per-row CTR at _SANITIZE_MAX_CTR ────────────────────────
    for _pass in range(20):
        excess = 0
        for i in range(n):
            cap = int(imps[i] * _SANITIZE_MAX_CTR)
            if clks[i] > cap:
                excess += clks[i] - cap
                clks[i] = cap
        if excess == 0:
            break
        headroom = [
            (int(imps[i] * _SANITIZE_MAX_CTR) - clks[i], i)
            for i in range(n)
            if int(imps[i] * _SANITIZE_MAX_CTR) > clks[i]
        ]
        if not headroom:
            clks[0] += excess
            break
        total_room  = sum(room for room, _ in headroom)
        distributed = 0
        for room, idx in headroom:
            share = round(excess * room / total_room)
            clks[idx] += share
            distributed += share
        leftover = excess - distributed
        if leftover:
            clks[headroom[0][1]] += leftover

    # Write corrected clicks + recalculate per-row CTR string
    for i, r in enumerate(rows):
        r[clk_key] = clks[i]
        r[ctr_key] = pct(clks[i], imps[i]) if imps[i] > 0 else "0.00%"

    # ── Step 4: Fix Grand Total dict to exact reference values ───────────────
    if total_dict is not None:
        total_dict[imp_key] = total_imp
        total_dict[clk_key] = total_clk
        total_dict[ctr_key] = ctr_reach

    return rows, total_dict


# ---------------------------------------------------------------------------
# Pre-write cross-sheet QC fix  (Step 2 — runs BEFORE Excel write)
# ---------------------------------------------------------------------------

def _pre_write_qc_fix(
    sheets: list,
    total_imp: int,
    total_clk: int,
    ctr_reach: str,
    imp_key: str = "Impressions",
    clk_key:  str = "Clicks",
    ctr_key:  str = "Click Rate (CTR)",
) -> list[str]:
    """
    Cross-sheet consistency check and fix on raw Python dicts.

    Runs AFTER all sheet builders + _sanitize_sheet_data(), and BEFORE any
    data is written to the Excel workbook.  This is the QC gate that prevents
    mismatches from ever reaching the file — no openpyxl formula-reading
    issues, no post-correct round-trips.

    For each sheet:
      1. Verify sum(row impressions) == total_imp  →  rescale if not
      2. Verify sum(row clicks)      == total_clk  →  rescale if not
      3. Recalculate per-row CTR string if anything changed
      4. Force GT dict to exact reference values (plain int, never formula)

    sheets : list of (name: str, rows: list[dict], total_dict: dict)
    Returns: list of fix descriptions — empty means all sheets were already clean.
    """
    fixes: list[str] = []

    for name, rows, total_dict in sheets:
        # Always force GT dict to exact reference values
        if total_dict is not None:
            total_dict[imp_key] = total_imp
            total_dict[clk_key] = total_clk
            total_dict[ctr_key] = ctr_reach

        if not rows:
            continue

        imps = [int(r.get(imp_key, 0) or 0) for r in rows]
        clks = [int(r.get(clk_key, 0) or 0) for r in rows]
        need_ctr_recalc = False

        # ── 1. Impressions ──────────────────────────────────────────────────
        imp_sum = sum(imps)
        if imp_sum != total_imp:
            imps = _rescale_to_total(imps, total_imp)
            for i, r in enumerate(rows):
                r[imp_key] = imps[i]
            fixes.append(f"[PRE-QC] {name}: Impressions {imp_sum:,} → {total_imp:,}")
            need_ctr_recalc = True

        # ── 2. Clicks ───────────────────────────────────────────────────────
        clk_sum = sum(clks)
        if clk_sum != total_clk:
            clks = _rescale_to_total(clks, total_clk)
            for i, r in enumerate(rows):
                r[clk_key] = clks[i]
            fixes.append(f"[PRE-QC] {name}: Clicks {clk_sum:,} → {total_clk:,}")
            need_ctr_recalc = True

        # ── 3. Recalculate per-row CTR if anything changed ──────────────────
        if need_ctr_recalc:
            for i, r in enumerate(rows):
                r[ctr_key] = pct(clks[i], imps[i]) if imps[i] > 0 else "0.00%"

    if fixes:
        print(f"[_pre_write_qc_fix] {len(fixes)} correction(s): {' | '.join(fixes)}")

    return fixes


# ---------------------------------------------------------------------------
# Excel styling
# ---------------------------------------------------------------------------

HEADER_BG = "00B0F0"
TOTAL_BG  = "9BC2E6"
_ALIGN_CACHE: dict[str, Alignment] = {}

def _thin_border() -> Border:
    s = Side(style="thin", color="000000")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(horizontal: str) -> Alignment:
    if horizontal not in _ALIGN_CACHE:
        _ALIGN_CACHE[horizontal] = Alignment(horizontal=horizontal, vertical="center")
    return _ALIGN_CACHE[horizontal]

def _style_header(ws, row: int, num_cols: int):
    fill = PatternFill("solid", fgColor=HEADER_BG)
    font = Font(bold=True, color="000000", size=11, name="Calibri")
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill; cell.font = font; cell.border = _thin_border()
    ws.row_dimensions[row].height = 14.50

def _style_total(ws, row: int, num_cols: int):
    fill = PatternFill("solid", fgColor=TOTAL_BG)
    font = Font(bold=True, color="000000", size=10, name="Calibri")
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill; cell.font = font; cell.border = _thin_border()

def _auto_fit(ws):
    for col in ws.columns:
        width = 10
        for idx, cell in enumerate(col):
            if idx >= 200: break
            width = max(width, len(str(cell.value or "")))
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 4, 50)

_NUMERIC_HEADERS = {
    "Impressions", "Clicks", "Reach", "Frequency",
    "Measurable Impressions", "Viewable Impressions",
    "Sum of Starts (Video)", "Sum of Complete Views (Video)",
}

def _is_numeric_like(val) -> bool:
    if val is None or isinstance(val, bool): return False
    if isinstance(val, (int, float)): return True
    s = str(val).strip().replace(",", "")
    if s.endswith("%"): s = s[:-1]
    try: float(s); return True
    except Exception: return False

def write_sheet(ws, headers: list[str], rows: list[dict],
                total_row: Optional[dict] = None,
                formulas: Optional[dict] = None,
                alignments=None):
    ws.sheet_view.showGridLines = True
    formulas   = formulas or {}
    alignments = alignments or {}
    col_map = {h: get_column_letter(i) for i, h in enumerate(headers, 1)}

    def _assign(cell, val, h_name=None, row_idx=None):
        if hasattr(val, "item"): val = val.item()
        if h_name in formulas and row_idx is not None:
            f = formulas[h_name]
            for href, let in col_map.items():
                f = f.replace(f"{{{href}}}", f"{let}{row_idx}")
            # Wrap division formulas with IFERROR so Impressions=0 → 0.00% not #DIV/0!
            cell.value = f"=IFERROR({f},0)"
            if h_name in ("CTR", "Click Rate (CTR)", "Viewability", "VCR (Completion Rate)"):
                cell.number_format = "0.00%"
            return
        if isinstance(val, str) and val.strip().endswith("%"):
            try:
                cell.value = float(val.strip()[:-1]) / 100.0
                cell.number_format = "0.00%"
            except Exception:
                cell.value = val
        elif isinstance(val, str) and val.lstrip("-").isdigit():
            cell.value = int(val)
            if h_name in _NUMERIC_HEADERS: cell.number_format = "0"
        else:
            cell.value = val
            if h_name in _NUMERIC_HEADERS and isinstance(val, (int, float)):
                cell.number_format = "0"

    # Header row
    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.alignment = _align("center")
    _style_header(ws, 1, len(headers))

    # Data rows
    for r_idx, row in enumerate(rows, 2):
        for c_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            raw  = row.get(h, "")
            _assign(cell, raw, h_name=h, row_idx=r_idx)
            cell.font   = Font(size=10, name="Calibri")
            cell.border = _thin_border()
            if c_idx == 1:
                cell.alignment = _align("left")
            elif h in formulas or h in _NUMERIC_HEADERS or _is_numeric_like(raw):
                cell.alignment = _align("right")
            elif isinstance(alignments, dict) and h in alignments:
                cell.alignment = _align(alignments[h])
            else:
                cell.alignment = _align("center")

    # Total row
    if total_row:
        t = 2 + len(rows)
        for c_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=t, column=c_idx)
            val  = total_row.get(h, "")
            if h in formulas:
                f = formulas[h]
                for href, let in col_map.items():
                    f = f.replace(f"{{{href}}}", f"{let}{t}")
                cell.value = f"={f}"; cell.number_format = "0.00%"
            elif h in _NUMERIC_HEADERS:
                # Write the actual pre-computed value (not a SUM formula) so
                # openpyxl data_only=True can read Grand Total without needing
                # Excel to recalculate — avoids GT=None/0 in QC checks.
                actual = total_row.get(h)
                if isinstance(actual, float) and actual.is_integer():
                    actual = int(actual)
                cell.value = actual
                cell.number_format = "#,##0"
            else:
                _assign(cell, val)
            if c_idx == 1:
                cell.alignment = _align("left")
            else:
                cell.alignment = _align("right")
        _style_total(ws, t, len(headers))

    _auto_fit(ws)
    try:
        ws.ignored_errors.append(IgnoredError(
            sqref="A1:ZZ5000",
            numberStoredAsText=True, evalError=True,
            emptyCellReference=True, calculatedColumn=True,
        ))
    except Exception:
        pass


# ---------------------------------------------------------------------------
# REACH sheet — dedicated writer
# ---------------------------------------------------------------------------

# Fixed column widths for REACH (characters)
_REACH_COL_WIDTHS = {
    "Impressions":      18,
    "Clicks":           14,
    "Click Rate (CTR)": 22,
    "Reach":            14,
    "Frequency":        12,
}

_REACH_HEADERS = ["Impressions", "Clicks", "Click Rate (CTR)", "Reach", "Frequency"]

def write_reach_sheet(ws, rows: list[dict]) -> None:
    """
    Dedicated writer for the REACH sheet.

    Layout
    ------
    Row 1  : Header  — blue fill (#00B0F0), bold 11pt Calibri, CENTER aligned, height 18 pt
    Row 2  : Data    — all columns RIGHT aligned, 10pt Calibri, height 16 pt
    No Grand Total row (only 1 data row).

    Column widths are fixed (not auto-fit) so the sheet always looks clean
    regardless of the actual numeric values.

    Alignment per column
    --------------------
    Impressions       → RIGHT  (it's a number; c_idx==1 override removed here)
    Clicks            → RIGHT
    Click Rate (CTR)  → RIGHT  (stored as =B2/A2 formula, formatted 0.00%)
    Reach             → RIGHT
    Frequency         → RIGHT  (hardcoded 3)
    """
    col_map = {h: get_column_letter(i) for i, h in enumerate(_REACH_HEADERS, 1)}

    # ── Row 1: Header ────────────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor=HEADER_BG)
    header_font = Font(bold=True, color="000000", size=11, name="Calibri")
    for c_idx, h in enumerate(_REACH_HEADERS, 1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.border    = _thin_border()
        cell.alignment = _align("center")
    ws.row_dimensions[1].height = 18

    # ── Row 2: Data ──────────────────────────────────────────────────────────
    data_font = Font(size=10, name="Calibri")
    row = rows[0] if rows else {}

    for c_idx, h in enumerate(_REACH_HEADERS, 1):
        cell  = ws.cell(row=2, column=c_idx)
        val   = row.get(h, "")
        let   = col_map[h]

        if h == "Click Rate (CTR)":
            # Formula: =B2/A2 (Clicks / Impressions)
            # QC reader falls back to computing from Impressions/Clicks if uncached.
            cell.value         = f"={col_map['Clicks']}2/{col_map['Impressions']}2"
            cell.number_format = "0.00%"
        elif h in ("Impressions", "Clicks", "Reach", "Frequency"):
            cell.value         = int(val) if val != "" else 0
            cell.number_format = "#,##0"
        else:
            cell.value = val

        cell.font      = data_font
        cell.border    = _thin_border()
        cell.alignment = _align("right")   # ALL columns right-aligned (numbers)

    ws.row_dimensions[2].height = 16

    # ── Column widths (fixed) ────────────────────────────────────────────────
    for c_idx, h in enumerate(_REACH_HEADERS, 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = _REACH_COL_WIDTHS[h]

    # ── Suppress Excel warnings ──────────────────────────────────────────────
    try:
        ws.ignored_errors.append(IgnoredError(
            sqref="A1:E2",
            numberStoredAsText=True,
            evalError=True,
            emptyCellReference=True,
        ))
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Shared helpers used by every dedicated sheet writer below
# ---------------------------------------------------------------------------

_HEADER_HEIGHT = 18   # pt — header row height
_DATA_HEIGHT   = 16   # pt — each data row height
_TOTAL_HEIGHT  = 16   # pt — grand-total row height

# Columns whose values are plain integers (formatted #,##0)
_INT_COLS = {
    "Impressions", "Clicks", "Reach", "Frequency",
    "Measurable Impressions", "Viewable Impressions",
    "Sum of Starts (Video)", "Sum of Complete Views (Video)",
}
# Columns that are percentages (formula or pre-calculated string)
_PCT_COLS = {"Click Rate (CTR)", "Viewability", "VCR (Completion Rate)"}


def _sh_header(ws, headers: list[str]) -> None:
    """Blue bold header row, center-aligned, height 18 pt."""
    fill = PatternFill("solid", fgColor=HEADER_BG)
    font = Font(bold=True, color="000000", size=11, name="Calibri")
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill      = fill
        cell.font      = font
        cell.border    = _thin_border()
        cell.alignment = _align("center")
    ws.row_dimensions[1].height = _HEADER_HEIGHT


def _sh_data_cell(cell, val, h: str, c_idx: int,
                   col1_align: str, col_map: dict,
                   pct_formulas: dict, row_idx: int,
                   other_align: str = "center") -> None:
    """
    Write value + number-format + alignment for one data-row cell.

    col1_align  : "left" | "center" | "right" — alignment for column 1 (label).
    other_align : "center" | "right" — alignment for all other columns (default "center").
    pct_formulas: dict of header → formula template, e.g.
                  {"Click Rate (CTR)": "{Clicks}/{Impressions}"}
    """
    cell.font   = Font(size=10, name="Calibri")
    cell.border = _thin_border()

    # ── value + number format ──────────────────────────────────────────────
    if h in pct_formulas:
        f = pct_formulas[h]
        for href, let in col_map.items():
            f = f.replace(f"{{{href}}}", f"{let}{row_idx}")
        cell.value         = f"={f}"
        cell.number_format = "0.00%"

    elif isinstance(val, str) and val.strip().endswith("%"):
        try:
            cell.value         = float(val.strip()[:-1]) / 100.0
            cell.number_format = "0.00%"
        except Exception:
            cell.value = val

    elif h in _INT_COLS:
        cell.value         = int(safe_int(val))
        cell.number_format = "0"

    else:
        cell.value = val

    # ── alignment ─────────────────────────────────────────────────────────
    cell.alignment = _align(col1_align if c_idx == 1 else other_align)


def _sh_total_row(ws, headers: list[str], total: dict,
                   row_num: int, col_map: dict,
                   pct_formulas: dict, last_data: int) -> None:
    """
    Write the Grand Total row with:
      - SUM() formulas for integer columns
      - SUM()/SUM() formulas for percentage columns
      - Label text (col 1) left-aligned
      - All other cells right-aligned
      - Font: 11pt bold (same as header — acts as a sheet footer)
    """
    fill = PatternFill("solid", fgColor=TOTAL_BG)
    font = Font(bold=True, color="000000", size=11, name="Calibri")
    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=c_idx)
        val  = total.get(h, "")

        if h in pct_formulas:
            # e.g. "{Clicks}/{Impressions}" → "=C{row_num}/B{row_num}"
            # References Grand Total cells directly instead of re-summing ranges
            f = pct_formulas[h]
            for href, let in col_map.items():
                f = f.replace(f"{{{href}}}", f"{let}{row_num}")
            cell.value         = f"={f}"
            cell.number_format = "0.00%"

        elif h in _INT_COLS:
            let = col_map.get(h, "")
            cell.value         = f"=SUM({let}2:{let}{last_data})"
            cell.number_format = "#,##0"

        elif isinstance(val, str) and val.strip().endswith("%"):
            try:
                cell.value         = float(val.strip()[:-1]) / 100.0
                cell.number_format = "0.00%"
            except Exception:
                cell.value = val

        else:
            cell.value = val

        cell.fill      = fill
        cell.font      = font
        cell.border    = _thin_border()
        cell.alignment = _align("left" if c_idx == 1 else "right")

    ws.row_dimensions[row_num].height = _TOTAL_HEIGHT


def _sh_col_widths(ws, headers: list[str], widths: dict) -> None:
    """Apply fixed column widths from a header→width dict."""
    for c, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(c)].width = widths.get(h, 14)


def _sh_suppress(ws) -> None:
    """Suppress openpyxl / Excel formula warnings."""
    try:
        ws.ignored_errors.append(IgnoredError(
            sqref="A1:ZZ5000",
            numberStoredAsText=True, evalError=True,
            emptyCellReference=True, calculatedColumn=True,
        ))
    except Exception:
        pass


# ---------------------------------------------------------------------------
# DATE sheet
# ---------------------------------------------------------------------------

_DATE_H_BANNER = [
    "Date", "Impressions", "Clicks", "Click Rate (CTR)",
    "Viewable Impressions", "Measurable Impressions", "Viewability",
]
_DATE_H_VIDEO = [
    "Date", "Impressions", "Clicks", "Click Rate (CTR)",
    "Viewable Impressions", "Measurable Impressions", "Viewability",
    "Sum of Starts (Video)", "Sum of Complete Views (Video)", "VCR (Completion Rate)",
]
_DATE_F_BANNER = {
    "Click Rate (CTR)": "{Clicks}/{Impressions}",
    "Viewability":      "{Viewable Impressions}/{Measurable Impressions}",
}
_DATE_F_VIDEO = {
    **_DATE_F_BANNER,
    "VCR (Completion Rate)": "{Sum of Complete Views (Video)}/{Sum of Starts (Video)}",
}
_DATE_W_BANNER = {
    "Date": 20, "Impressions": 18, "Clicks": 14,
    "Click Rate (CTR)": 20, "Viewable Impressions": 22,
    "Measurable Impressions": 24, "Viewability": 16,
}
_DATE_W_VIDEO = {
    **_DATE_W_BANNER,
    "Sum of Starts (Video)": 24,
    "Sum of Complete Views (Video)": 28,
    "VCR (Completion Rate)": 22,
}


def write_date_sheet(ws, rows: list[dict], total_row: dict, is_banner: bool) -> None:
    """
    Dedicated writer for the DATE sheet.

    Columns (Banner — 7)  : Date | Impressions | Clicks | CTR |
                            Viewable Imp | Measurable Imp | Viewability
    Columns (Video — 10)  : + Sum of Starts | Sum of Complete Views | VCR

    Alignment
    ---------
    Date                          → LEFT
    Impressions / Clicks          → Center  #,##0
    Click Rate (CTR)              → Center  0.00%  formula =C{r}/B{r}
    Viewable / Measurable Imp     → Center  #,##0
    Viewability                   → Center  0.00%  formula
    Sum of Starts / Complete Views→ Center  #,##0  (video only)
    VCR (Completion Rate)         → Center  0.00%  formula (video only)
    Grand Total row col-1         → LEFT
    """
    headers  = _DATE_H_BANNER if is_banner else _DATE_H_VIDEO
    formulas = _DATE_F_BANNER if is_banner else _DATE_F_VIDEO
    widths   = _DATE_W_BANNER if is_banner else _DATE_W_VIDEO
    col_map  = {h: get_column_letter(i) for i, h in enumerate(headers, 1)}

    _sh_header(ws, headers)
    ws.cell(row=1, column=1).alignment = _align("center")   # Date header → CENTER

    for r_idx, row in enumerate(rows, 2):
        for c_idx, h in enumerate(headers, 1):
            _sh_data_cell(
                ws.cell(row=r_idx, column=c_idx),
                row.get(h, ""), h, c_idx,
                "center", col_map, formulas, r_idx,
            )
        ws.row_dimensions[r_idx].height = _DATA_HEIGHT

    if total_row and rows:
        t = 2 + len(rows)
        _sh_total_row(ws, headers, total_row, t, col_map, formulas, t - 1)

    _sh_col_widths(ws, headers, widths)
    _sh_suppress(ws)


# ---------------------------------------------------------------------------
# APP URL sheet
# ---------------------------------------------------------------------------

_APP_URL_H = ["App/URL", "Impressions", "Clicks", "Click Rate (CTR)"]
_APP_URL_F = {"Click Rate (CTR)": "{Clicks}/{Impressions}"}
_APP_URL_W = {"App/URL": 42, "Impressions": 18, "Clicks": 14, "Click Rate (CTR)": 20}


def write_app_url_sheet(ws, rows: list[dict], total_row: dict) -> None:
    """
    Dedicated writer for the APP URL sheet.

    Alignment
    ---------
    App/URL      → LEFT   (domain text — can be long)
    Impressions  → RIGHT  #,##0
    Clicks       → RIGHT  #,##0
    CTR          → RIGHT  0.00%  formula
    Grand Total  → col-1 LEFT, rest RIGHT
    """
    col_map = {h: get_column_letter(i) for i, h in enumerate(_APP_URL_H, 1)}
    _sh_header(ws, _APP_URL_H)
    ws.cell(row=1, column=1).alignment = _align("left")   # App/URL header → LEFT
    for r_idx, row in enumerate(rows, 2):
        for c_idx, h in enumerate(_APP_URL_H, 1):
            _sh_data_cell(ws.cell(row=r_idx, column=c_idx),
                          row.get(h, ""), h, c_idx,
                          "left", col_map, _APP_URL_F, r_idx,
                          other_align="right")
        ws.row_dimensions[r_idx].height = _DATA_HEIGHT
    if total_row and rows:
        t = 2 + len(rows)
        _sh_total_row(ws, _APP_URL_H, total_row, t, col_map, _APP_URL_F, t - 1)
    _sh_col_widths(ws, _APP_URL_H, _APP_URL_W)
    _sh_suppress(ws)


# ---------------------------------------------------------------------------
# TIME OF DAY sheet
# ---------------------------------------------------------------------------

_TOD_H = ["Time of Day", "Impressions", "Clicks", "Click Rate (CTR)"]
_TOD_F = {"Click Rate (CTR)": "{Clicks}/{Impressions}"}
_TOD_W = {"Time of Day": 16, "Impressions": 18, "Clicks": 14, "Click Rate (CTR)": 20}


def write_time_of_day_sheet(ws, rows: list[dict], total_row: dict) -> None:
    """
    Dedicated writer for the TIME OF DAY sheet.

    Alignment
    ---------
    Time of Day  → CENTER (hour labels 0–23, centered looks cleaner than left)
    Impressions  → RIGHT  #,##0
    Clicks       → RIGHT  #,##0
    CTR          → RIGHT  0.00%  formula
    Grand Total  → col-1 LEFT, rest RIGHT
    """
    col_map = {h: get_column_letter(i) for i, h in enumerate(_TOD_H, 1)}
    _sh_header(ws, _TOD_H)
    ws.cell(row=1, column=1).alignment = _align("left")   # Time of Day header → LEFT
    for r_idx, row in enumerate(rows, 2):
        for c_idx, h in enumerate(_TOD_H, 1):
            _sh_data_cell(ws.cell(row=r_idx, column=c_idx),
                          row.get(h, ""), h, c_idx,
                          "left", col_map, _TOD_F, r_idx,
                          other_align="right")
        ws.row_dimensions[r_idx].height = _DATA_HEIGHT
    if total_row and rows:
        t = 2 + len(rows)
        _sh_total_row(ws, _TOD_H, total_row, t, col_map, _TOD_F, t - 1)
    _sh_col_widths(ws, _TOD_H, _TOD_W)
    _sh_suppress(ws)


# ---------------------------------------------------------------------------
# EXCHANGE sheet
# ---------------------------------------------------------------------------

_EXCHANGE_H = ["Exchange", "Impressions", "Clicks", "Click Rate (CTR)"]
_EXCHANGE_F = {"Click Rate (CTR)": "{Clicks}/{Impressions}"}
_EXCHANGE_W = {"Exchange": 32, "Impressions": 18, "Clicks": 14, "Click Rate (CTR)": 20}


def write_exchange_sheet(ws, rows: list[dict], total_row: dict) -> None:
    """
    Dedicated writer for the EXCHANGE sheet.

    Alignment
    ---------
    Exchange     → LEFT   (company names e.g. "Google Ad Manager")
    Impressions  → RIGHT  #,##0
    Clicks       → RIGHT  #,##0
    CTR          → RIGHT  0.00%  formula
    Grand Total  → col-1 LEFT, rest RIGHT
    """
    col_map = {h: get_column_letter(i) for i, h in enumerate(_EXCHANGE_H, 1)}
    _sh_header(ws, _EXCHANGE_H)
    ws.cell(row=1, column=1).alignment = _align("left")   # Exchange header → LEFT
    for r_idx, row in enumerate(rows, 2):
        for c_idx, h in enumerate(_EXCHANGE_H, 1):
            _sh_data_cell(ws.cell(row=r_idx, column=c_idx),
                          row.get(h, ""), h, c_idx,
                          "left", col_map, _EXCHANGE_F, r_idx,
                          other_align="right")
        ws.row_dimensions[r_idx].height = _DATA_HEIGHT
    if total_row and rows:
        t = 2 + len(rows)
        _sh_total_row(ws, _EXCHANGE_H, total_row, t, col_map, _EXCHANGE_F, t - 1)
    _sh_col_widths(ws, _EXCHANGE_H, _EXCHANGE_W)
    _sh_suppress(ws)


# ---------------------------------------------------------------------------
# DEVICE sheet
# ---------------------------------------------------------------------------

_DEVICE_H = ["Device Type", "Impressions", "Clicks", "Click Rate (CTR)"]
_DEVICE_F = {"Click Rate (CTR)": "{Clicks}/{Impressions}"}
_DEVICE_W = {"Device Type": 20, "Impressions": 18, "Clicks": 14, "Click Rate (CTR)": 20}


def write_device_sheet(ws, rows: list[dict], total_row: dict) -> None:
    """
    Dedicated writer for the DEVICE sheet.

    Alignment
    ---------
    Device Type  → LEFT   (Desktop / Smart Phone / Tablet)
    Impressions  → RIGHT  #,##0
    Clicks       → RIGHT  #,##0
    CTR          → RIGHT  0.00%  formula
    Grand Total  → col-1 LEFT, rest RIGHT
    """
    col_map = {h: get_column_letter(i) for i, h in enumerate(_DEVICE_H, 1)}
    _sh_header(ws, _DEVICE_H)
    ws.cell(row=1, column=1).alignment = _align("left")   # Device Type header → LEFT
    for r_idx, row in enumerate(rows, 2):
        for c_idx, h in enumerate(_DEVICE_H, 1):
            _sh_data_cell(ws.cell(row=r_idx, column=c_idx),
                          row.get(h, ""), h, c_idx,
                          "left", col_map, _DEVICE_F, r_idx,
                          other_align="right")
        ws.row_dimensions[r_idx].height = _DATA_HEIGHT
    if total_row and rows:
        t = 2 + len(rows)
        _sh_total_row(ws, _DEVICE_H, total_row, t, col_map, _DEVICE_F, t - 1)
    _sh_col_widths(ws, _DEVICE_H, _DEVICE_W)
    _sh_suppress(ws)


# ---------------------------------------------------------------------------
# CREATIVE sheet
# ---------------------------------------------------------------------------

_CREATIVE_H = ["Creative", "Impressions", "Clicks", "Click Rate (CTR)"]
_CREATIVE_F = {"Click Rate (CTR)": "{Clicks}/{Impressions}"}
_CREATIVE_W = {"Creative": 38, "Impressions": 18, "Clicks": 14, "Click Rate (CTR)": 20}


def write_creative_sheet(ws, rows: list[dict], total_row: Optional[dict]) -> None:
    """
    Dedicated writer for the CREATIVE sheet.

    Alignment
    ---------
    Creative     → LEFT   (creative name / ad name — can be long)
    Impressions  → RIGHT  #,##0
    Clicks       → RIGHT  #,##0
    CTR          → RIGHT  0.00%  formula
    Grand Total  → col-1 LEFT, rest RIGHT
    Note: rows may be empty when no Creative column found in source file.
    """
    col_map = {h: get_column_letter(i) for i, h in enumerate(_CREATIVE_H, 1)}
    _sh_header(ws, _CREATIVE_H)
    ws.cell(row=1, column=1).alignment = _align("left")   # Creative header → LEFT
    for r_idx, row in enumerate(rows, 2):
        for c_idx, h in enumerate(_CREATIVE_H, 1):
            _sh_data_cell(ws.cell(row=r_idx, column=c_idx),
                          row.get(h, ""), h, c_idx,
                          "left", col_map, _CREATIVE_F, r_idx,
                          other_align="right")
        ws.row_dimensions[r_idx].height = _DATA_HEIGHT
    if total_row and rows:
        t = 2 + len(rows)
        _sh_total_row(ws, _CREATIVE_H, total_row, t, col_map, _CREATIVE_F, t - 1)
    _sh_col_widths(ws, _CREATIVE_H, _CREATIVE_W)
    _sh_suppress(ws)


# ---------------------------------------------------------------------------
# CITY sheet
# ---------------------------------------------------------------------------

_CITY_H = ["City", "Impressions", "Clicks", "Click Rate (CTR)"]
_CITY_F = {"Click Rate (CTR)": "{Clicks}/{Impressions}"}
_CITY_W = {"City": 24, "Impressions": 18, "Clicks": 14, "Click Rate (CTR)": 20}


def write_city_sheet(ws, rows: list[dict], total_row: dict) -> None:
    """
    Dedicated writer for the CITY sheet.

    Alignment
    ---------
    City         → LEFT   (city names)
    Impressions  → RIGHT  #,##0
    Clicks       → RIGHT  #,##0
    CTR          → RIGHT  0.00%  formula
    Grand Total  → col-1 LEFT, rest RIGHT
    """
    col_map = {h: get_column_letter(i) for i, h in enumerate(_CITY_H, 1)}
    _sh_header(ws, _CITY_H)
    ws.cell(row=1, column=1).alignment = _align("left")   # City header → LEFT
    for r_idx, row in enumerate(rows, 2):
        for c_idx, h in enumerate(_CITY_H, 1):
            _sh_data_cell(ws.cell(row=r_idx, column=c_idx),
                          row.get(h, ""), h, c_idx,
                          "left", col_map, _CITY_F, r_idx,
                          other_align="right")
        ws.row_dimensions[r_idx].height = _DATA_HEIGHT
    if total_row and rows:
        t = 2 + len(rows)
        _sh_total_row(ws, _CITY_H, total_row, t, col_map, _CITY_F, t - 1)
    _sh_col_widths(ws, _CITY_H, _CITY_W)
    _sh_suppress(ws)


# ---------------------------------------------------------------------------
# AGE sheet
# ---------------------------------------------------------------------------

_AGE_H = ["Age", "Impressions", "Clicks", "Click Rate (CTR)"]
_AGE_F = {"Click Rate (CTR)": "{Clicks}/{Impressions}"}
_AGE_W = {"Age": 14, "Impressions": 18, "Clicks": 14, "Click Rate (CTR)": 20}


def write_age_sheet(ws, rows: list[dict], total_row: dict) -> None:
    """
    Dedicated writer for the AGE sheet.

    Alignment
    ---------
    Age          → CENTER  (ranges like "18-24", "25-34" look better centered)
    Impressions  → RIGHT   #,##0
    Clicks       → RIGHT   #,##0
    CTR          → RIGHT   0.00%  formula
    Grand Total  → col-1 LEFT, rest RIGHT
    """
    col_map = {h: get_column_letter(i) for i, h in enumerate(_AGE_H, 1)}
    _sh_header(ws, _AGE_H)
    ws.cell(row=1, column=1).alignment = _align("left")   # Age header → LEFT
    for r_idx, row in enumerate(rows, 2):
        for c_idx, h in enumerate(_AGE_H, 1):
            _sh_data_cell(ws.cell(row=r_idx, column=c_idx),
                          row.get(h, ""), h, c_idx,
                          "left", col_map, _AGE_F, r_idx,
                          other_align="right")
        ws.row_dimensions[r_idx].height = _DATA_HEIGHT
    if total_row and rows:
        t = 2 + len(rows)
        _sh_total_row(ws, _AGE_H, total_row, t, col_map, _AGE_F, t - 1)
    _sh_col_widths(ws, _AGE_H, _AGE_W)
    _sh_suppress(ws)


# ---------------------------------------------------------------------------
# GENDER sheet
# ---------------------------------------------------------------------------

_GENDER_H = ["Gender", "Impressions", "Clicks", "Click Rate (CTR)"]
_GENDER_F = {"Click Rate (CTR)": "{Clicks}/{Impressions}"}
_GENDER_W = {"Gender": 14, "Impressions": 18, "Clicks": 14, "Click Rate (CTR)": 20}


def write_gender_sheet(ws, rows: list[dict], total_row: dict) -> None:
    """
    Dedicated writer for the GENDER sheet.

    Alignment
    ---------
    Gender       → CENTER  (Male / Female — short labels, centered looks cleaner)
    Impressions  → RIGHT   #,##0
    Clicks       → RIGHT   #,##0
    CTR          → RIGHT   0.00%  formula
    Grand Total  → col-1 LEFT, rest RIGHT
    """
    col_map = {h: get_column_letter(i) for i, h in enumerate(_GENDER_H, 1)}
    _sh_header(ws, _GENDER_H)
    ws.cell(row=1, column=1).alignment = _align("left")   # Gender header → LEFT
    for r_idx, row in enumerate(rows, 2):
        for c_idx, h in enumerate(_GENDER_H, 1):
            _sh_data_cell(ws.cell(row=r_idx, column=c_idx),
                          row.get(h, ""), h, c_idx,
                          "left", col_map, _GENDER_F, r_idx,
                          other_align="right")
        ws.row_dimensions[r_idx].height = _DATA_HEIGHT
    if total_row and rows:
        t = 2 + len(rows)
        _sh_total_row(ws, _GENDER_H, total_row, t, col_map, _GENDER_F, t - 1)
    _sh_col_widths(ws, _GENDER_H, _GENDER_W)
    _sh_suppress(ws)


# ---------------------------------------------------------------------------
# Sheet builders (ported directly from App.py)
# ---------------------------------------------------------------------------

def build_sheet1_reach(total_imp: int, total_clk: int) -> list[dict]:
    ctr   = pct(total_clk, total_imp)
    freq  = 3
    reach = int(total_imp / freq) + random.randint(200, 300)
    return [{"Impressions": total_imp, "Clicks": total_clk,
             "Click Rate (CTR)": ctr, "Reach": reach, "Frequency": freq}]


def build_sheet2_date(df: pd.DataFrame, total_imp: int, total_clk: int,
                       ctr_reach: str, is_banner: bool = False):
    rows = []
    sum_imp = sum_clk = sum_view = sum_meas = sum_starts = sum_comp = 0
    vcr_weighted = vcr_imp_total = 0

    # ── Use actual impressions & clicks from input sheet if available ─────────
    # If the input sheet already has Impressions/Clicks values, use them directly.
    # Only randomise when the source data is missing or all-zero.
    n = len(df)
    _src_imps = [safe_int(r.get("Impressions", 0)) for _, r in df.iterrows()] if n > 0 else []
    _src_clks = [safe_int(r.get("Clicks", 0))      for _, r in df.iterrows()] if n > 0 else []
    _has_imp  = sum(_src_imps) > 0
    _has_clk  = sum(_src_clks) > 0

    if n > 0 and _has_imp:
        # Input has real Impressions — use them directly (rescale to match total_imp if needed)
        _imp_sum = sum(_src_imps)
        if abs(_imp_sum - total_imp) <= 1:
            rand_imps = _src_imps           # already matches exactly
        else:
            rand_imps = _largest_remainder(_src_imps, _imp_sum, total_imp)
    elif n > 0:
        # No Impressions in source → randomise
        imp_weights = [random.uniform(0.5, 1.5) for _ in range(n)]
        rand_imps   = _largest_remainder(imp_weights, sum(imp_weights), total_imp)
    else:
        rand_imps = []

    if n > 0 and _has_clk:
        # Input has real Clicks — use them directly (rescale to match total_clk if needed)
        _clk_sum = sum(_src_clks)
        if abs(_clk_sum - total_clk) <= 1:
            rand_clks = _src_clks
        else:
            rand_clks = _largest_remainder(_src_clks, _clk_sum, total_clk)
    elif n > 0:
        # No Clicks in source → randomise
        clk_weights = [random.uniform(0.5, 1.5) for _ in range(n)]
        rand_clks   = _largest_remainder(clk_weights, sum(clk_weights), total_clk)
    else:
        rand_clks = []

    for row_i, (_, r) in enumerate(df.iterrows()):
        imp = rand_imps[row_i] if rand_imps else safe_int(r.get("Impressions", 0))
        clk = rand_clks[row_i] if rand_clks else safe_int(r.get("Clicks", 0))
        date_val = r.get("Date", "")
        # pandas Timestamp / Python datetime → format directly
        if hasattr(date_val, "strftime"):
            date_str = date_val.strftime("%d %B, %Y")
        else:
            try:
                float(date_val)          # Excel serial number
                date_str = serial_to_date(date_val)
            except (ValueError, TypeError):
                date_str = str(date_val)

        sum_imp += imp; sum_clk += clk

        if is_banner:
            meas = round(imp * random.uniform(0.90, 0.96))
            view = round(meas * random.uniform(0.75, 0.85))
            sum_view += view; sum_meas += meas
            rows.append({"Date": date_str, "Impressions": imp, "Clicks": clk,
                         "Click Rate (CTR)": pct(clk, imp),
                         "Viewable Impressions": view,
                         "Measurable Impressions": meas,
                         "Viewability": pct(view, meas)})
        else:
            view   = safe_int(r.get("Viewable Impressions", 0))
            meas   = safe_int(r.get("Measurable Impressions", 0))
            starts = safe_int(r.get("Sum of Starts (Video)") or r.get("Start views") or 0)
            comps  = safe_int(r.get("Sum of Complete Views (Video)") or r.get("Complete Views") or 0)
            # VCR may be a decimal (0.75) or a percent string ("75%") — normalise to decimal
            vcr_val = r.get("VCR (Completion Rate)") or r.get("Video Completion Rate (VCR)") or 0
            if isinstance(vcr_val, str) and str(vcr_val).strip().endswith("%"):
                vcr_raw = safe_float(str(vcr_val).replace("%", "")) / 100
            else:
                vcr_raw = safe_float(vcr_val)

            # Source file may not have video-specific metrics (only Date/Imp/Clicks).
            # Generate synthetic values from impression count — same approach as banner branch.
            if imp > 0:
                if meas == 0:
                    meas = round(imp * random.uniform(0.90, 0.98))
                if view == 0:
                    view = round(meas * random.uniform(0.80, 0.92))
                if starts == 0:
                    starts = round(imp * random.uniform(0.80, 0.95))
                if comps == 0:
                    comps = round(starts * random.uniform(0.75, 0.85))

            sum_view += view; sum_meas += meas
            sum_starts += starts; sum_comp += comps
            vcr_weighted += vcr_raw * imp; vcr_imp_total += imp
            vcr_pct = pct(comps, starts)
            rows.append({"Date": date_str, "Impressions": imp, "Clicks": clk,
                         "Click Rate (CTR)": pct(clk, imp),
                         "Viewable Impressions": view,
                         "Measurable Impressions": meas,
                         "Viewability": pct(view, meas),
                         "Sum of Starts (Video)": starts,
                         "Sum of Complete Views (Video)": comps,
                         "VCR (Completion Rate)": vcr_pct})

    if is_banner:
        total = {"Date": "Grand Total", "Impressions": sum_imp, "Clicks": sum_clk,
                 "Click Rate (CTR)": ctr_reach,
                 "Viewable Impressions": sum_view,
                 "Measurable Impressions": sum_meas,
                 "Viewability": pct(sum_view, sum_meas)}
    else:
        total = {"Date": "Grand Total", "Impressions": sum_imp, "Clicks": sum_clk,
                 "Click Rate (CTR)": ctr_reach,
                 "Viewable Impressions": sum_view,
                 "Measurable Impressions": sum_meas,
                 "Viewability": pct(sum_view, sum_meas),
                 "Sum of Starts (Video)": sum_starts,
                 "Sum of Complete Views (Video)": sum_comp,
                 "VCR (Completion Rate)": pct(sum_comp, sum_starts)}
    return rows, total


def build_sheet3_timeofday(total_imp: int, total_clk: int, ctr_reach: str):
    hour_weights = {
        0:0.5,1:0.4,2:0.3,3:0.3,4:0.4,5:0.7,6:1.2,7:1.5,
        8:2.0,9:2.5,10:2.3,11:2.4,12:2.2,13:1.8,14:1.6,15:1.7,
        16:1.9,17:2.6,18:2.8,19:2.5,20:2.0,21:1.6,22:1.2,23:0.8
    }
    noisy = [hour_weights[h] * (0.85 + random.random() * 0.30) for h in range(24)]
    total_w = sum(noisy)
    hourly_imp = [int((w / total_w) * total_imp) for w in noisy]
    diff = total_imp - sum(hourly_imp)
    for _ in range(abs(diff)):
        hourly_imp[random.randint(0, 23)] += 1 if diff > 0 else -1
    hourly_imp = deduplicate_preserving_sum(hourly_imp, gap=1)

    hourly_clk = []
    for imp in hourly_imp:
        if imp >= 180:
            c_min = math.ceil(imp * 0.0035)
            c_max = math.floor(imp * 0.0056)
            hourly_clk.append(random.randint(c_min, max(c_min, c_max)))
        else:
            hourly_clk.append(0)

    current  = sum(hourly_clk)
    diff_clk = total_clk - current
    eligible = [i for i, imp in enumerate(hourly_imp) if imp >= 180]
    iterations = 0
    while diff_clk != 0 and eligible and iterations < 5000:
        iterations += 1
        idx = random.choice(eligible)
        imp = hourly_imp[idx]; clk = hourly_clk[idx]
        if diff_clk > 0:
            if (clk + 1) / imp <= 0.0056:
                hourly_clk[idx] += 1; diff_clk -= 1
        elif diff_clk < 0:
            if clk > 1 and (clk - 1) / imp >= 0.0035:
                hourly_clk[idx] -= 1; diff_clk += 1

    rows = [{"Time of Day": h, "Impressions": hourly_imp[h],
              "Clicks": hourly_clk[h], "Click Rate (CTR)": pct(hourly_clk[h], hourly_imp[h])}
             for h in range(24)]
    total = {"Time of Day": "Grand Total", "Impressions": total_imp,
             "Clicks": total_clk, "Click Rate (CTR)": ctr_reach}
    return rows, total


def build_sheet4_age(total_imp: int, total_clk: int, ctr_reach: str):
    pct_1824 = rand_float(0.20, 0.25)
    pct_4554 = rand_float(0.09, 0.12)
    remaining = max(0.60, min(0.74, 1 - pct_1824 - pct_4554))
    split = rand_float(0.45, 0.55)
    pct_2534 = remaining * split
    pct_3544 = remaining - pct_2534

    imp_1824 = round(total_imp * pct_1824)
    imp_4554 = round(total_imp * pct_4554)
    rem_imp  = total_imp - imp_1824 - imp_4554
    imp_2534 = round(rem_imp * split)
    imp_3544 = rem_imp - imp_2534
    imp_2534 += total_imp - (imp_1824 + imp_2534 + imp_3544 + imp_4554)

    groups = [{"age": "18-24", "imp": imp_1824}, {"age": "25-34", "imp": imp_2534},
              {"age": "35-44", "imp": imp_3544}, {"age": "45-54", "imp": imp_4554}]
    clicks = [round(g["imp"] * rand_float(0.0035, 0.0056)) if g["imp"] >= 180 else 0 for g in groups]
    c_sum  = sum(clicks)
    if c_sum > 0:
        clicks = [round(c * total_clk / c_sum) if groups[i]["imp"] >= 180 else 0 for i, c in enumerate(clicks)]
    drift = total_clk - sum(clicks)
    clicks[clicks.index(max(clicks))] += drift

    rows = [{"Age": g["age"], "Impressions": g["imp"], "Clicks": clicks[i],
              "Click Rate (CTR)": pct(clicks[i], g["imp"])} for i, g in enumerate(groups)]
    total = {"Age": "Grand Total", "Impressions": total_imp, "Clicks": total_clk, "Click Rate (CTR)": ctr_reach}
    return rows, total


def build_sheet5_gender(total_imp: int, total_clk: int, ctr_reach: str):
    female_pct = rand_float(0.30, 0.40)
    imp_female = round(total_imp * female_pct)
    imp_male   = total_imp - imp_female
    male_pct   = imp_male / total_imp
    if male_pct < 0.58:
        imp_male   = round(total_imp * rand_float(0.58, 0.61))
        imp_female = total_imp - imp_male
    elif male_pct > 0.65:
        imp_male   = round(total_imp * rand_float(0.62, 0.65))
        imp_female = total_imp - imp_male
    imp_male += total_imp - (imp_male + imp_female)

    groups = [{"gender": "Male", "imp": imp_male}, {"gender": "Female", "imp": imp_female}]
    clicks = [round(g["imp"] * rand_float(0.0035, 0.0056)) if g["imp"] >= 180 else 0 for g in groups]
    g_sum  = sum(clicks)
    if g_sum > 0:
        clicks = [round(c * total_clk / g_sum) if groups[i]["imp"] >= 180 else 0 for i, c in enumerate(clicks)]
    drift = total_clk - sum(clicks)
    clicks[clicks.index(max(clicks))] += drift

    rows = [{"Gender": g["gender"], "Impressions": g["imp"], "Clicks": clicks[i],
              "Click Rate (CTR)": pct(clicks[i], g["imp"])} for i, g in enumerate(groups)]
    total = {"Gender": "Grand Total", "Impressions": total_imp, "Clicks": total_clk, "Click Rate (CTR)": ctr_reach}
    return rows, total


def build_sheet6_device(total_imp: int, total_clk: int, ctr_reach: str):
    tablet_pct  = rand_float(0.07, 0.10)
    desktop_shr = rand_float(0.45, 0.55)
    rem         = 1 - tablet_pct
    imp_tablet  = round(total_imp * tablet_pct)
    imp_desktop = round(total_imp * rem * desktop_shr)
    imp_mobile  = total_imp - imp_tablet - imp_desktop
    imp_mobile += total_imp - (imp_desktop + imp_mobile + imp_tablet)

    groups = [{"device": "Desktop",     "imp": imp_desktop},
              {"device": "Smart Phone", "imp": imp_mobile},
              {"device": "Tablet",      "imp": imp_tablet}]
    clicks = [max(0, int(g["imp"] * rand_float(0.0035, 0.0056))) if g["imp"] >= 180 else 0 for g in groups]
    d_sum  = sum(clicks)
    if d_sum > 0:
        clicks = [int(c * total_clk / d_sum) if groups[i]["imp"] >= 180 else 0 for i, c in enumerate(clicks)]
    drift = total_clk - sum(clicks)
    clicks[clicks.index(max(clicks))] += drift

    rows = [{"Device Type": g["device"], "Impressions": g["imp"], "Clicks": clicks[i],
              "Click Rate (CTR)": pct(clicks[i], g["imp"])} for i, g in enumerate(groups)]
    total = {"Device Type": "Grand Total", "Impressions": total_imp, "Clicks": total_clk, "Click Rate (CTR)": ctr_reach}
    return rows, total




def build_sheet7_exchange(total_imp: int, total_clk: int, ctr_reach: str):
    tier1_other = ["BidSwitch","Index Exchange","Magnite DV+","PubMatic"]
    tier2 = ["Criteo Commerce Grid","Equativ","InMobi","Media.net","Nexxen (fka Unruly)","OpenX","Sovrn"]
    tier3 = ["Microsoft Monetize","Epsilon Core Private Exchange","Yieldmo","TripleLift",
             "Kargo","Improve Digital","TeadsTv","GumGum","Adform"]

    ex_t1 = int(total_imp * rand_float(0.80, 0.85))
    ex_t2 = int(total_imp * rand_float(0.08, 0.13))
    ex_t3 = total_imp - ex_t1 - ex_t2
    google = int(ex_t1 * rand_float(0.35, 0.45))
    t1_other = ex_t1 - google

    def split_across(names, total):
        remaining = total; result = []
        for i, name in enumerate(names):
            if i == len(names) - 1:
                imp = remaining
            else:
                imp = max(1, rand_int(1, int((remaining - (len(names)-1-i)) * 0.6)))
            remaining -= imp
            result.append({"name": name, "imp": imp})
        return result

    exchanges = ([{"name": "Google Ad Manager", "imp": google}]
                 + split_across(tier1_other, t1_other)
                 + split_across(tier2, ex_t2)
                 + split_across(tier3, ex_t3))
    drift = total_imp - sum(e["imp"] for e in exchanges)
    exchanges[0]["imp"] += drift

    clicks = [round(e["imp"] * rand_float(0.0035, 0.0056)) if e["imp"] >= 180 else 0 for e in exchanges]
    e_sum  = sum(clicks)
    if e_sum > 0:
        clicks = [round(c * total_clk / e_sum) if exchanges[i]["imp"] >= 180 else 0 for i, c in enumerate(clicks)]
    e_drift = total_clk - sum(clicks)
    max_idx = max(range(len(exchanges)), key=lambda i: exchanges[i]["imp"])
    clicks[max_idx] += e_drift

    rows = sorted([{"Exchange": e["name"], "Impressions": e["imp"], "Clicks": clicks[i],
                    "Click Rate (CTR)": pct(clicks[i], e["imp"])} for i, e in enumerate(exchanges)],
                  key=lambda x: x["Exchange"])
    total = {"Exchange": "Grand Total", "Impressions": total_imp, "Clicks": total_clk, "Click Rate (CTR)": ctr_reach}
    return rows, total


# ---------------------------------------------------------------------------
# Creative / City / App helpers
# ---------------------------------------------------------------------------

def _find_li_col(frame: pd.DataFrame) -> Optional[str]:
    """Return the 'Line Item' column name from a DataFrame, or None."""
    for col in frame.columns:
        if "line item" in str(col).strip().lower():
            return col
    return None


def _li_tokens(val) -> set:
    """
    Extract comparable tokens from a Line Item string.
    Strips leading 'LI12345|' prefix and normalises to lowercase words.
    Safe against None / float NaN values.
    """
    if val is None:
        return set()
    try:
        if isinstance(val, float) and math.isnan(val):
            return set()
    except Exception:
        pass
    val = str(val).strip()
    if not val or val.lower() in {"nan", "none", ""}:
        return set()
    val = re.sub(r'^[A-Za-z]{2}\d+\|', '', val).strip().lower()
    return set(re.split(r'[\s\-|,_]+', val)) - {'', '-'}


def build_sheet9_creative(df: pd.DataFrame, total_imp: int, total_clk: int, ctr_reach: str,
                           df2: Optional[pd.DataFrame] = None, li_hint: str = ""):
    """Creative sheet — group by Creative column."""
    _CREATIVE_KEYS = {"creative", "ad name", "ad content", "ad"}

    def _find_creative_col(frame: pd.DataFrame) -> Optional[str]:
        for col in frame.columns:
            if str(col).strip().lower() in _CREATIVE_KEYS:
                return col
        return None

    source = df
    creative_col = _find_creative_col(df)
    if creative_col is None and df2 is not None:
        creative_col = _find_creative_col(df2)
        if creative_col is not None:
            source = df2

    if creative_col is None:
        return [], {"Creative": "Grand Total", "Impressions": total_imp,
                    "Clicks": total_clk, "Click Rate (CTR)": ctr_reach}

    if source is not df:
        li_col_df  = _find_li_col(df)
        li_col_src = _find_li_col(source)
        _filtered  = False

        if li_col_df and li_col_src:
            df_li_raw  = (df[li_col_df].dropna()
                           .apply(lambda v: str(v).strip())
                           .replace({"nan": None, "None": None, "": None})
                           .dropna().unique())
            df_li_norm = {v.lower() for v in df_li_raw if v}
            df_li_tok  = [t for v in df_li_raw if v for t in [_li_tokens(v)] if t]

            src_li = source[li_col_src].fillna("").astype(str).str.strip()
            exact_mask   = src_li.str.lower().isin(df_li_norm)
            src_li_clean = src_li.str.replace(r'^[A-Za-z]{2}\d+\|', '', regex=True).str.strip()
            prefix_mask  = src_li_clean.str.lower().isin(df_li_norm)

            def _token_match(val) -> bool:
                toks = _li_tokens(val)
                if not toks: return False
                for ref in df_li_tok:
                    if ref and len(toks & ref) / max(len(ref), 1) >= 0.6:
                        return True
                return False
            token_mask = src_li.apply(_token_match)

            if exact_mask.any():
                source = source[exact_mask].copy(); _filtered = True
            elif prefix_mask.any():
                source = source[prefix_mask].copy(); _filtered = True
            elif token_mask.any():
                source = source[token_mask].copy(); _filtered = True

        if not _filtered and li_hint and li_col_src:
            hint_toks = _li_tokens(li_hint)
            if hint_toks:
                counts  = source[li_col_src].apply(lambda v: len(hint_toks & _li_tokens(v)))
                max_cnt = counts.max()
                if max_cnt > 0:
                    source = source[counts == max_cnt].copy(); _filtered = True

        if not _filtered and li_hint and creative_col:
            hint_toks = _li_tokens(li_hint)
            if hint_toks:
                counts  = source[creative_col].apply(lambda v: len(hint_toks & _li_tokens(str(v))))
                max_cnt = counts.max()
                if max_cnt > 0:
                    source = source[counts == max_cnt].copy()

    df_copy = source.copy()
    df_copy["_imp"] = df_copy.get("Impressions", pd.Series([0] * len(df_copy))).apply(safe_float)
    df_copy["_clk"] = df_copy.get("Clicks",      pd.Series([0] * len(df_copy))).apply(safe_float)
    df_copy["_creative"] = df_copy[creative_col].astype(str).str.strip()

    grouped = (df_copy.groupby("_creative", sort=True)
               .agg(raw_imp=("_imp", "sum"), raw_clk=("_clk", "sum"))
               .reset_index())
    grouped = grouped[grouped["_creative"].str.lower() != "nan"]
    grouped = grouped.sort_values("_creative")

    if grouped.empty:
        return [], {"Creative": "Grand Total", "Impressions": total_imp,
                    "Clicks": total_clk, "Click Rate (CTR)": ctr_reach}

    # Randomise distribution across creatives each run
    n_cr        = len(grouped)
    weights     = [random.uniform(0.5, 1.5) for _ in range(n_cr)]
    total_w     = sum(weights) or 1.0
    cr_imps     = _largest_remainder(weights, total_w, total_imp)
    clk_weights = [random.uniform(0.5, 1.5) for _ in range(n_cr)]
    clk_w       = sum(clk_weights) or 1.0
    cr_clks     = _largest_remainder(clk_weights, clk_w, total_clk)

    imp_drift = total_imp - sum(cr_imps)
    if imp_drift: cr_imps[cr_imps.index(max(cr_imps))] += imp_drift
    clk_drift = total_clk - sum(cr_clks)
    if clk_drift: cr_clks[cr_clks.index(max(cr_clks))] += clk_drift

    rows  = [{"Creative": grouped.iloc[i]["_creative"], "Impressions": cr_imps[i],
              "Clicks": cr_clks[i], "Click Rate (CTR)": pct(cr_clks[i], cr_imps[i])}
             for i in range(len(grouped))]
    total = {"Creative": "Grand Total", "Impressions": total_imp,
             "Clicks": total_clk, "Click Rate (CTR)": ctr_reach}
    return rows, total


# ---------------------------------------------------------------------------
# City sheet  (pivot-aware)
# ---------------------------------------------------------------------------

def build_sheet8_city(total_imp: int, total_clk: int, ctr_reach: str,
                       city_sheet_names: list,
                       df2=None, li_hint: str = ""):
    """
    City sheet — city list built from Pivot Report (df2) when available.

    Pivot path  : match li_hint tokens against Pivot Line Item column,
                  collect unique cities, map to DB weights (fallback weight=1.0).
                  No city cap when pivot data is available.
    Fallback    : use DB cities capped at 60.
    """
    pivot_cities = []
    if df2 is not None and not df2.empty and li_hint:
        li_col   = next((c for c in df2.columns if "line item" in str(c).lower()), None)
        city_col = next((c for c in df2.columns
                         if str(c).strip().lower() in ("city", "cities", "geo", "location")), None)

        if li_col and city_col:
            hint_lower = str(li_hint).strip().lower()
            _AD_WORDS = {"banner", "video", "display", "native", "audio"}
            _tokens = [
                p.strip()
                for p in re.split(r"[-|]", hint_lower)
                if p.strip() and p.strip() not in _AD_WORDS
                and not re.match(r"^li\d+$", p.strip())
            ]

            def _li_matches(pivot_li_val):
                pl = str(pivot_li_val).strip().lower()
                return bool(_tokens) and all(t in pl for t in _tokens)

            mask    = df2[li_col].astype(str).apply(_li_matches)
            matched = df2[mask]
            if not matched.empty:
                raw_cities = matched[city_col].dropna().astype(str).str.strip().tolist()
                seen = set()
                _SKIP_CITY = {"unknown", "nan", "none", "", "n/a", "na", "other", "others"}
                for city in raw_cities:
                    key = city.lower()
                    if key not in seen and key not in _SKIP_CITY:
                        seen.add(key)
                        pivot_cities.append(city)

    db_cities_raw = load_city_db_sheet(city_sheet_names)
    db_map = {str(c["name"]).strip().lower(): c for c in db_cities_raw}

    city_list = []
    if pivot_cities:
        for city_name in pivot_cities:
            key = str(city_name).strip().lower()
            if key in db_map:
                city_list.append(db_map[key])
            else:
                city_list.append({"name": str(city_name).strip(), "weight": 1.0})
    else:
        city_list = sorted(db_cities_raw, key=lambda x: x["weight"], reverse=True)[:60]
        if not city_list:
            city_list = [{"name": c, "weight": 1.0} for c in
                         ["Sydney","Melbourne","Brisbane","Perth","Adelaide","Canberra"]]

    # Remove "Unknown" / blank city names before sorting
    _INVALID_CITY = {"unknown", "nan", "none", "", "n/a", "na", "other", "others"}
    city_list = [
        c for c in city_list
        if str(c["name"]).strip().lower() not in _INVALID_CITY
    ]

    city_list.sort(key=lambda x: str(x["name"]).lower())
    weights   = [max(c["weight"], 1.0) for c in city_list]
    total_w   = sum(weights)
    city_imps = _largest_remainder(weights, total_w, total_imp)
    city_imps = deduplicate_preserving_sum(city_imps, gap=1)

    city_clks = [round(imp * rand_float(0.0035, 0.0056)) if imp >= 180 else 0
                 for imp in city_imps]
    c_sum = sum(city_clks) or 1
    city_clks = [round(c * total_clk / c_sum) if city_imps[i] >= 180 else 0
                 for i, c in enumerate(city_clks)]
    drift = total_clk - sum(city_clks)
    if drift:
        city_clks[max(range(len(city_imps)), key=lambda i: city_imps[i])] += drift

    rows  = [{"City": city_list[i]["name"], "Impressions": city_imps[i],
              "Clicks": city_clks[i], "Click Rate (CTR)": pct(city_clks[i], city_imps[i])}
             for i in range(len(city_list))]
    total = {"City": "Grand Total", "Impressions": total_imp, "Clicks": total_clk,
             "Click Rate (CTR)": pct(total_clk, total_imp)}
    return rows, total


# ---------------------------------------------------------------------------
# App / URL sheet
# ---------------------------------------------------------------------------

def build_sheet10_apps(total_imp: int, total_clk: int,
                        language_sheet_names: list,
                        user_urls_text: str):
    """App/URL sheet — combine DB URLs + user URLs, distribute impressions."""
    _JUNK = {"nan", "none", "", "app/url", "app", "url", "site", "domain"}

    user_apps = [_clean_url(u) for u in user_urls_text.splitlines()
                 if _clean_url(u) not in _JUNK]
    user_set  = set(user_apps)
    db_urls   = [u for u in get_urls_from_sheets(language_sheet_names)
                 if u not in user_set and u not in _JUNK]
    random.shuffle(db_urls)

    if total_imp < 50_000:
        target = random.randint(40, 50);   top_lo, top_hi = 10_000, 12_000
    elif total_imp < 70_000:
        target = random.randint(50, 60);   top_lo, top_hi = 12_000, 15_000
    elif total_imp < 100_000:
        target = random.randint(60, 70);   top_lo, top_hi = 15_000, 20_000
    elif total_imp < 500_000:
        target = random.randint(110, 120); top_lo, top_hi = 18_000, 35_000
    elif total_imp < 1_000_000:
        target = random.randint(110, 120); top_lo, top_hi = 40_000, 50_000
    else:
        target = random.randint(110, 120); top_lo, top_hi = 70_000, 80_000

    n_user  = len(user_apps)
    filler  = db_urls[:max(0, target - n_user)]
    all_apps = user_apps + filler
    if not all_apps:
        all_apps = ["facebook.com", "youtube.com", "instagram.com"]
    n = len(all_apps)

    top_hi = min(top_hi, int(total_imp * 0.48)) if total_imp > 0 else top_hi
    top_lo = max(1, min(top_lo, top_hi - 500))
    target_top = random.randint(top_lo, top_hi)

    other_weights = []
    for i in range(1, n):
        if   i < 8:  w = random.uniform(0.55, 0.90)
        elif i < 20: w = random.uniform(0.08, 0.25)
        elif i < 85: w = random.uniform(0.003, 0.04)
        else:        w = random.uniform(0.0005, 0.006)
        other_weights.append(w)

    remainder_imp = max(0, total_imp - target_top)
    sum_oth = sum(other_weights) or 1.0
    if sum_oth > 0 and remainder_imp > 0:
        scale = remainder_imp / sum_oth
        other_weights = [w * scale for w in other_weights]

    imp_weights = [float(target_top)] + other_weights
    w_total     = sum(imp_weights)
    app_imps    = _largest_remainder(imp_weights, w_total, total_imp)
    app_imps    = deduplicate_preserving_sum(app_imps, gap=7)

    raw_clks = [imp * rand_float(0.0025, 0.0085) if imp >= 180 else 0.0 for imp in app_imps]
    for i in range(min(n_user, n)):
        if app_imps[i] >= 180:
            raw_clks[i] = app_imps[i] * rand_float(0.0045, 0.0052)

    w_clk    = sum(raw_clks) or float(n)
    app_clks = _largest_remainder(raw_clks, w_clk, total_clk)

    drift = total_clk - sum(app_clks)
    if drift > 0:
        app_clks[0] += drift
    elif drift < 0:
        for i in range(abs(drift)):
            idx = max(range(n), key=lambda j: app_clks[j])
            if app_clks[idx] > 1: app_clks[idx] -= 1

    # ── Cap individual CTRs at 0.90% and redistribute excess ─────────────────
    # Prevents any single URL from showing >0.90% CTR regardless of how
    # campaign-level CTR or random weight factors align.
    _MAX_ROW_CTR = 0.009   # 0.90% hard ceiling per URL row
    for _cap_pass in range(20):
        excess = 0
        for i in range(n):
            cap = int(app_imps[i] * _MAX_ROW_CTR)
            if app_clks[i] > cap:
                excess += app_clks[i] - cap
                app_clks[i] = cap
        if excess == 0:
            break
        # Spread excess to URLs that still have headroom below the cap
        headroom = [
            (int(app_imps[i] * _MAX_ROW_CTR) - app_clks[i], i)
            for i in range(n)
            if int(app_imps[i] * _MAX_ROW_CTR) > app_clks[i]
        ]
        if not headroom:
            app_clks[0] += excess   # nowhere left — safety valve
            break
        total_room = sum(r for r, _ in headroom)
        distributed = 0
        for r, idx in headroom:
            share = round(excess * r / total_room)
            app_clks[idx] += share
            distributed += share
        leftover = excess - distributed
        if leftover:
            app_clks[headroom[0][1]] += leftover

    rows = [{"App/URL": all_apps[i], "Impressions": app_imps[i],
             "Clicks": app_clks[i], "Click Rate (CTR)": pct(app_clks[i], app_imps[i])}
            for i in range(n)]
    rows.sort(key=lambda x: x["Impressions"], reverse=True)
    total = {"App/URL": "Grand Total", "Impressions": total_imp,
             "Clicks": total_clk, "Click Rate (CTR)": pct(total_clk, total_imp)}
    return rows, total


# ---------------------------------------------------------------------------
# Banner daily-report format parser
# ---------------------------------------------------------------------------

def _parse_banner_format(file_bytes: bytes):
    """
    Detect and parse the special banner daily-report Excel format.
    Data may start at ANY column (some files have leading empty columns).
    Scans all cells to find 'daily report' / 'date' keywords.

    Returns (date_df, total_imp, total_clk, banner_li_hint) if detected, else None.
    """
    try:
        raw = pd.read_excel(io.BytesIO(file_bytes), header=None)
    except Exception:
        return None

    total_imp       = 0
    total_clk       = 0
    date_header_idx = None
    data_start_col  = 0
    banner_li_hint  = ""

    for i, row in raw.iterrows():
        for j in range(len(row)):
            val = row.iloc[j]
            v   = str(val).strip().lower() if (val is not None and str(val).strip() not in ("", "nan")) else ""
            if v == "daily report":
                total_imp      = safe_int(row.iloc[j + 1]) if j + 1 < len(row) else 0
                total_clk      = safe_int(row.iloc[j + 2]) if j + 2 < len(row) else 0
                data_start_col = j
            if v == "date":
                date_header_idx = i
                data_start_col  = j
                break
        if date_header_idx is not None:
            break

    if date_header_idx is None or total_imp == 0:
        return None

    # Extract line item name: first non-empty row, column data_start+1
    for _, row2 in raw.iterrows():
        li_candidate_col = data_start_col + 1
        if li_candidate_col < len(row2):
            candidate = row2.iloc[li_candidate_col]
            if candidate is not None and str(candidate).strip() not in ("", "nan"):
                banner_li_hint = str(candidate).strip()
                break

    # Read column names from the "Date" header row starting at data_start_col
    header_row = raw.iloc[date_header_idx]
    col_names  = []
    for k in range(data_start_col, len(header_row)):
        h = header_row.iloc[k]
        col_names.append(
            str(h).strip() if (h is not None and str(h).strip() not in ("nan", "")) else None
        )

    date_rows = []
    for i in range(int(date_header_idx) + 1, len(raw)):
        row      = raw.iloc[i]
        date_val = row.iloc[data_start_col]
        if date_val is None or str(date_val).strip().lower() in ("", "nan", "none"):
            continue
        row_dict = {}
        for k, col_name in enumerate(col_names):
            actual_col = data_start_col + k
            if col_name and actual_col < len(row):
                row_dict[col_name] = row.iloc[actual_col]
        date_rows.append(row_dict)

    if not date_rows:
        return None

    date_df = pd.DataFrame(date_rows)
    if total_imp == 0 and "Impressions" in date_df.columns:
        total_imp = int(date_df["Impressions"].apply(safe_float).sum())
    if total_clk == 0 and "Clicks" in date_df.columns:
        total_clk = int(date_df["Clicks"].apply(safe_float).sum())

    return date_df, total_imp, total_clk, banner_li_hint


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def generate_report(
    campaign_file_bytes: bytes,
    campaign_filename: str,
    language_sheet_names: list,
    city_sheet_names: list,
    user_urls_text: str = "",
    mode: str = "video",
    creative_file_bytes: Optional[bytes] = None,
) -> bytes:
    """Generate a multi-sheet Excel report and return it as bytes."""
    is_banner = (mode.lower() == "banner")
    ext = campaign_filename.rsplit(".", 1)[-1].lower()

    banner_parsed  = None
    banner_li_hint = ""
    if ext not in ("csv",):
        banner_parsed = _parse_banner_format(campaign_file_bytes)

    if banner_parsed is not None:
        df_date, total_imp, total_clk, banner_li_hint = banner_parsed
        li_stem = banner_li_hint or campaign_filename.rsplit(".", 1)[0]
        df1 = pd.DataFrame({
            "Line Item Name": [li_stem],
            "Impressions":    [total_imp],
            "Clicks":         [total_clk],
        })
        # Do NOT override is_banner here — respect the mode passed by the user.
        # Box-format files can be either Banner or Video campaigns.
    else:
        if ext == "csv":
            df1 = pd.read_csv(io.BytesIO(campaign_file_bytes))
        else:
            df1 = pd.read_excel(io.BytesIO(campaign_file_bytes), header=0)
            _KEY_COLS = {"Date", "Impressions", "Clicks"}
            if not _KEY_COLS.intersection(df1.columns):
                df1 = pd.read_excel(io.BytesIO(campaign_file_bytes), header=1)

        if df1.empty:
            raise ValueError("Campaign file contains no data.")

        total_imp = int(df1.get("Impressions", pd.Series([0])).apply(safe_float).sum())
        total_clk = int(df1.get("Clicks",      pd.Series([0])).apply(safe_float).sum())

        if total_imp == 0:
            total_imp = random.randint(150_000, 300_000)
        if total_clk == 0:
            total_clk = round(total_imp * rand_float(0.0040, 0.0055))

        _VIDEO_METRIC_COLS = [
            "Viewable Impressions", "Measurable Impressions",
            "Sum of Starts (Video)", "Sum of Complete Views (Video)",
        ]
        _video_totals = {}
        for _vc in _VIDEO_METRIC_COLS:
            _col = df1.get(_vc)
            if _col is not None:
                _val = int(_col.apply(safe_float).sum())
                if _val > 0:
                    _video_totals[_vc] = _val

        if "Date" not in df1.columns:
            import datetime as _dt
            _n = 30
            _weights = [1.0] * _n
            _base = _dt.date.today() - _dt.timedelta(days=_n)
            df_date = pd.DataFrame({
                "Date":        [(_base + _dt.timedelta(days=i)).strftime("%d %B, %Y") for i in range(_n)],
                "Impressions": _largest_remainder(_weights, float(_n), total_imp),
                "Clicks":      _largest_remainder(_weights, float(_n), total_clk),
            })
            for _vc, _total in _video_totals.items():
                df_date[_vc] = _largest_remainder(_weights, float(_n), _total)
        else:
            df_date = df1

    ctr_reach = pct(total_clk, total_imp)

    s1 = build_sheet1_reach(total_imp, total_clk)
    # REACH is a single summary row — no multi-row sanitization needed

    s2, s2t = build_sheet2_date(df_date, total_imp, total_clk, ctr_reach, is_banner)
    s2, s2t = _sanitize_sheet_data(s2, s2t, total_imp, total_clk, ctr_reach)

    s3, s3t = build_sheet3_timeofday(total_imp, total_clk, ctr_reach)
    s3, s3t = _sanitize_sheet_data(s3, s3t, total_imp, total_clk, ctr_reach)

    s4, s4t = build_sheet4_age(total_imp, total_clk, ctr_reach)
    s4, s4t = _sanitize_sheet_data(s4, s4t, total_imp, total_clk, ctr_reach)

    s5, s5t = build_sheet5_gender(total_imp, total_clk, ctr_reach)
    s5, s5t = _sanitize_sheet_data(s5, s5t, total_imp, total_clk, ctr_reach)

    s6, s6t = build_sheet6_device(total_imp, total_clk, ctr_reach)
    s6, s6t = _sanitize_sheet_data(s6, s6t, total_imp, total_clk, ctr_reach)

    s7, s7t = build_sheet7_exchange(total_imp, total_clk, ctr_reach)
    s7, s7t = _sanitize_sheet_data(s7, s7t, total_imp, total_clk, ctr_reach)

    df2 = None
    if creative_file_bytes:
        try:
            buf2 = io.BytesIO(creative_file_bytes)
            df2 = pd.read_excel(buf2)
        except Exception as e:
            print(f"[report] creative_file_bytes load failed: {e}")

    # li_hint: priority -> banner_li_hint -> sheet name -> filename stem
    if banner_parsed is not None and banner_li_hint:
        li_hint = banner_li_hint
    else:
        try:
            import openpyxl as _opxl
            _wb_tmp = _opxl.load_workbook(
                io.BytesIO(campaign_file_bytes), read_only=True, data_only=True
            )
            _sheet0 = _wb_tmp.sheetnames[0].strip() if _wb_tmp.sheetnames else ""
            _wb_tmp.close()
            li_hint = _sheet0 if (_sheet0 and not re.match(r"^sheet[_\s]?\d*$", _sheet0, re.I)) \
                      else (campaign_filename.rsplit(".", 1)[0] if "." in campaign_filename else campaign_filename)
        except Exception:
            li_hint = campaign_filename.rsplit(".", 1)[0] if "." in campaign_filename else campaign_filename

    s9, s9t = build_sheet9_creative(df1, total_imp, total_clk, ctr_reach, df2, li_hint=li_hint)
    s9, s9t = _sanitize_sheet_data(s9, s9t, total_imp, total_clk, ctr_reach)

    s8, s8t = build_sheet8_city(total_imp, total_clk, ctr_reach, city_sheet_names,
                                 df2=df2, li_hint=li_hint)
    s8, s8t = _sanitize_sheet_data(s8, s8t, total_imp, total_clk, ctr_reach)

    s10, s10t = build_sheet10_apps(total_imp, total_clk, language_sheet_names, user_urls_text)
    s10, s10t = _sanitize_sheet_data(s10, s10t, total_imp, total_clk, ctr_reach)

    # ── Step 2: Pre-write cross-sheet QC fix ────────────────────────────────
    # Check all sheets against REACH reference on raw Python dicts —
    # BEFORE any data touches the workbook.  Eliminates openpyxl
    # formula-reading issues and removes the need for post-write auto-correct.
    _pre_write_qc_fix(
        sheets=[
            ("DATE",        s2,  s2t),
            ("TIME OF DAY", s3,  s3t),
            ("AGE",         s4,  s4t),
            ("GENDER",      s5,  s5t),
            ("DEVICE",      s6,  s6t),
            ("EXCHANGE",    s7,  s7t),
            ("CREATIVE",    s9,  s9t if s9 else None),
            ("CITY",        s8,  s8t),
            ("APP URL",     s10, s10t),
        ],
        total_imp=total_imp,
        total_clk=total_clk,
        ctr_reach=ctr_reach,
    )

    # Build workbook
    wb = Workbook()

    ws1 = wb.active; ws1.title = "REACH"
    write_reach_sheet(ws1, s1)

    ws2 = wb.create_sheet("DATE")
    write_date_sheet(ws2, s2, s2t, is_banner)

    ws10 = wb.create_sheet("APP URL")
    write_app_url_sheet(ws10, s10, s10t)

    ws3 = wb.create_sheet("TIME OF DAY")
    write_time_of_day_sheet(ws3, s3, s3t)

    ws7 = wb.create_sheet("EXCHANGE")
    write_exchange_sheet(ws7, s7, s7t)

    ws6 = wb.create_sheet("DEVICE")
    write_device_sheet(ws6, s6, s6t)

    ws9 = wb.create_sheet("CREATIVE")
    write_creative_sheet(ws9, s9, s9t if s9 else None)

    ws8 = wb.create_sheet("CITY")
    write_city_sheet(ws8, s8, s8t)

    ws4 = wb.create_sheet("AGE")
    write_age_sheet(ws4, s4, s4t)

    ws5 = wb.create_sheet("GENDER")
    write_gender_sheet(ws5, s5, s5t)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()
