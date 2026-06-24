"""
Final Report router
===================
POST /final-report/process
  Accepts: multipart/form-data  →  file (.xlsx / .xls)
  Returns: JSON list of processed files
    [{"name": str, "filename": str, "ad_type": "Video"|"Banner"}, ...]

GET /final-report/download/{filename}
  Returns: individual .xlsx file

GET /final-report/languages
  Returns: list of distinct sheet_name values from app_url_reference

GET /final-report/cities
  Returns: list of distinct sheet_name values from city_reference

GET /final-report/reports/{id}/reach-data
  Returns: Impressions, Clicks, CTR from REACH sheet of saved report

POST /final-report/reports/{id}/send-to-qc
  Marks report as in_qc, saves verified Impressions / Clicks / CTR
"""
from __future__ import annotations

import io
import logging
import os
import re
from datetime import datetime, timezone
from typing import Optional

from fastapi import APIRouter, Body, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse, Response

from database.crm_db import crm_engine
from services.excel_splitter import split_excel_to_files
from services.report_generator import generate_report

def _get_ctr_conn():
    """Return a raw DBAPI connection from the shared crm_engine (no hardcoded creds)."""
    return crm_engine.raw_connection()

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/final-report", tags=["final-report"])

_FINAL_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "final_report_outputs",
)
os.makedirs(_FINAL_DIR, exist_ok=True)


def _ensure_report_store():
    """Create final_report_store table + QC columns if they don't exist."""
    try:
        conn = _get_ctr_conn()
        cur  = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS final_report_store (
                id              SERIAL PRIMARY KEY,
                campaign_name   TEXT,
                report_filename TEXT NOT NULL,
                generated_at    TIMESTAMPTZ DEFAULT NOW(),
                file_data       BYTEA NOT NULL
            )
        """)
        conn.commit()

        # QC columns — safe to run repeatedly (ignore if already exists)
        qc_cols = [
            ("qc_status",       "VARCHAR DEFAULT 'draft'"),
            ("qc_impressions",  "BIGINT"),
            ("qc_clicks",       "BIGINT"),
            ("qc_ctr",          "VARCHAR"),
            ("qc_submitted_at", "TIMESTAMPTZ"),
            ("qc_results",      "TEXT"),   # JSON array of check results
        ]
        for col, col_type in qc_cols:
            try:
                cur.execute(f"ALTER TABLE final_report_store ADD COLUMN IF NOT EXISTS {col} {col_type};")
                conn.commit()
            except Exception:
                try:
                    conn.rollback()
                    # SQLite fallback (no IF NOT EXISTS support)
                    cur.execute(f"ALTER TABLE final_report_store ADD COLUMN {col} {col_type};")
                    conn.commit()
                except Exception:
                    conn.rollback()

        cur.close(); conn.close()
    except Exception as e:
        logger.warning("_ensure_report_store failed: %s", e)

_ensure_report_store()


def _save_report_to_db(campaign_name: str, report_filename: str, file_data: bytes) -> int:
    """Insert report into DB, return new row id."""
    conn = _get_ctr_conn()
    cur  = conn.cursor()
    cur.execute(
        """
        INSERT INTO final_report_store (campaign_name, report_filename, file_data)
        VALUES (%s, %s, %s) RETURNING id
        """,
        (campaign_name, report_filename, file_data),
    )
    row_id = cur.fetchone()[0]
    conn.commit()
    cur.close(); conn.close()
    return row_id

# Video detection keywords (for name-based fallback)
_VIDEO_KW = {"video", "vid", "15s", "30s", "6s", "bumper", "outstream", "instream"}
# Column headers that only appear in video reports (lowercase for case-insensitive match)
_VIDEO_COLS = {
    # Standard names used in report output
    "sum of starts (video)", "sum of complete views (video)", "vcr (completion rate)",
    # Legacy / variant names
    "start views", "complete views", "video completion rate (vcr)", "vcr",
    # Additional variants
    "video starts", "video completions", "video completion rate",
}

def _detect_ad_type_from_buf(buf_bytes: bytes) -> str:
    """
    Detect Video vs Banner by reading column headers from the Excel bytes.
    Video files contain columns like 'Sum of Starts (Video)', 'VCR (Completion Rate)', etc.
    Scans first 3 rows — box-format files have a label row before the real header row.
    Falls back to 'Banner' if columns can't be read.
    """
    try:
        import io as _io
        import openpyxl as _xl
        wb = _xl.load_workbook(_io.BytesIO(buf_bytes), read_only=True, data_only=True)
        ws = wb.active
        # Scan up to 3 rows — row 1 may be a box label, row 2 the real header
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            headers = {str(c).strip().lower() for c in row if c is not None}
            if headers & _VIDEO_COLS:
                wb.close()
                return "Video"
        wb.close()
    except Exception:
        pass
    return "Banner"

def _detect_ad_type(name: str) -> str:
    """Name-based fallback detection."""
    lower = name.lower()
    for kw in _VIDEO_KW:
        if kw in lower:
            return "Video"
    return "Banner"

def _safe_filename(name: str) -> str:
    """Strip characters that are invalid in filenames."""
    return re.sub(r'[\\/*?:"<>|]', "_", name).strip() or "Sheet"


@router.post("/process")
async def final_report_process(file: UploadFile = File(...)):
    ext = (file.filename or "").rsplit(".", 1)[-1].lower()
    if ext not in {"xlsx", "xls"}:
        raise HTTPException(status_code=400, detail="Only .xlsx and .xls files accepted.")

    raw = await file.read()
    try:
        blocks = split_excel_to_files(raw, file.filename or "upload.xlsx")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.exception("split_excel_to_files failed")
        raise HTTPException(status_code=500, detail=f"Processing error: {e}")

    stem = (file.filename or "output").rsplit(".", 1)[0]
    results = []

    for block in blocks:
        sheet_name = block["sheet_name"]
        safe_name  = _safe_filename(sheet_name)
        out_name   = f"{stem}_{safe_name}.xlsx"
        out_path   = os.path.join(_FINAL_DIR, out_name)

        # Save to disk
        buf_bytes = block["buf"].getvalue()
        try:
            with open(out_path, "wb") as f_out:
                f_out.write(buf_bytes)
        except Exception as e:
            logger.warning("Could not save final report file %s: %s", out_name, e)

        # Detect ad type:
        # 1. Try full untruncated raw name (B1 for box-type, Line Item value for LI-type)
        raw_name = block.get("raw_name", sheet_name)
        ad_type  = _detect_ad_type(raw_name)
        # 2. If name has no video keywords, fall back to column-header scan
        if ad_type == "Banner":
            ad_type = _detect_ad_type_from_buf(buf_bytes)

        results.append({
            "name":     sheet_name,
            "filename": out_name,
            "ad_type":  ad_type,
        })

    return results


@router.get("/languages")
def get_languages():
    """Return distinct sheet names from app_url_reference (ordered alphabetically)."""
    try:
        conn = _get_ctr_conn()
        cur  = conn.cursor()
        cur.execute("""
            SELECT DISTINCT sheet_name
            FROM   app_url_reference
            ORDER  BY sheet_name
        """)
        names = [row[0] for row in cur.fetchall()]
        cur.close()
        conn.close()
        return {"languages": names}
    except Exception as e:
        logger.exception("get_languages failed")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/cities")
def get_cities():
    """Return distinct sheet names from city_reference (ordered by total impressions desc)."""
    try:
        conn = _get_ctr_conn()
        cur  = conn.cursor()
        cur.execute("""
            SELECT   sheet_name
            FROM     city_reference
            GROUP BY sheet_name
            ORDER BY SUM(COALESCE(potential_impressions, 0)) DESC NULLS LAST
        """)
        names = [row[0] for row in cur.fetchall()]
        cur.close()
        conn.close()
        return {"cities": names}
    except Exception as e:
        logger.exception("get_cities failed")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/generate")
async def generate_final_report(
    filename:  str  = Form(...),        # split file already on disk in final_report_outputs/
    languages: str  = Form(""),         # comma-separated language sheet names
    cities:    str  = Form(""),         # comma-separated city sheet names
    app_urls:  str  = Form(""),         # newline-separated URLs from textarea
    mode:      str  = Form("video"),
    file2:     UploadFile = File(None), # optional: second file with Creative column
):
    """
    Generate the 10-sheet campaign report Excel from a previously split file.
    Returns JSON with report_id, filename, qc_status, qc_results, corrections.
    Frontend downloads via GET /final-report/reports/{id}/download.
    """
    safe_name = os.path.basename(filename)
    src_path  = os.path.join(_FINAL_DIR, safe_name)
    if not os.path.isfile(src_path):
        raise HTTPException(status_code=404, detail=f"Split file not found: {safe_name}")

    lang_list = [s.strip() for s in languages.split(",") if s.strip()]
    city_list = [s.strip() for s in cities.split(",")    if s.strip()]

    try:
        with open(src_path, "rb") as f:
            file_bytes = f.read()

        creative_bytes = None
        if file2 and file2.filename:
            creative_bytes = await file2.read()

        report_bytes = generate_report(
            campaign_file_bytes  = file_bytes,
            campaign_filename    = safe_name,
            language_sheet_names = lang_list,
            city_sheet_names     = city_list,
            user_urls_text       = app_urls,
            mode                 = mode,
            creative_file_bytes  = creative_bytes,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.exception("generate_report failed")
        raise HTTPException(status_code=500, detail=f"Report generation error: {e}")

    stem          = safe_name.rsplit(".", 1)[0]
    report_name   = f"report_{stem}.xlsx"
    campaign_name = stem

    # ── Save to DB first (get report_id) ─────────────────────────────────────
    report_id = None
    try:
        report_id = _save_report_to_db(campaign_name, report_name, report_bytes)
    except Exception as e:
        logger.warning("Could not save report to DB: %s", e)

    # ── Auto-QC: read REACH values directly from the generated Excel ──────────
    qc_status     = "draft"
    qc_results    = {}
    corrections   = []

    try:
        import openpyxl as _opxl
        import json as _json

        # Read Impressions, Clicks, CTR, Reach, Frequency from REACH sheet
        impressions = clicks = 0
        ctr_raw = ""
        reach = frequency = None

        _wb_tmp = _opxl.load_workbook(io.BytesIO(report_bytes), read_only=True, data_only=True)
        if "REACH" in _wb_tmp.sheetnames:
            _ws = _wb_tmp["REACH"]
            _it = _ws.iter_rows(values_only=True)
            _hdrs = [str(h).strip() if h is not None else "" for h in next(_it, [])]
            _dr   = next(_it, None)
            if _dr:
                _rd        = dict(zip(_hdrs, _dr))
                impressions = int(_rd.get("Impressions") or 0)
                clicks      = int(_rd.get("Clicks")      or 0)
                _ctr_cell   = _rd.get("Click Rate (CTR)")
                # Formula result may be None (uncached) — compute from imp/clicks as fallback
                if _ctr_cell is not None:
                    ctr_raw = str(_ctr_cell)
                elif impressions > 0:
                    ctr_raw = f"{clicks / impressions * 100:.2f}%"
                else:
                    ctr_raw = "0.00%"
                reach       = _rd.get("Reach")
                frequency   = _rd.get("Frequency") or 3
        _wb_tmp.close()

        # Run all QC checks
        all_qc = _run_all_qc(
            report_bytes, impressions, clicks, ctr_raw,
            reach, frequency,
        )

        # Auto-correct fixable failures
        corrected_bytes, corrections, all_qc = _auto_correct_report(report_bytes, all_qc)
        final_bytes = corrected_bytes if corrections else report_bytes

        # If corrections were applied, re-run QC on the corrected file for true final status
        if corrections:
            try:
                all_qc = _run_all_qc(
                    corrected_bytes, impressions, clicks, ctr_raw, reach, frequency,
                )
            except Exception as _re_err:
                logger.warning("Re-run QC after correction failed: %s", _re_err)

        # Overall status from final (post-correction) QC results
        # If corrections were applied the report is at best-effort quality — cap at "warning"
        all_checks = [c for checks in all_qc.values() for c in checks]
        has_fail   = any(c["status"] == "FAIL"    for c in all_checks)
        has_warn   = any(c["status"] == "WARNING" for c in all_checks)
        if has_fail:
            qc_status = "warning" if corrections else "rejected"
        elif has_warn:
            qc_status = "warning"
        else:
            qc_status = "approved"

        qc_results  = all_qc
        qc_json     = _json.dumps(all_qc)
        now         = datetime.now(timezone.utc)

        # Update DB row with QC results + corrected file
        if report_id:
            try:
                conn = _get_ctr_conn()
                cur  = conn.cursor()
                cur.execute(
                    """
                    UPDATE final_report_store
                    SET    qc_status       = %s,
                           qc_impressions  = %s,
                           qc_clicks       = %s,
                           qc_ctr          = %s,
                           qc_submitted_at = %s,
                           qc_results      = %s,
                           file_data       = %s
                    WHERE  id = %s
                    """,
                    (qc_status, impressions, clicks, ctr_raw,
                     now, qc_json, final_bytes, report_id),
                )
                conn.commit()
                cur.close(); conn.close()
            except Exception as e:
                logger.warning("Could not update QC results in DB: %s", e)

    except Exception as e:
        logger.warning("Auto-QC failed (report still saved): %s", e)

    return {
        "success":      True,
        "report_id":    report_id,
        "filename":     report_name,
        "qc_status":    qc_status,
        "qc_results":   qc_results,
        "corrections":  corrections,
    }


@router.get("/reports")
def list_saved_reports():
    """Return list of all saved reports from DB (without file_data)."""
    try:
        conn = _get_ctr_conn()
        cur  = conn.cursor()
        cur.execute("""
            SELECT id, campaign_name, report_filename,
                   generated_at,
                   octet_length(file_data) AS file_size,
                   COALESCE(qc_status, 'draft') AS qc_status,
                   qc_impressions,
                   qc_clicks,
                   qc_ctr,
                   qc_submitted_at
            FROM   final_report_store
            ORDER  BY generated_at DESC
        """)
        rows = cur.fetchall()
        cur.close(); conn.close()
        return [
            {
                "id":              r[0],
                "campaign_name":   r[1],
                "filename":        r[2],
                "generated_at":    r[3].isoformat() if r[3] else None,
                "file_size":       r[4],
                "qc_status":       r[5],
                "qc_impressions":  r[6],
                "qc_clicks":       r[7],
                "qc_ctr":          r[8],
                "qc_submitted_at": r[9].isoformat() if r[9] else None,
            }
            for r in rows
        ]
    except Exception as e:
        logger.exception("list_saved_reports failed")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/reports/{report_id}/download")
def download_saved_report(report_id: int):
    """Stream a saved report from DB by its ID."""
    try:
        conn = _get_ctr_conn()
        cur  = conn.cursor()
        cur.execute(
            "SELECT report_filename, file_data FROM final_report_store WHERE id = %s",
            (report_id,),
        )
        row = cur.fetchone()
        cur.close(); conn.close()
    except Exception as e:
        logger.exception("download_saved_report failed")
        raise HTTPException(status_code=500, detail=str(e))

    if not row:
        raise HTTPException(status_code=404, detail=f"Report {report_id} not found")

    fname, fdata = row
    return Response(
        content    = bytes(fdata),
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers    = {"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.delete("/reports/{report_id}")
def delete_saved_report(report_id: int):
    """Delete a saved report from DB."""
    try:
        conn = _get_ctr_conn()
        cur  = conn.cursor()
        cur.execute("DELETE FROM final_report_store WHERE id = %s RETURNING id", (report_id,))
        deleted = cur.fetchone()
        conn.commit()
        cur.close(); conn.close()
    except Exception as e:
        logger.exception("delete_saved_report failed")
        raise HTTPException(status_code=500, detail=str(e))

    if not deleted:
        raise HTTPException(status_code=404, detail=f"Report {report_id} not found")
    return {"deleted": report_id}


@router.post("/update-app-urls")
async def update_app_urls(file: UploadFile = File(...)):
    """
    Replace app_url_reference rows from an uploaded Excel file.
    Each sheet = one language. Extracts URLs from 3 boxes (cols B, F, J).
    Skips the 'Summary' sheet.
    Returns counts per sheet.
    """
    ext = (file.filename or "").rsplit(".", 1)[-1].lower()
    if ext not in {"xlsx", "xls"}:
        raise HTTPException(status_code=400, detail="Only .xlsx / .xls accepted.")

    raw = await file.read()

    try:
        import openpyxl as _xl
        from psycopg2.extras import execute_values as _ev

        wb = _xl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        results = []

        conn = _get_ctr_conn()
        cur  = conn.cursor()

        for sheet_name in wb.sheetnames:
            if sheet_name.strip().lower() == "summary":
                continue

            ws   = wb[sheet_name]
            rows_iter = list(ws.iter_rows(values_only=True))
            if not rows_iter:
                continue

            # Skip header row (row 0), extract from data rows
            data_rows = rows_iter[1:]

            # Box columns: (id_col, url_col)
            boxes = [(0, 1), (4, 5), (8, 9)]
            records = []
            for row in data_rows:
                for id_ci, url_ci in boxes:
                    uid  = row[id_ci]  if len(row) > id_ci  else None
                    url  = row[url_ci] if len(row) > url_ci else None
                    if url is None:
                        continue
                    url_str = str(url).strip()
                    if not url_str or url_str.lower() in ("none", "nan", "sites"):
                        continue
                    records.append((sheet_name.strip(), int(uid) if uid else None, url_str))

            # DELETE existing rows for this sheet_name, then INSERT fresh
            cur.execute(
                "DELETE FROM app_url_reference WHERE sheet_name = %s",
                (sheet_name.strip(),),
            )
            if records:
                _ev(
                    cur,
                    "INSERT INTO app_url_reference (sheet_name, url_id, url) VALUES %s",
                    records,
                    page_size=500,
                )
            conn.commit()
            results.append({"sheet": sheet_name.strip(), "inserted": len(records)})
            logger.info("app_url_reference updated: %s → %d rows", sheet_name.strip(), len(records))

        cur.close()
        conn.close()
        wb.close()

        total = sum(r["inserted"] for r in results)
        return {"success": True, "sheets_updated": len(results), "total_urls": total, "details": results}

    except HTTPException:
        raise
    except Exception as e:
        logger.exception("update_app_urls failed")
        raise HTTPException(status_code=500, detail=f"Update failed: {e}")


@router.get("/download")
def download_final_report_file(file: str):
    safe_name = os.path.basename(file)   # strip any path traversal
    filepath  = os.path.join(_FINAL_DIR, safe_name)
    if not os.path.isfile(filepath):
        raise HTTPException(status_code=404, detail=f"File not found: {safe_name}")
    return FileResponse(
        filepath,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}"'},
    )


# ── QC endpoints ──────────────────────────────────────────────────────────────

@router.get("/reports/{report_id}/reach-data")
def get_reach_data(report_id: int):
    """
    Read the REACH sheet from the saved report Excel and return
    Impressions, Clicks, Click Rate (CTR).
    """
    try:
        conn = _get_ctr_conn()
        cur  = conn.cursor()
        cur.execute(
            "SELECT report_filename, file_data FROM final_report_store WHERE id = %s",
            (report_id,),
        )
        row = cur.fetchone()
        cur.close(); conn.close()
    except Exception as e:
        logger.exception("get_reach_data DB fetch failed")
        raise HTTPException(status_code=500, detail=str(e))

    if not row:
        raise HTTPException(status_code=404, detail=f"Report {report_id} not found")

    fname, fdata = row
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(bytes(fdata)), read_only=True, data_only=True)

        if "REACH" not in wb.sheetnames:
            raise HTTPException(status_code=422, detail="REACH sheet not found in this report")

        ws = wb["REACH"]
        rows_iter = ws.iter_rows(values_only=True)

        # Row 1 = headers, Row 2 = data
        headers = [str(h).strip() if h is not None else "" for h in next(rows_iter, [])]
        data_row = next(rows_iter, None)
        wb.close()

        if data_row is None:
            raise HTTPException(status_code=422, detail="REACH sheet has no data row")

        row_dict = dict(zip(headers, data_row))

        impressions = row_dict.get("Impressions") or row_dict.get("impressions") or 0
        clicks      = row_dict.get("Clicks")      or row_dict.get("clicks")      or 0
        ctr         = row_dict.get("Click Rate (CTR)") or row_dict.get("CTR") or "0.00%"

        return {
            "report_id":   report_id,
            "filename":    fname,
            "impressions": int(impressions) if impressions else 0,
            "clicks":      int(clicks)      if clicks      else 0,
            "ctr":         str(ctr),
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("get_reach_data parse failed")
        raise HTTPException(status_code=500, detail=f"Could not read REACH sheet: {e}")


@router.post("/reports/{report_id}/send-to-qc")
def send_to_qc(
    report_id:      int,
    impressions:    int           = Body(...),
    clicks:         int           = Body(...),
    ctr:            str           = Body(...),
    campaign_start: Optional[str] = Body(None),
    campaign_end:   Optional[str] = Body(None),
):
    """
    Run comprehensive QC checks on all sheets of the report.
    Returns grouped check results per sheet plus overall APPROVED / WARNING / REJECTED status.
    """
    try:
        conn = _get_ctr_conn()
        cur  = conn.cursor()
        cur.execute(
            "SELECT id, file_data FROM final_report_store WHERE id = %s",
            (report_id,),
        )
        row = cur.fetchone()
        if not row:
            cur.close(); conn.close()
            raise HTTPException(status_code=404, detail=f"Report {report_id} not found")

        file_data = row[1]

        # Read Reach + Frequency from REACH sheet for REACH QC
        reach = frequency = None
        try:
            import openpyxl as _opxl
            _wb = _opxl.load_workbook(io.BytesIO(bytes(file_data)), read_only=True, data_only=True)
            if "REACH" in _wb.sheetnames:
                _ws = _wb["REACH"]
                _it = _ws.iter_rows(values_only=True)
                _hdrs = [str(h).strip() if h is not None else "" for h in next(_it, [])]
                _dr   = next(_it, None)
                if _dr:
                    _rd = dict(zip(_hdrs, _dr))
                    reach     = _rd.get("Reach") or _rd.get("reach")
                    frequency = _rd.get("Frequency") or _rd.get("frequency") or 3
            _wb.close()
        except Exception as _e:
            logger.warning("Could not read Reach/Frequency: %s", _e)

        # Run all QC checks
        all_qc = _run_all_qc(
            bytes(file_data), impressions, clicks, ctr,
            reach, frequency, campaign_start, campaign_end
        )

        # ── Auto-correct fixable failures ────────────────────────────
        corrected_bytes, corrections_log, all_qc = _auto_correct_report(
            bytes(file_data), all_qc
        )

        # Compute overall status (CORRECTED counts as PASS for overall)
        all_checks = [c for checks in all_qc.values() for c in checks]
        if any(c["status"] == "FAIL" for c in all_checks):
            overall = "rejected"
        elif any(c["status"] == "WARNING" for c in all_checks):
            overall = "warning"
        else:
            overall = "approved"

        import json as _json
        qc_json = _json.dumps(all_qc)

        now = datetime.now(timezone.utc)

        if corrections_log:
            # Save corrected file + QC results
            cur.execute(
                """
                UPDATE final_report_store
                SET    qc_status       = %s,
                       qc_impressions  = %s,
                       qc_clicks       = %s,
                       qc_ctr          = %s,
                       qc_submitted_at = %s,
                       qc_results      = %s,
                       file_data       = %s
                WHERE  id = %s
                """,
                (overall, impressions, clicks, ctr, now, qc_json,
                 corrected_bytes, report_id),
            )
        else:
            cur.execute(
                """
                UPDATE final_report_store
                SET    qc_status       = %s,
                       qc_impressions  = %s,
                       qc_clicks       = %s,
                       qc_ctr          = %s,
                       qc_submitted_at = %s,
                       qc_results      = %s
                WHERE  id = %s
                """,
                (overall, impressions, clicks, ctr, now, qc_json, report_id),
            )
        conn.commit()
        cur.close(); conn.close()

        return {
            "success":       True,
            "report_id":     report_id,
            "qc_status":     overall,
            "qc_results":    all_qc,
            "corrections":   corrections_log,
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("send_to_qc failed")
        raise HTTPException(status_code=500, detail=str(e))


# ─── QC helper utilities ──────────────────────────────────────────────────────

def _qsafe(v):
    """Parse any cell value to float, returns 0.0 on failure."""
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    try: return float(str(v).replace(",", "").replace("%", "").strip())
    except: return 0.0


def _qpct(v):
    """
    Parse a percentage value to float representing the percentage number.
    e.g.  "80.5%"  ->  80.5   (already has % sign → keep as-is)
          0.805    ->  80.5   (decimal < 1 → multiply by 100)
          80.5     ->  80.5
    """
    if v is None: return 0.0
    if isinstance(v, (int, float)):
        f = float(v)
        return round(f * 100, 4) if 0 < abs(f) < 1.0 else f
    # String path — check if it already carries a "%" sign
    has_pct = "%" in str(v)
    s = str(v).replace("%", "").replace(",", "").strip()
    try:
        f = float(s)
        # If the string had a "%" it is already expressed as percentage — don't multiply
        if has_pct:
            return f
        return round(f * 100, 4) if 0 < abs(f) < 1.0 else f
    except: return 0.0


def _qread(wb, name):
    """
    Read sheet from workbook.
    Returns (headers, data_rows, grand_total_dict) or (None, [], None) if missing.
    Grand Total row detected by first cell containing 'grand total'.

    Formula-GT fix: older reports write Grand Total as =SUM(...) formulas.
    openpyxl data_only=True returns None for uncached formulas.
    After reading, if a GT numeric key is None, we backfill it by summing
    the corresponding data rows so QC checks don't produce false FAILs.
    """
    if name not in wb.sheetnames:
        return None, [], None
    ws = wb[name]
    it = ws.iter_rows(values_only=True)
    hdrs_raw = next(it, [])
    headers  = [str(h).strip() if h is not None else "" for h in hdrs_raw]
    data, gt = [], None
    for row in it:
        if not any(v is not None for v in row): continue
        rd    = dict(zip(headers, row))
        first = str(row[0]).strip().lower() if row and row[0] is not None else ""
        if "grand total" in first or first == "total":
            gt = rd
        else:
            data.append(rd)
    # ── Formula-GT backfill ──────────────────────────────────────────────────
    # If GT row exists but a numeric key is None (SUM formula, uncached),
    # compute it from the data rows so QC doesn't produce false GT=0 FAILs.
    if gt is not None and data:
        for h in headers:
            if gt.get(h) is None:
                col_sum = sum(_qsafe(r.get(h)) for r in data)
                if col_sum > 0:
                    gt[h] = col_sum
    # ─────────────────────────────────────────────────────────────────────────
    return headers, data, gt


def _qchk(check, formula, expected, actual, status):
    """Build a standardised check result dict."""
    return {"check": check, "formula": formula,
            "expected": expected, "actual": actual, "status": status}


# ─── Cross-sheet Grand Total check ───────────────────────────────────────────

# Each tuple: (workbook sheet name, impressions column header, clicks column header)
_CROSS_SHEET_TARGETS = [
    ("DATE",        "Impressions", "Clicks"),
    ("APP URL",     "Impressions", "Clicks"),
    ("TIME OF DAY", "Impressions", "Clicks"),
    ("EXCHANGE",    "Impressions", "Clicks"),
    ("DEVICE",      "Impressions", "Clicks"),
    ("CREATIVE",    "Impressions", "Clicks"),
    ("CITY",        "Impressions", "Clicks"),
    ("AGE",         "Impressions", "Clicks"),
    ("GENDER",      "Impressions", "Clicks"),
]


def _rescale_to_total(values: list, target: int) -> list:
    """
    Proportionally rescale a list of non-negative integers so they sum exactly
    to `target`.  Uses largest-remainder (Hamilton) method for rounding.
    """
    current = sum(values)
    if current == 0:
        base, rem = divmod(target, max(len(values), 1))
        return [base + (1 if i < rem else 0) for i in range(len(values))]
    ratio   = target / current
    scaled  = [v * ratio for v in values]
    floors  = [int(s) for s in scaled]
    order   = sorted(range(len(scaled)), key=lambda i: scaled[i] - floors[i], reverse=True)
    diff    = target - sum(floors)
    for i in range(diff):
        floors[order[i]] += 1
    return floors


def _run_cross_sheet_gt_qc(wb, ref_imp: int, ref_clk: int) -> list:
    """
    Check Grand Total Impressions, Clicks, and CTR on every data sheet against
    the REACH reference.  Returns list of check dicts for "CROSS-SHEET GT".

    Structure:
      1. Grand Total summary (3 checks): Impressions / Clicks / CTR from REACH
      2. Per-sheet checks: Impressions match + Clicks match for each sheet
    """
    results = []

    # ── Compute reference CTR ──────────────────────────────────────────────────
    ref_ctr_pct = round(ref_clk / ref_imp * 100, 4) if ref_imp > 0 else 0.0
    ref_ctr_str = f"{ref_ctr_pct:.2f}%"

    # ── 1. Grand Total summary checks (always shown at top) ───────────────────
    results.append(_qchk(
        "Grand Total — Impressions",
        "Total Impressions across all sheets (from REACH)",
        f"{ref_imp:,}",
        f"{ref_imp:,}" if ref_imp > 0 else "0",
        "PASS" if ref_imp > 0 else "WARNING",
    ))
    results.append(_qchk(
        "Grand Total — Clicks",
        "Total Clicks across all sheets (from REACH)",
        f"{ref_clk:,}",
        f"{ref_clk:,}" if ref_clk > 0 else "0",
        "PASS" if ref_clk > 0 else "WARNING",
    ))
    results.append(_qchk(
        "Grand Total — Click Rate (CTR)",
        "Clicks / Impressions × 100  (from REACH)",
        ref_ctr_str,
        ref_ctr_str if ref_imp > 0 else "—",
        "PASS" if ref_imp > 0 else "WARNING",
    ))

    if ref_imp <= 0:
        return results

    # ── 2. Per-sheet Impressions + Clicks checks ───────────────────────────────
    for sh_name, imp_col, clk_col in _CROSS_SHEET_TARGETS:
        # Resolve workbook sheet name (exact → case-insensitive → partial)
        ws_name = None
        sh_upper = sh_name.upper()
        for sn in wb.sheetnames:
            if sn.upper() == sh_upper:
                ws_name = sn; break
        if ws_name is None:
            for sn in wb.sheetnames:
                if sh_upper in sn.upper():
                    ws_name = sn; break
        if ws_name is None:
            continue   # Sheet absent — handled by its own per-sheet check

        _hdrs, _rows, _gt = _qread(wb, ws_name)
        if _hdrs is None or not _rows:
            continue

        # ── Impressions check ─────────────────────────────────────────────────
        row_imp_sum = int(sum(_qsafe(r.get(imp_col, 0)) for r in _rows))
        gt_imp      = int(_qsafe(_gt.get(imp_col, 0))) if _gt else 0
        sheet_imp   = gt_imp if gt_imp > 0 else row_imp_sum
        diff_imp    = abs(sheet_imp - ref_imp)
        ok_imp      = diff_imp <= 1
        results.append(_qchk(
            f"Cross-Sheet GT — {sh_name}  (Impressions)",
            f"Sheet Grand Total Impressions = {ref_imp:,}",
            f"{ref_imp:,}",
            f"GT = {sheet_imp:,}  diff = {diff_imp:,}" if not ok_imp
            else f"GT = {sheet_imp:,}  ✓",
            "PASS" if ok_imp else "FAIL",
        ))

        # ── Clicks check ──────────────────────────────────────────────────────
        if ref_clk > 0:
            row_clk_sum = int(sum(_qsafe(r.get(clk_col, 0)) for r in _rows))
            gt_clk      = int(_qsafe(_gt.get(clk_col, 0))) if _gt else 0
            sheet_clk   = gt_clk if gt_clk > 0 else row_clk_sum
            diff_clk    = abs(sheet_clk - ref_clk)
            ok_clk      = diff_clk <= 1
            results.append(_qchk(
                f"Cross-Sheet GT — {sh_name}  (Clicks)",
                f"Sheet Grand Total Clicks = {ref_clk:,}",
                f"{ref_clk:,}",
                f"GT = {sheet_clk:,}  diff = {diff_clk:,}" if not ok_clk
                else f"GT = {sheet_clk:,}  ✓",
                "PASS" if ok_clk else "FAIL",
            ))

        # ── CTR check ─────────────────────────────────────────────────────────
        ctr_key = next(
            (h for h in (_hdrs or []) if "click rate" in h.lower() or h.lower() == "ctr"),
            None,
        )
        if ctr_key and _gt:
            sh_ctr_raw = _gt.get(ctr_key)
            sh_ctr_pct = _qpct(sh_ctr_raw)
            # Fallback: compute from GT impressions/clicks
            if sh_ctr_pct == 0 and sheet_imp > 0:
                sh_ctr_pct = round(
                    (gt_clk if ref_clk > 0 else row_clk_sum) / sheet_imp * 100, 4
                ) if (ref_clk > 0 or row_clk_sum > 0) else 0.0
            diff_ctr = abs(sh_ctr_pct - ref_ctr_pct)
            ok_ctr   = diff_ctr < 0.1
            results.append(_qchk(
                f"Cross-Sheet GT — {sh_name}  (CTR)",
                f"Sheet CTR = REACH {ref_ctr_str}",
                ref_ctr_str,
                f"{sh_ctr_pct:.2f}%  diff = {diff_ctr:.3f}%" if not ok_ctr
                else f"{sh_ctr_pct:.2f}%  ✓",
                "PASS" if ok_ctr else "WARNING",
            ))

    return results


# ─── Auto-correction engine ───────────────────────────────────────────────────

def _auto_correct_report(file_bytes: bytes, qc_results: dict):
    """
    Apply automatic corrections to fixable QC failures in the Excel report.

    Correctable failures
    ────────────────────
    • Grand Total rows              — replace =SUM() formulas with computed values
    • Date Ordering (DATE)          — sort rows chronologically
    • Impression Sorting (APP URL)  — sort rows by Impressions descending
    • Cross-Sheet GT mismatch       — proportionally rescale row impressions/clicks
                                      in any sheet whose total ≠ REACH reference

    Returns
    ───────
    (corrected_bytes, corrections_log, updated_qc_results)
      corrections_log  : [{"sheet", "check", "action"}]
      updated_qc_results: same structure as qc_results with FAIL→CORRECTED for fixed items
    """
    import io as _io
    from openpyxl import load_workbook as _lwb

    wb   = _lwb(_io.BytesIO(file_bytes), data_only=True)
    log  = []
    qc_updated = {sn: [dict(c) for c in checks] for sn, checks in qc_results.items()}

    # ── helper: find workbook sheet for a QC sheet label ─────────────
    _ALIASES = {
        "REACH":       ["REACH", "Reach"],
        "DATE":        ["DATE", "Date"],
        "BANNER":      ["BANNER", "Banner"],
        "VIDEO":       ["VIDEO", "Video"],
        "APP URL":     ["APP URL", "App URL", "App/URL"],
        "TIME OF DAY": ["TIME OF DAY", "Time of Day", "Time Of Day"],
    }

    def _ws_for(qc_sheet):
        for name in _ALIASES.get(qc_sheet, [qc_sheet]):
            if name in wb.sheetnames:
                return wb[name]
        ql = qc_sheet.lower()
        for sn in wb.sheetnames:
            if ql in sn.lower():
                return wb[sn]
        return None

    def _mark_corrected(sheet_name, check_name, action):
        log.append({"sheet": sheet_name, "check": check_name, "action": action})
        for chk in qc_updated.get(sheet_name, []):
            if chk["check"] == check_name and chk["status"] == "FAIL":
                chk["status"]  = "CORRECTED"
                chk["actual"] += f"  →  {action}"
                break

    for sheet_name, checks in qc_results.items():
        failed = {c["check"]: c for c in checks if c["status"] == "FAIL"}
        if not failed:
            continue

        # ── Fix D: Cross-sheet Grand Total alignment ──────────────────────────
        # The "CROSS-SHEET GT" key has no matching workbook sheet — handle here.
        if sheet_name == "CROSS-SHEET GT":
            # Read reference impressions + clicks from REACH sheet
            _ref_imp = _ref_clk = 0
            _reach_ws = _ws_for("REACH")
            if _reach_ws:
                _it2 = _reach_ws.iter_rows(values_only=True)
                _rhdrs2 = [str(h).strip() if h is not None else ""
                            for h in next(_it2, [])]
                _rdr2 = next(_it2, None)
                if _rdr2:
                    _rd2     = dict(zip(_rhdrs2, _rdr2))
                    _ref_imp = int(_rd2.get("Impressions") or 0)
                    _ref_clk = int(_rd2.get("Clicks")      or 0)

            if _ref_imp > 0:
                for _chk_name in list(failed.keys()):
                    # Skip summary rows and non-Impressions sub-checks
                    # Only "Cross-Sheet GT — SHEETNAME  (Impressions)" triggers a rewrite
                    if not _chk_name.startswith("Cross-Sheet GT —"):
                        continue
                    if "(Impressions)" not in _chk_name:
                        continue
                    # Extract target sheet: "Cross-Sheet GT — CITY  (Impressions)" → "CITY"
                    _target = (_chk_name
                                .replace("Cross-Sheet GT —", "")
                                .replace("(Impressions)", "")
                                .strip())
                    _tws = _ws_for(_target)
                    if _tws is None:
                        continue

                    _t_all = list(_tws.iter_rows())
                    if not _t_all:
                        continue

                    _t_hdrs = [str(c.value).strip() if c.value is not None else ""
                                for c in _t_all[0]]
                    _t_hidx = {h: i for i, h in enumerate(_t_hdrs) if h}
                    _imp_ci = _t_hidx.get("Impressions")
                    _clk_ci = _t_hidx.get("Clicks")
                    _ctr_ci = _t_hidx.get("Click Rate (CTR)")
                    if _imp_ci is None:
                        continue

                    # Separate data rows from Grand Total row
                    _gt_ri    = None
                    _data_ris = []
                    for _ri, _row in enumerate(_t_all[1:], 1):
                        _fv = str(_row[0].value or "").strip().lower()
                        if "grand total" in _fv or _fv == "total":
                            _gt_ri = _ri
                        else:
                            _data_ris.append(_ri)

                    if not _data_ris:
                        continue

                    # Current impressions of data rows
                    _old_imps = [int(_t_all[_ri][_imp_ci].value or 0)
                                  for _ri in _data_ris]
                    _cur_sum  = sum(_old_imps)
                    if _cur_sum == 0 or abs(_cur_sum - _ref_imp) <= 1:
                        continue

                    # Proportionally rescale impressions to match ref
                    _new_imps = _rescale_to_total(_old_imps, _ref_imp)

                    # Proportionally rescale clicks to match ref
                    _new_clks = None
                    if _clk_ci is not None and _ref_clk > 0:
                        _old_clks = [int(_t_all[_ri][_clk_ci].value or 0)
                                      for _ri in _data_ris]
                        _old_clk_sum = sum(_old_clks)
                        if _old_clk_sum > 0:
                            _new_clks = _rescale_to_total(_old_clks, _ref_clk)

                    # Write rescaled values + recalculate CTR per row
                    for _j, _ri in enumerate(_data_ris):
                        _tws.cell(row=_ri + 1, column=_imp_ci + 1,
                                   value=_new_imps[_j])
                        if _new_clks is not None:
                            _tws.cell(row=_ri + 1, column=_clk_ci + 1,
                                       value=_new_clks[_j])
                            if _ctr_ci is not None and _new_imps[_j] > 0:
                                _new_ctr = _new_clks[_j] / _new_imps[_j] * 100
                                _tws.cell(row=_ri + 1, column=_ctr_ci + 1,
                                           value=f"{_new_ctr:.2f}%")

                    # Update Grand Total row to reference values
                    if _gt_ri is not None:
                        _tws.cell(row=_gt_ri + 1, column=_imp_ci + 1,
                                   value=_ref_imp)
                        if _new_clks is not None and _clk_ci is not None:
                            _tws.cell(row=_gt_ri + 1, column=_clk_ci + 1,
                                       value=_ref_clk)

                    _mark_corrected(
                        sheet_name, _chk_name,
                        f"Row impressions rescaled {_cur_sum:,} → {_ref_imp:,}; "
                        f"Grand Total updated",
                    )

            continue  # Done with CROSS-SHEET GT — skip normal per-sheet fixes

        ws = _ws_for(sheet_name)
        if ws is None:
            continue

        all_rows = list(ws.iter_rows())
        if not all_rows:
            continue

        headers = [str(c.value).strip() if c.value is not None else "" for c in all_rows[0]]
        h_idx   = {h: i for i, h in enumerate(headers) if h}

        # Find grand-total row and data rows
        gt_row_i  = None   # 0-based index in all_rows
        data_end_i = len(all_rows) - 1
        for i, row in enumerate(all_rows[1:], 1):
            fv = str(row[0].value or "").strip().lower()
            if "grand total" in fv or fv == "total":
                gt_row_i   = i
                data_end_i = i - 1
                break
        data_rows = all_rows[1: data_end_i + 1]

        # ── Fix A: Date Ordering ─────────────────────────────────────
        if sheet_name == "DATE" and "Date Ordering" in failed:
            date_ci = h_idx.get("Date", 0)
            vals = [(r[date_ci].value, [c.value for c in r]) for r in data_rows]
            try:
                sorted_vals = sorted(vals, key=lambda x: x[0] if x[0] else "")
                if [v[0] for v in sorted_vals] != [v[0] for v in vals]:
                    for i, (_, row_vals) in enumerate(sorted_vals):
                        for j, v in enumerate(row_vals):
                            ws.cell(row=i + 2, column=j + 1, value=v)
                    _mark_corrected(sheet_name, "Date Ordering",
                                    "Rows sorted chronologically")
            except Exception:
                pass

        # ── Fix B: Impression Sorting (APP URL) ─────────────────────
        if sheet_name == "APP URL" and "Impression Sorting" in failed:
            imp_ci = h_idx.get("Impressions")
            if imp_ci is not None:
                vals = [(float(r[imp_ci].value or 0), [c.value for c in r])
                        for r in data_rows]
                sorted_vals = sorted(vals, key=lambda x: x[0], reverse=True)
                if [v[0] for v in sorted_vals] != [v[0] for v in vals]:
                    for i, (_, row_vals) in enumerate(sorted_vals):
                        for j, v in enumerate(row_vals):
                            ws.cell(row=i + 2, column=j + 1, value=v)
                    _mark_corrected(sheet_name, "Impression Sorting",
                                    "URLs re-sorted by Impressions (highest first)")

        # ── Fix C: Grand Total recalculation (with row-rescaling if needed) ───
        # Read REACH reference once per sheet that has GT failures
        _fixc_ref_imp = _fixc_ref_clk = 0
        if gt_row_i is not None and any("Grand Total" in n for n in failed):
            _rws_c = _ws_for("REACH")
            if _rws_c:
                _it_c = _rws_c.iter_rows(values_only=True)
                _rh_c = [str(h).strip() if h is not None else "" for h in next(_it_c, [])]
                _dr_c = next(_it_c, None)
                if _dr_c:
                    _rd_c = dict(zip(_rh_c, _dr_c))
                    _fixc_ref_imp = int(_rd_c.get("Impressions") or 0)
                    _fixc_ref_clk = int(_rd_c.get("Clicks")      or 0)

        if gt_row_i is not None:
            for check_name in list(failed.keys()):
                if "Grand Total" not in check_name:
                    continue
                col_label = check_name.replace("Grand Total —", "").strip()
                col_ci = None
                for h, idx in h_idx.items():
                    if h.lower() == col_label.lower() or col_label.lower() in h.lower():
                        col_ci = idx
                        break
                if col_ci is None:
                    continue

                # Current row sum
                col_sum = sum(
                    float(r[col_ci].value)
                    for r in data_rows
                    if isinstance(r[col_ci].value, (int, float))
                )
                if col_sum <= 0:
                    continue

                # Determine reference target for this column
                col_lower = col_label.lower()
                if "impression" in col_lower and _fixc_ref_imp > 0:
                    ref_target = _fixc_ref_imp
                elif col_lower == "clicks" and _fixc_ref_clk > 0:
                    ref_target = _fixc_ref_clk
                else:
                    ref_target = int(col_sum)   # no reference → use row sum as-is

                action_parts = []

                # If rows don't sum to the reference, rescale them proportionally
                if abs(col_sum - ref_target) > 1:
                    old_vals = [int(r[col_ci].value or 0) for r in data_rows]
                    new_vals = _rescale_to_total(old_vals, ref_target)
                    for i, row in enumerate(data_rows):
                        ws.cell(row=row[0].row, column=col_ci + 1, value=new_vals[i])
                        # Recalculate per-row CTR if we changed impressions or clicks
                        _ctr_ci2 = h_idx.get("Click Rate (CTR)")
                        if _ctr_ci2 is not None:
                            _imp_ci2 = h_idx.get("Impressions")
                            _clk_ci2 = h_idx.get("Clicks")
                            if _imp_ci2 is not None and _clk_ci2 is not None:
                                _i2 = int(ws.cell(row=row[0].row, column=_imp_ci2 + 1).value or 0)
                                _c2 = int(ws.cell(row=row[0].row, column=_clk_ci2 + 1).value or 0)
                                if _i2 > 0:
                                    ws.cell(row=row[0].row, column=_ctr_ci2 + 1,
                                             value=f"{_c2 / _i2 * 100:.2f}%")
                    action_parts.append(f"rows rescaled {int(col_sum):,} → {ref_target:,}")
                    col_sum = ref_target

                # Write the corrected Grand Total cell
                new_val = int(col_sum) if float(col_sum).is_integer() else round(col_sum, 4)
                ws.cell(row=gt_row_i + 1, column=col_ci + 1, value=new_val)
                action_parts.append(f"GT = {int(col_sum):,}")
                _mark_corrected(sheet_name, check_name, "  |  ".join(action_parts))

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read(), log, qc_updated


# ─── Per-sheet QC functions ───────────────────────────────────────────────────

def _run_reach_qc(impressions, clicks, ctr_raw, reach, frequency):
    results = []
    freq = int(frequency) if frequency else 3

    # 1. Reach Formula
    if reach is not None:
        try:
            ri     = int(reach)
            offset = ri - (impressions / freq if freq else 0)
            ok     = 200 <= offset <= 300
            results.append(_qchk(
                "Reach Formula", "(Impressions / Frequency) + offset = Reach",
                "offset 200 – 300",
                f"offset={offset:.0f}  |  ({impressions:,} / {freq}) + {offset:.0f} = {ri:,}",
                "PASS" if ok else "FAIL",
            ))
        except Exception:
            results.append(_qchk("Reach Formula", "(Impressions / Frequency) + offset = Reach",
                                  "offset 200 – 300", str(reach), "FAIL"))
    else:
        results.append(_qchk("Reach Formula", "(Impressions / Frequency) + offset = Reach",
                              "offset 200 – 300", "Reach value not found", "FAIL"))

    # 2. Frequency
    results.append(_qchk("Frequency", "Frequency = 3 (fixed)", "3", str(freq),
                          "PASS" if freq == 3 else "FAIL"))

    # 3. CTR
    try:
        act_ctr = _qpct(ctr_raw)
        exp_ctr = (clicks / impressions * 100) if impressions else 0.0
        ok      = abs(act_ctr - exp_ctr) < 0.1
        results.append(_qchk("CTR Calculation", "Clicks / Impressions x 100",
                              f"{exp_ctr:.2f}%", ctr_raw or "—",
                              "PASS" if ok else "FAIL"))
    except Exception:
        results.append(_qchk("CTR Calculation", "Clicks / Impressions x 100",
                              "calculated", ctr_raw or "—", "FAIL"))

    # 4. Grand Total fields present
    ok = impressions > 0 and clicks > 0 and bool((ctr_raw or "").strip())
    results.append(_qchk(
        "Grand Total Fields", "Impressions, Clicks, CTR all present", "All non-zero",
        f"Imp={impressions:,}  Clk={clicks:,}  CTR={ctr_raw or 'missing'}",
        "PASS" if ok else "FAIL",
    ))
    return results


def _run_date_qc(wb, campaign_start=None, campaign_end=None):
    results = []
    headers, rows, gt = _qread(wb, "DATE")
    if headers is None:
        return [_qchk("DATE Sheet", "Sheet exists in workbook", "Present", "Sheet not found", "FAIL")]

    dates   = [str(r.get("Date", "")).strip() for r in rows if r.get("Date")]
    imp_key = next((h for h in headers if "impression" in h.lower()), None)
    clk_key = next((h for h in headers if h.lower() == "clicks"), None)

    # 1. Duplicate dates
    seen, dups = set(), []
    for d in dates:
        dl = d.lower()
        if dl and dl not in ("nan", "none"):
            if dl in seen: dups.append(d)
            else: seen.add(dl)
    results.append(_qchk(
        "Duplicate Dates", "No date appears more than once",
        "No duplicates",
        f"Duplicates found: {dups[:5]}" if dups else f"No duplicates ({len(dates)} dates)",
        "FAIL" if dups else "PASS",
    ))

    # 2. Missing dates (gap detection)
    import datetime as _dt
    _FMTS = ["%d %B, %Y", "%d %B %Y", "%Y-%m-%d", "%d/%m/%Y"]
    parsed = []
    for d in dates:
        p = None
        for fmt in _FMTS:
            try: p = _dt.datetime.strptime(d.strip(), fmt); break
            except: pass
        parsed.append(p)
    valid = sorted(p for p in parsed if p is not None)
    if len(valid) >= 2:
        expected_count = (valid[-1] - valid[0]).days + 1
        missing_count  = expected_count - len(valid)
        results.append(_qchk(
            "Missing Dates", "No gaps in the date sequence",
            "Continuous daily records",
            f"Range: {valid[0].strftime('%d %b')} – {valid[-1].strftime('%d %b')} "
            f"({expected_count} days expected, {len(valid)} found, {missing_count} missing)",
            "FAIL" if missing_count > 0 else "PASS",
        ))
    else:
        results.append(_qchk("Missing Dates", "No gaps in date sequence",
                              "Continuous", f"Only {len(valid)} parseable dates", "WARNING"))

    # 3. Date ordering
    if len(valid) >= 2:
        in_order = all(valid[i] <= valid[i+1] for i in range(len(valid)-1))
        results.append(_qchk(
            "Date Ordering", "Dates in chronological order", "Ascending",
            "Chronological" if in_order else "Dates out of order",
            "PASS" if in_order else "FAIL",
        ))

    # 4. Grand Total – Impressions
    if gt and imp_key:
        s = sum(_qsafe(r.get(imp_key)) for r in rows)
        g = _qsafe(gt.get(imp_key))
        d = abs(s - g)
        results.append(_qchk("Grand Total — Impressions", "Sum(rows) = Grand Total",
                              f"{g:,.0f}", f"Sum={s:,.0f}  GT={g:,.0f}  diff={d:,.0f}",
                              "PASS" if d <= 1 else "FAIL"))
    else:
        results.append(_qchk("Grand Total — Impressions", "Sum = Grand Total", "Match",
                              "Grand Total row or Impressions column missing", "FAIL"))

    # 5. Grand Total – Clicks
    if gt and clk_key:
        s = sum(_qsafe(r.get(clk_key)) for r in rows)
        g = _qsafe(gt.get(clk_key))
        d = abs(s - g)
        results.append(_qchk("Grand Total — Clicks", "Sum(rows) = Grand Total",
                              f"{g:,.0f}", f"Sum={s:,.0f}  GT={g:,.0f}  diff={d:,.0f}",
                              "PASS" if d <= 1 else "FAIL"))

    # 6. Campaign date range (if provided)
    if campaign_start and campaign_end and valid:
        cs = ce = None
        for fmt in _FMTS:
            try: cs = _dt.datetime.strptime(str(campaign_start).strip(), fmt); break
            except: pass
        for fmt in _FMTS:
            try: ce = _dt.datetime.strptime(str(campaign_end).strip(), fmt); break
            except: pass
        if cs and ce:
            ok = valid[0] >= cs and valid[-1] <= ce
            results.append(_qchk(
                "Campaign Date Range",
                f"All dates within {campaign_start} – {campaign_end}",
                f"{campaign_start} to {campaign_end}",
                f"Report: {valid[0].strftime('%d %b %Y')} – {valid[-1].strftime('%d %b %Y')}",
                "PASS" if ok else "FAIL",
            ))
        else:
            results.append(_qchk("Campaign Date Range", "Dates within campaign range",
                                  "In range", "Could not parse provided campaign dates", "WARNING"))
    elif campaign_start or campaign_end:
        results.append(_qchk("Campaign Date Range", "Dates within campaign range",
                              "In range", "Both start and end dates required — skipped", "WARNING"))

    return results


def _run_campaign_qc(wb, is_video):
    """Viewability (banner/video) and VCR (video only) checks from the DATE sheet."""
    results = []
    headers, rows, gt = _qread(wb, "DATE")
    if headers is None:
        return [_qchk("DATE Sheet", "Sheet present", "Present", "Sheet not found", "FAIL")]

    hdr_lower = {h.lower(): h for h in headers}

    # Viewability
    view_key = hdr_lower.get("viewable impressions")
    meas_key = hdr_lower.get("measurable impressions")
    if view_key and meas_key:
        tv = sum(_qsafe(r.get(view_key)) for r in rows)
        tm = sum(_qsafe(r.get(meas_key)) for r in rows)
        vp = (tv / tm * 100) if tm else 0.0
        ok = 75.0 <= vp <= 86.0
        lbl = "PASS" if ok else ("WARNING" if (65 <= vp < 75 or 86 < vp <= 92) else "FAIL")
        results.append(_qchk(
            f"{'Video' if is_video else 'Banner'} Viewability",
            "Viewable Impressions / Measurable Impressions x 100",
            "75% – 86%",
            f"{vp:.2f}%  (Viewable={tv:,.0f} / Measurable={tm:,.0f})",
            lbl,
        ))
        # GT – Viewable
        if gt:
            g = _qsafe(gt.get(view_key)); d = abs(tv - g)
            results.append(_qchk("Grand Total — Viewable Impressions", "Sum = Grand Total",
                                  f"{g:,.0f}", f"Sum={tv:,.0f}  GT={g:,.0f}  diff={d:,.0f}",
                                  "PASS" if d <= 1 else "FAIL"))
            g = _qsafe(gt.get(meas_key)); d = abs(tm - g)
            results.append(_qchk("Grand Total — Measurable Impressions", "Sum = Grand Total",
                                  f"{g:,.0f}", f"Sum={tm:,.0f}  GT={g:,.0f}  diff={d:,.0f}",
                                  "PASS" if d <= 1 else "FAIL"))
    else:
        results.append(_qchk(
            f"{'Video' if is_video else 'Banner'} Viewability",
            "Viewable / Measurable Impressions x 100", "75% – 86%",
            "Viewable/Measurable Impressions columns not found in DATE sheet", "WARNING",
        ))

    if is_video:
        # VCR
        vcr_key    = next((v for k, v in hdr_lower.items() if "vcr" in k or "completion rate" in k), None)
        starts_key = next((v for k, v in hdr_lower.items() if "start" in k), None)
        comp_key   = next((v for k, v in hdr_lower.items() if "complete" in k), None)

        if vcr_key:
            vals = [_qpct(r.get(vcr_key)) for r in rows if r.get(vcr_key) is not None]
            if vals:
                avg  = sum(vals) / len(vals)
                ok   = 75.0 <= avg <= 86.0
                lbl  = "PASS" if ok else ("WARNING" if (65 <= avg < 75 or 86 < avg <= 92) else "FAIL")
                outs = [f"{v:.1f}%" for v in vals if not (75 <= v <= 86)]
                results.append(_qchk(
                    "VCR (Video Completion Rate)",
                    "Sum Complete Views / Sum Starts x 100", "75% – 86%",
                    f"Avg={avg:.2f}%"
                    + (f"  |  Outliers: {', '.join(outs[:5])}" if outs else "  |  All in range"),
                    lbl,
                ))
            else:
                results.append(_qchk("VCR", "75% – 86%", "In range", "No VCR values found", "WARNING"))
        elif starts_key and comp_key:
            ts = sum(_qsafe(r.get(starts_key)) for r in rows)
            tc = sum(_qsafe(r.get(comp_key)) for r in rows)
            vcr = (tc / ts * 100) if ts else 0.0
            ok  = 75.0 <= vcr <= 86.0
            lbl = "PASS" if ok else ("WARNING" if (65 <= vcr < 75 or 86 < vcr <= 92) else "FAIL")
            results.append(_qchk(
                "VCR (Video Completion Rate)", "Complete Views / Starts x 100", "75% – 86%",
                f"{vcr:.2f}%  (Starts={ts:,.0f}  Completes={tc:,.0f})", lbl,
            ))
        else:
            results.append(_qchk("VCR (Video Completion Rate)", "Complete / Starts x 100",
                                  "75% – 86%", "VCR / Starts / Completes columns not found", "WARNING"))

        # GT for video metrics
        if gt and starts_key and comp_key:
            for label, key in [("Video Starts", starts_key), ("Complete Views", comp_key)]:
                total = sum(_qsafe(r.get(key)) for r in rows)
                g     = _qsafe(gt.get(key)); d = abs(total - g)
                results.append(_qchk(
                    f"Grand Total — {label}", "Sum = Grand Total",
                    f"{g:,.0f}", f"Sum={total:,.0f}  GT={g:,.0f}  diff={d:,.0f}",
                    "PASS" if d <= 1 else "FAIL",
                ))

    return results


def _run_app_url_qc(wb):
    results = []
    headers, rows, gt = _qread(wb, "APP URL")
    if headers is None:
        return [_qchk("APP URL Sheet", "Sheet exists", "Present", "Sheet not found", "FAIL")]

    hdr_lower = {h.lower(): h for h in headers}
    url_key = next((v for k, v in hdr_lower.items() if "app" in k or "url" in k), None)
    imp_key = next((v for k, v in hdr_lower.items() if "impression" in k), None)
    clk_key = next((v for k, v in hdr_lower.items() if k == "clicks"), None)
    ctr_key = next((v for k, v in hdr_lower.items() if "click rate" in k or k == "ctr"), None)

    if not url_key or not imp_key:
        return [_qchk("APP URL Sheet", "Required columns present",
                       "App/URL + Impressions", f"Headers: {headers}", "FAIL")]

    imps = [_qsafe(r.get(imp_key)) for r in rows]
    urls = [str(r.get(url_key, "")).strip() for r in rows]

    # 1. Sorted by Impressions descending
    sorted_ok = all(imps[i] >= imps[i+1] for i in range(len(imps)-1))
    results.append(_qchk(
        "Impression Sorting", "Apps/URLs sorted highest to lowest impressions",
        "Descending order",
        "Correctly sorted" if sorted_ok else f"Out of order — sample: {imps[:4]}",
        "PASS" if sorted_ok else "FAIL",
    ))

    # 2. Duplicate URLs
    seen_u, dup_u = set(), []
    for u in urls:
        ul = u.lower()
        if ul and ul not in ("nan", "none"):
            if ul in seen_u: dup_u.append(u)
            else: seen_u.add(ul)
    results.append(_qchk(
        "Duplicate Site Check", "Each site/URL appears only once",
        "No duplicates",
        f"Duplicates: {dup_u[:5]}" if dup_u else f"No duplicates ({len(urls)} sites)",
        "FAIL" if dup_u else "PASS",
    ))

    # 3. Duplicate Impression Values
    imp_cnt = {}
    for v in imps:
        if v > 0: imp_cnt[v] = imp_cnt.get(v, 0) + 1
    dup_i = [v for v, c in imp_cnt.items() if c > 1]
    results.append(_qchk(
        "Duplicate Impression Values", "No two sites share the same impression count",
        "All unique",
        f"Duplicate values: {[f'{v:,.0f}' for v in dup_i[:5]]}" if dup_i else "No duplicates",
        "WARNING" if dup_i else "PASS",
    ))

    # 4. CTR Range Check (0.35% – 0.37%) with Impression adjustment hint
    _CTR_LO, _CTR_HI = 0.35, 0.37   # target CTR band

    if ctr_key and clk_key:
        ctr_vals = []
        for u, r in zip(urls, rows):
            ctr_v = _qpct(r.get(ctr_key))
            clk_v = _qsafe(r.get(clk_key))
            imp_v = _qsafe(r.get(imp_key))
            ctr_vals.append((u[:50], ctr_v, clk_v, imp_v))

        out_of_range = [(u, c, clk, imp) for u, c, clk, imp in ctr_vals
                        if c > 0 and not (_CTR_LO <= c <= _CTR_HI)]

        if out_of_range:
            # Build per-site suggestion: target impressions to hit 0.35%–0.37%
            suggestions = []
            for u, c, clk, imp in out_of_range[:5]:
                if clk > 0:
                    imp_for_hi  = int(clk / (_CTR_LO / 100))   # CTR=0.35% → max imps
                    imp_for_lo  = int(clk / (_CTR_HI / 100))   # CTR=0.37% → min imps
                    suggestions.append(
                        f"{u} (CTR={c:.3f}%, Clk={int(clk):,}) "
                        f"→ Adjust Impressions to {imp_for_lo:,}–{imp_for_hi:,}"
                    )
            actual_txt = (f"{len(out_of_range)} site(s) outside 0.35%–0.37%:  "
                          + "  |  ".join(suggestions))
        else:
            actual_txt = f"All {len(ctr_vals)} sites within 0.35%–0.37%"

        results.append(_qchk(
            "CTR Range Check",
            "Clicks / Impressions × 100 should be 0.35% – 0.37%",
            "0.35% – 0.37%",
            actual_txt,
            "WARNING" if out_of_range else "PASS",
        ))
    elif ctr_key:
        # No clicks column — just check CTR value alone
        ctr_vals_simple = [(u[:50], _qpct(r.get(ctr_key))) for u, r in zip(urls, rows)]
        out_of_range = [(u, c) for u, c in ctr_vals_simple if c > 0 and not (_CTR_LO <= c <= _CTR_HI)]
        results.append(_qchk(
            "CTR Range Check", "CTR should be 0.35% – 0.37%", "0.35% – 0.37%",
            (f"{len(out_of_range)} site(s) outside range: "
             + ", ".join(f"{u} ({c:.3f}%)" for u, c in out_of_range[:3]))
            if out_of_range else "All sites within 0.35%–0.37%",
            "WARNING" if out_of_range else "PASS",
        ))
    else:
        results.append(_qchk("CTR Range Check", "0.35% – 0.37%",
                              "0.35% – 0.37%", "CTR column not found", "WARNING"))

    # 6. Grand Total – Impressions
    if gt and imp_key:
        s = sum(imps); g = _qsafe(gt.get(imp_key)); d = abs(s - g)
        results.append(_qchk("Grand Total — Impressions", "Sum = Grand Total",
                              f"{g:,.0f}", f"Sum={s:,.0f}  GT={g:,.0f}  diff={d:,.0f}",
                              "PASS" if d <= 1 else "FAIL"))
    if gt and clk_key:
        s = sum(_qsafe(r.get(clk_key)) for r in rows)
        g = _qsafe(gt.get(clk_key)); d = abs(s - g)
        results.append(_qchk("Grand Total — Clicks", "Sum = Grand Total",
                              f"{g:,.0f}", f"Sum={s:,.0f}  GT={g:,.0f}  diff={d:,.0f}",
                              "PASS" if d <= 1 else "FAIL"))

    return results


def _run_exchange_qc(wb):
    results = []
    headers, rows, gt = _qread(wb, "EXCHANGE")
    if headers is None:
        return [_qchk("EXCHANGE Sheet", "Sheet exists", "Present", "Sheet not found", "FAIL")]

    hdr_lower  = {h.lower(): h for h in headers}
    exch_key   = next((v for k, v in hdr_lower.items() if "exchange" in k), None)
    imp_key    = next((v for k, v in hdr_lower.items() if "impression" in k), None)
    clk_key    = next((v for k, v in hdr_lower.items() if k == "clicks"), None)
    ctr_key    = next((v for k, v in hdr_lower.items() if "click rate" in k or k == "ctr"), None)

    if not exch_key or not imp_key:
        return [_qchk("EXCHANGE Sheet", "Required columns present",
                       "Exchange + Impressions", f"Headers: {headers}", "FAIL")]

    imps  = [_qsafe(r.get(imp_key))             for r in rows]
    names = [str(r.get(exch_key, "")).strip()    for r in rows]

    # ── 1. Google Ad Manager has highest impressions ───────────────────────────
    google_idx = next(
        (i for i, n in enumerate(names) if "google" in n.lower()),
        None,
    )
    if google_idx is not None:
        google_imp  = imps[google_idx]
        others_max  = max((v for i, v in enumerate(imps) if i != google_idx), default=0)
        google_ok   = google_imp >= others_max
        results.append(_qchk(
            "Google Ad Manager — Highest Impressions",
            "Google Ad Manager should have the highest impression count",
            "Google > all others",
            (f"Google={google_imp:,.0f}  next highest={others_max:,.0f}  ✓"
             if google_ok
             else f"Google={google_imp:,.0f}  next highest={others_max:,.0f}  ✗"),
            "PASS" if google_ok else "FAIL",
        ))
    else:
        results.append(_qchk(
            "Google Ad Manager — Highest Impressions",
            "Google Ad Manager should have the highest impression count",
            "Google > all others",
            "Google Ad Manager row not found in EXCHANGE sheet",
            "WARNING",
        ))

    # ── 2. No duplicate impression values ─────────────────────────────────────
    imp_cnt = {}
    for v in imps:
        if v > 0:
            imp_cnt[v] = imp_cnt.get(v, 0) + 1
    dup_imps = [v for v, c in imp_cnt.items() if c > 1]
    results.append(_qchk(
        "Duplicate Impression Values",
        "No two exchanges share the same impression count",
        "All unique",
        (f"Duplicate values: {[f'{v:,.0f}' for v in dup_imps[:5]]}"
         if dup_imps else f"No duplicates ({len(rows)} exchanges)"),
        "FAIL" if dup_imps else "PASS",
    ))

    # ── 3. CTR must not exceed 1% on any row ──────────────────────────────────
    if ctr_key:
        high_ctr = [
            (names[i], _qpct(r.get(ctr_key)))
            for i, r in enumerate(rows)
            if _qpct(r.get(ctr_key)) > 1.0
        ]
        results.append(_qchk(
            "CTR Threshold (> 1%)",
            "No exchange row CTR should exceed 1%",
            "CTR ≤ 1%",
            (f"{len(high_ctr)} exchange(s) above 1%: "
             + ", ".join(f"{n} ({c:.3f}%)" for n, c in high_ctr[:5]))
            if high_ctr else f"All {len(rows)} exchanges within threshold",
            "FAIL" if high_ctr else "PASS",
        ))
    else:
        results.append(_qchk("CTR Threshold (> 1%)", "CTR ≤ 1%",
                              "Within limit", "CTR column not found", "WARNING"))

    # ── 4. Grand Total — Impressions ──────────────────────────────────────────
    if gt and imp_key:
        s = sum(imps); g = _qsafe(gt.get(imp_key)); d = abs(s - g)
        results.append(_qchk("Grand Total — Impressions", "Sum = Grand Total",
                              f"{g:,.0f}", f"Sum={s:,.0f}  GT={g:,.0f}  diff={d:,.0f}",
                              "PASS" if d <= 1 else "FAIL"))

    # ── 5. Grand Total — Clicks ───────────────────────────────────────────────
    if gt and clk_key:
        s = sum(_qsafe(r.get(clk_key)) for r in rows)
        g = _qsafe(gt.get(clk_key)); d = abs(s - g)
        results.append(_qchk("Grand Total — Clicks", "Sum = Grand Total",
                              f"{g:,.0f}", f"Sum={s:,.0f}  GT={g:,.0f}  diff={d:,.0f}",
                              "PASS" if d <= 1 else "FAIL"))

    return results


def _run_device_qc(wb):
    results = []
    headers, rows, gt = _qread(wb, "DEVICE")
    if headers is None:
        return [_qchk("DEVICE Sheet", "Sheet exists", "Present", "Sheet not found", "FAIL")]

    hdr_lower  = {h.lower(): h for h in headers}
    dev_key    = next((v for k, v in hdr_lower.items() if "device" in k), None)
    imp_key    = next((v for k, v in hdr_lower.items() if "impression" in k), None)
    clk_key    = next((v for k, v in hdr_lower.items() if k == "clicks"), None)
    ctr_key    = next((v for k, v in hdr_lower.items() if "click rate" in k or k == "ctr"), None)

    if not dev_key or not imp_key:
        return [_qchk("DEVICE Sheet", "Required columns present",
                       "Device Type + Impressions", f"Headers: {headers}", "FAIL")]

    imps  = [_qsafe(r.get(imp_key))          for r in rows]
    names = [str(r.get(dev_key, "")).strip()  for r in rows]

    # ── 1. Desktop or Smart Phone has highest impressions (not Tablet) ─────────
    if imps:
        top_idx  = imps.index(max(imps))
        top_name = names[top_idx]
        top_ok   = any(k in top_name.lower() for k in ("desktop", "smart", "phone", "mobile"))
        results.append(_qchk(
            "Top Device — Highest Impressions",
            "Desktop or Smart Phone should have the highest impression count (not Tablet)",
            "Desktop or Smart Phone",
            f"{top_name} = {imps[top_idx]:,.0f}  {'✓' if top_ok else '✗'}",
            "PASS" if top_ok else "FAIL",
        ))

    # ── 2. No duplicate impression values ─────────────────────────────────────
    imp_cnt  = {}
    for v in imps:
        if v > 0:
            imp_cnt[v] = imp_cnt.get(v, 0) + 1
    dup_imps = [v for v, c in imp_cnt.items() if c > 1]
    results.append(_qchk(
        "Duplicate Impression Values",
        "No two device types share the same impression count",
        "All unique",
        (f"Duplicate values: {[f'{v:,.0f}' for v in dup_imps[:5]]}"
         if dup_imps else f"No duplicates ({len(rows)} devices)"),
        "FAIL" if dup_imps else "PASS",
    ))

    # ── 3. CTR must not exceed 1% on any row ──────────────────────────────────
    if ctr_key:
        high_ctr = [
            (names[i], _qpct(r.get(ctr_key)))
            for i, r in enumerate(rows)
            if _qpct(r.get(ctr_key)) > 1.0
        ]
        results.append(_qchk(
            "CTR Threshold (> 1%)",
            "No device type CTR should exceed 1%",
            "CTR ≤ 1%",
            (f"{len(high_ctr)} device(s) above 1%: "
             + ", ".join(f"{n} ({c:.3f}%)" for n, c in high_ctr[:5]))
            if high_ctr else f"All {len(rows)} devices within threshold",
            "FAIL" if high_ctr else "PASS",
        ))
    else:
        results.append(_qchk("CTR Threshold (> 1%)", "CTR ≤ 1%",
                              "Within limit", "CTR column not found", "WARNING"))

    # ── 4. Grand Total — Impressions ──────────────────────────────────────────
    if gt and imp_key:
        s = sum(imps); g = _qsafe(gt.get(imp_key)); d = abs(s - g)
        results.append(_qchk("Grand Total — Impressions", "Sum = Grand Total",
                              f"{g:,.0f}", f"Sum={s:,.0f}  GT={g:,.0f}  diff={d:,.0f}",
                              "PASS" if d <= 1 else "FAIL"))

    # ── 5. Grand Total — Clicks ───────────────────────────────────────────────
    if gt and clk_key:
        s = sum(_qsafe(r.get(clk_key)) for r in rows)
        g = _qsafe(gt.get(clk_key)); d = abs(s - g)
        results.append(_qchk("Grand Total — Clicks", "Sum = Grand Total",
                              f"{g:,.0f}", f"Sum={s:,.0f}  GT={g:,.0f}  diff={d:,.0f}",
                              "PASS" if d <= 1 else "FAIL"))

    return results


def _run_city_qc(wb):
    results = []
    headers, rows, gt = _qread(wb, "CITY")
    if headers is None:
        return [_qchk("CITY Sheet", "Sheet exists", "Present", "Sheet not found", "FAIL")]

    hdr_lower = {h.lower(): h for h in headers}
    city_key  = next((v for k, v in hdr_lower.items() if "city" in k), None)
    imp_key   = next((v for k, v in hdr_lower.items() if "impression" in k), None)
    clk_key   = next((v for k, v in hdr_lower.items() if k == "clicks"), None)
    ctr_key   = next((v for k, v in hdr_lower.items() if "click rate" in k or k == "ctr"), None)

    if not city_key or not imp_key:
        return [_qchk("CITY Sheet", "Required columns present",
                       "City + Impressions", f"Headers: {headers}", "FAIL")]

    imps  = [_qsafe(r.get(imp_key))         for r in rows]
    names = [str(r.get(city_key, "")).strip() for r in rows]

    # ── 1. No "Unknown" city names ────────────────────────────────────────────
    _INVALID = {"unknown", "nan", "none", "", "n/a", "na", "other", "others"}
    bad_names = [n for n in names if n.lower() in _INVALID]
    results.append(_qchk(
        "No Unknown City Names",
        "City names must be real — no 'Unknown', blank, or placeholder values",
        "All valid city names",
        (f"Invalid names found: {bad_names[:5]}"
         if bad_names else f"All {len(rows)} city names are valid"),
        "FAIL" if bad_names else "PASS",
    ))

    # ── 2. Cities sorted A → Z ────────────────────────────────────────────────
    valid_names = [n for n in names if n.lower() not in _INVALID]
    az_ok = all(
        valid_names[i].lower() <= valid_names[i + 1].lower()
        for i in range(len(valid_names) - 1)
    )
    results.append(_qchk(
        "City Ordering (A → Z)",
        "Cities should be sorted alphabetically A to Z",
        "Ascending alphabetical",
        "Correctly sorted A → Z" if az_ok
        else f"Out of order — e.g. '{valid_names[next((i for i in range(len(valid_names)-1) if valid_names[i].lower() > valid_names[i+1].lower()), 0)]}' before '{valid_names[next((i for i in range(len(valid_names)-1) if valid_names[i].lower() > valid_names[i+1].lower()), 0)+1]}'",
        "PASS" if az_ok else "FAIL",
    ))

    # ── 3. CTR must not exceed 1% on any row ──────────────────────────────────
    if ctr_key:
        high_ctr = [
            (names[i], _qpct(r.get(ctr_key)))
            for i, r in enumerate(rows)
            if _qpct(r.get(ctr_key)) > 1.0
        ]
        results.append(_qchk(
            "CTR Threshold (> 1%)",
            "No city CTR should exceed 1%",
            "CTR ≤ 1%",
            (f"{len(high_ctr)} city/cities above 1%: "
             + ", ".join(f"{n} ({c:.3f}%)" for n, c in high_ctr[:5]))
            if high_ctr else f"All {len(rows)} cities within threshold",
            "FAIL" if high_ctr else "PASS",
        ))
    else:
        results.append(_qchk("CTR Threshold (> 1%)", "CTR ≤ 1%",
                              "Within limit", "CTR column not found", "WARNING"))

    # ── 4. Grand Total — Impressions ──────────────────────────────────────────
    if gt and imp_key:
        s = sum(imps); g = _qsafe(gt.get(imp_key)); d = abs(s - g)
        results.append(_qchk("Grand Total — Impressions", "Sum = Grand Total",
                              f"{g:,.0f}", f"Sum={s:,.0f}  GT={g:,.0f}  diff={d:,.0f}",
                              "PASS" if d <= 1 else "FAIL"))

    # ── 5. Grand Total — Clicks ───────────────────────────────────────────────
    if gt and clk_key:
        s = sum(_qsafe(r.get(clk_key)) for r in rows)
        g = _qsafe(gt.get(clk_key)); d = abs(s - g)
        results.append(_qchk("Grand Total — Clicks", "Sum = Grand Total",
                              f"{g:,.0f}", f"Sum={s:,.0f}  GT={g:,.0f}  diff={d:,.0f}",
                              "PASS" if d <= 1 else "FAIL"))

    return results


def _run_creative_qc(wb):
    results = []
    headers, rows, gt = _qread(wb, "CREATIVE")
    if headers is None:
        return [_qchk("CREATIVE Sheet", "Sheet exists", "Present", "Sheet not found", "FAIL")]

    hdr_lower    = {h.lower(): h for h in headers}
    creative_key = next((v for k, v in hdr_lower.items() if "creative" in k), None)
    imp_key      = next((v for k, v in hdr_lower.items() if "impression" in k), None)
    clk_key      = next((v for k, v in hdr_lower.items() if k == "clicks"), None)
    ctr_key      = next((v for k, v in hdr_lower.items() if "click rate" in k or k == "ctr"), None)

    if not creative_key or not imp_key:
        return [_qchk("CREATIVE Sheet", "Required columns present",
                       "Creative + Impressions", f"Headers: {headers}", "FAIL")]

    imps  = [_qsafe(r.get(imp_key))                for r in rows]
    clks  = [_qsafe(r.get(clk_key, 0))             for r in rows] if clk_key else [0] * len(rows)
    names = [str(r.get(creative_key, "")).strip()   for r in rows]

    # ── 1. No duplicate impression values ─────────────────────────────────────
    imp_cnt  = {}
    for v in imps:
        if v > 0:
            imp_cnt[v] = imp_cnt.get(v, 0) + 1
    dup_imps = [v for v, c in imp_cnt.items() if c > 1]
    results.append(_qchk(
        "Duplicate Impression Values",
        "No two creatives share the same impression count",
        "All unique",
        (f"Duplicate values: {[f'{v:,.0f}' for v in dup_imps[:5]]}"
         if dup_imps else f"No duplicates ({len(rows)} creatives)"),
        "FAIL" if dup_imps else "PASS",
    ))

    # ── 2. CTR must not exceed 1% on any row ──────────────────────────────────
    if ctr_key:
        high_ctr = [
            (names[i][:50], _qpct(r.get(ctr_key)))
            for i, r in enumerate(rows)
            if _qpct(r.get(ctr_key)) > 1.0
        ]
        results.append(_qchk(
            "CTR Threshold (> 1%)",
            "No creative CTR should exceed 1%",
            "CTR ≤ 1%",
            (f"{len(high_ctr)} creative(s) above 1%: "
             + ", ".join(f"{n} ({c:.3f}%)" for n, c in high_ctr[:5]))
            if high_ctr else f"All {len(rows)} creatives within threshold",
            "FAIL" if high_ctr else "PASS",
        ))
    else:
        results.append(_qchk("CTR Threshold (> 1%)", "CTR ≤ 1%",
                              "Within limit", "CTR column not found", "WARNING"))

    # ── 3. Grand Total — Impressions ──────────────────────────────────────────
    if gt and imp_key:
        s = sum(imps); g = _qsafe(gt.get(imp_key)); d = abs(s - g)
        results.append(_qchk("Grand Total — Impressions", "Sum = Grand Total",
                              f"{g:,.0f}", f"Sum={s:,.0f}  GT={g:,.0f}  diff={d:,.0f}",
                              "PASS" if d <= 1 else "FAIL"))

    # ── 4. Grand Total — Clicks ───────────────────────────────────────────────
    if gt and clk_key:
        s = sum(clks); g = _qsafe(gt.get(clk_key)); d = abs(s - g)
        results.append(_qchk("Grand Total — Clicks", "Sum = Grand Total",
                              f"{g:,.0f}", f"Sum={s:,.0f}  GT={g:,.0f}  diff={d:,.0f}",
                              "PASS" if d <= 1 else "FAIL"))

    # ── 5–7. Tier CTR-weight checks ───────────────────────────────────────────
    _TIER1 = {"250x250","300x250","300x600","160x600","728x90",
               "970x90","970x250","120x600"}
    _TIER2 = {"300x50","320x50","320x100","336x280","320x480"}
    _TIER3 = {"320x250","480x320","768x102","1024x768"}

    def _tier_of(name: str):
        nl = name.lower()
        for sz in _TIER1:
            if sz in nl: return 1
        for sz in _TIER2:
            if sz in nl: return 2
        for sz in _TIER3:
            if sz in nl: return 3
        return 0   # unclassified

    tiers = [_tier_of(n) for n in names]
    any_classified = any(t > 0 for t in tiers)

    if not any_classified:
        # Creative names don't contain size strings — skip tier checks
        results.append(_qchk(
            "Tier Distribution (Clicks)",
            "Tier 1 45-55% | Tier 2 30-40% | Tier 3 ~20%",
            "Size dimensions in creative names",
            "Cannot determine tier — no size strings (e.g. '300x250') found in creative names",
            "WARNING",
        ))
    else:
        total_clk = sum(clks) or 1
        t1_clk = sum(clks[i] for i, t in enumerate(tiers) if t == 1)
        t2_clk = sum(clks[i] for i, t in enumerate(tiers) if t == 2)
        t3_clk = sum(clks[i] for i, t in enumerate(tiers) if t == 3)

        t1_pct = t1_clk / total_clk * 100
        t2_pct = t2_clk / total_clk * 100
        t3_pct = t3_clk / total_clk * 100

        # Tier 1 — 45% to 55%
        t1_ok = 45.0 <= t1_pct <= 55.0
        results.append(_qchk(
            "Tier 1 — Click Weight (45%–55%)",
            "Tier 1 sizes (300x250, 728x90, 970x250 …) should drive 45–55% of clicks",
            "45% – 55%",
            f"{t1_pct:.1f}%  ({t1_clk:,.0f} clicks  |  {sum(1 for t in tiers if t==1)} creatives)",
            "PASS" if t1_ok else "FAIL",
        ))

        # Tier 2 — 30% to 40%
        t2_ok = 30.0 <= t2_pct <= 40.0
        results.append(_qchk(
            "Tier 2 — Click Weight (30%–40%)",
            "Tier 2 sizes (320x50, 300x50, 320x480 …) should drive 30–40% of clicks",
            "30% – 40%",
            f"{t2_pct:.1f}%  ({t2_clk:,.0f} clicks  |  {sum(1 for t in tiers if t==2)} creatives)",
            "PASS" if t2_ok else "FAIL",
        ))

        # Tier 3 — ~20% (allow 15–25%)
        t3_ok = 15.0 <= t3_pct <= 25.0
        results.append(_qchk(
            "Tier 3 — Click Weight (~20%)",
            "Tier 3 sizes (320x250, 480x320, 1024x768 …) should drive ~20% of clicks",
            "15% – 25%",
            f"{t3_pct:.1f}%  ({t3_clk:,.0f} clicks  |  {sum(1 for t in tiers if t==3)} creatives)",
            "PASS" if t3_ok else "FAIL",
        ))

        # Tier 3 — no zero-click rows
        t3_zero = [names[i][:50] for i, t in enumerate(tiers) if t == 3 and clks[i] == 0]
        results.append(_qchk(
            "Tier 3 — No Zero-Click Rows",
            "Every Tier 3 creative must have at least 1 click",
            "Clicks > 0 for all Tier 3",
            (f"{len(t3_zero)} Tier 3 creative(s) with 0 clicks: "
             + ", ".join(t3_zero[:5]))
            if t3_zero else f"All Tier 3 creatives have clicks",
            "FAIL" if t3_zero else "PASS",
        ))

    return results


def _run_age_qc(wb):
    results = []
    headers, rows, gt = _qread(wb, "AGE")
    if headers is None:
        return [_qchk("AGE Sheet", "Sheet exists", "Present", "Sheet not found", "FAIL")]

    hdr_lower = {h.lower(): h for h in headers}
    age_key   = next((v for k, v in hdr_lower.items() if "age" in k), None)
    imp_key   = next((v for k, v in hdr_lower.items() if "impression" in k), None)
    clk_key   = next((v for k, v in hdr_lower.items() if k == "clicks"), None)
    ctr_key   = next((v for k, v in hdr_lower.items() if "click rate" in k or k == "ctr"), None)

    if not age_key or not imp_key:
        return [_qchk("AGE Sheet", "Required columns present",
                       "Age + Impressions", f"Headers: {headers}", "FAIL")]

    imps  = [_qsafe(r.get(imp_key))        for r in rows]
    names = [str(r.get(age_key, "")).strip() for r in rows]
    total_imp = sum(imps) or 1

    # ── 1. No duplicate impression values ─────────────────────────────────────
    imp_cnt  = {}
    for v in imps:
        if v > 0:
            imp_cnt[v] = imp_cnt.get(v, 0) + 1
    dup_imps = [v for v, c in imp_cnt.items() if c > 1]
    results.append(_qchk(
        "Duplicate Impression Values",
        "No two age groups share the same impression count",
        "All unique",
        (f"Duplicate values: {[f'{v:,.0f}' for v in dup_imps[:5]]}"
         if dup_imps else f"No duplicates ({len(rows)} age groups)"),
        "FAIL" if dup_imps else "PASS",
    ))

    # ── 2. CTR must not exceed 1% on any row ──────────────────────────────────
    if ctr_key:
        high_ctr = [
            (names[i], _qpct(r.get(ctr_key)))
            for i, r in enumerate(rows)
            if _qpct(r.get(ctr_key)) > 1.0
        ]
        results.append(_qchk(
            "CTR Threshold (> 1%)",
            "No age group CTR should exceed 1%",
            "CTR ≤ 1%",
            (f"{len(high_ctr)} age group(s) above 1%: "
             + ", ".join(f"{n} ({c:.3f}%)" for n, c in high_ctr[:5]))
            if high_ctr else f"All {len(rows)} age groups within threshold",
            "FAIL" if high_ctr else "PASS",
        ))
    else:
        results.append(_qchk("CTR Threshold (> 1%)", "CTR ≤ 1%",
                              "Within limit", "CTR column not found", "WARNING"))

    # ── 3. Grand Total — Impressions ──────────────────────────────────────────
    if gt and imp_key:
        s = sum(imps); g = _qsafe(gt.get(imp_key)); d = abs(s - g)
        results.append(_qchk("Grand Total — Impressions", "Sum = Grand Total",
                              f"{g:,.0f}", f"Sum={s:,.0f}  GT={g:,.0f}  diff={d:,.0f}",
                              "PASS" if d <= 1 else "FAIL"))

    # ── 4. Grand Total — Clicks ───────────────────────────────────────────────
    if gt and clk_key:
        s = sum(_qsafe(r.get(clk_key)) for r in rows)
        g = _qsafe(gt.get(clk_key)); d = abs(s - g)
        results.append(_qchk("Grand Total — Clicks", "Sum = Grand Total",
                              f"{g:,.0f}", f"Sum={s:,.0f}  GT={g:,.0f}  diff={d:,.0f}",
                              "PASS" if d <= 1 else "FAIL"))

    # ── 5. Unique impression percentage per age group ─────────────────────────
    # Each age group must have a distinct % share of total impressions.
    # Also surfaces the actual breakdown so reviewers can spot skewed distributions.
    pct_map   = {names[i]: round(imps[i] / total_imp * 100, 2) for i in range(len(rows))}
    pct_vals  = list(pct_map.values())
    pct_seen  = {}
    dup_pcts  = []
    for name, pv in pct_map.items():
        pv_r = round(pv, 1)          # 1-decimal bucket for near-duplicate detection
        if pv_r in pct_seen:
            dup_pcts.append(f"{name} & {pct_seen[pv_r]} both ≈ {pv_r}%")
        else:
            pct_seen[pv_r] = name

    breakdown = "  |  ".join(f"{n}: {p:.1f}%" for n, p in pct_map.items())
    results.append(_qchk(
        "Unique Impression % per Age Group",
        "Each age group should have a distinct impression share — no two groups at the same %",
        "All percentages unique",
        (f"Duplicate %: {dup_pcts[:3]}  |  Breakdown: {breakdown}"
         if dup_pcts else f"All unique  |  {breakdown}"),
        "FAIL" if dup_pcts else "PASS",
    ))

    return results


def _run_age_gender_qc(wb, sheet_name: str):
    """
    Shared QC logic for AGE and GENDER sheets.
    Checks: duplicate impressions, CTR ≤ 1%, grand total, unique impression %.
    """
    results = []
    headers, rows, gt = _qread(wb, sheet_name)
    if headers is None:
        return [_qchk(f"{sheet_name} Sheet", "Sheet exists", "Present",
                       "Sheet not found", "FAIL")]

    hdr_lower   = {h.lower(): h for h in headers}
    label_key   = headers[0] if headers else sheet_name
    imp_key     = next((v for k, v in hdr_lower.items() if "impression" in k), None)
    clk_key     = next((v for k, v in hdr_lower.items() if k == "clicks"), None)
    ctr_key     = next((v for k, v in hdr_lower.items() if "click rate" in k or k == "ctr"), None)

    if not imp_key:
        return [_qchk(f"{sheet_name} Sheet", "Impressions column present",
                       "Found", "Impressions column not found", "FAIL")]

    imps  = [_qsafe(r.get(imp_key))           for r in rows]
    clks  = [_qsafe(r.get(clk_key, 0))        for r in rows] if clk_key else [0] * len(rows)
    names = [str(r.get(label_key, "")).strip() for r in rows]
    total_imp = sum(imps) or 1

    # ── 1. No duplicate impression values ─────────────────────────────────────
    imp_cnt  = {}
    for v in imps:
        if v > 0:
            imp_cnt[v] = imp_cnt.get(v, 0) + 1
    dup_imps = [v for v, c in imp_cnt.items() if c > 1]
    results.append(_qchk(
        "Duplicate Impression Values",
        f"No two {sheet_name.lower()} groups share the same impression count",
        "All unique",
        (f"Duplicate values: {[f'{v:,.0f}' for v in dup_imps[:5]]}"
         if dup_imps else f"No duplicates ({len(rows)} groups)"),
        "FAIL" if dup_imps else "PASS",
    ))

    # ── 2. CTR must not exceed 1% on any row ──────────────────────────────────
    if ctr_key:
        high_ctr = [
            (names[i], _qpct(r.get(ctr_key)))
            for i, r in enumerate(rows)
            if _qpct(r.get(ctr_key)) > 1.0
        ]
        results.append(_qchk(
            "CTR Threshold (> 1%)",
            f"No {sheet_name.lower()} group CTR should exceed 1%",
            "CTR ≤ 1%",
            (f"{len(high_ctr)} group(s) above 1%: "
             + ", ".join(f"{n} ({c:.3f}%)" for n, c in high_ctr[:5]))
            if high_ctr else f"All {len(rows)} groups within threshold",
            "FAIL" if high_ctr else "PASS",
        ))
    else:
        results.append(_qchk("CTR Threshold (> 1%)", "CTR ≤ 1%",
                              "Within limit", "CTR column not found", "WARNING"))

    # ── 3. Grand Total — Impressions ──────────────────────────────────────────
    if gt and imp_key:
        s = sum(imps); g = _qsafe(gt.get(imp_key)); d = abs(s - g)
        results.append(_qchk("Grand Total — Impressions", "Sum = Grand Total",
                              f"{g:,.0f}", f"Sum={s:,.0f}  GT={g:,.0f}  diff={d:,.0f}",
                              "PASS" if d <= 1 else "FAIL"))

    # ── 4. Grand Total — Clicks ───────────────────────────────────────────────
    if gt and clk_key:
        s = sum(clks); g = _qsafe(gt.get(clk_key)); d = abs(s - g)
        results.append(_qchk("Grand Total — Clicks", "Sum = Grand Total",
                              f"{g:,.0f}", f"Sum={s:,.0f}  GT={g:,.0f}  diff={d:,.0f}",
                              "PASS" if d <= 1 else "FAIL"))

    # ── 5. Unique impression percentage per group ──────────────────────────────
    # Each group's % share of total impressions must be unique
    pct_vals   = [round(v / total_imp * 100, 2) for v in imps]
    pct_cnt    = {}
    for p in pct_vals:
        pct_cnt[p] = pct_cnt.get(p, 0) + 1
    dup_pcts   = [p for p, c in pct_cnt.items() if c > 1]

    pct_summary = "  |  ".join(
        f"{names[i]} = {pct_vals[i]:.2f}%" for i in range(len(rows))
    )
    results.append(_qchk(
        "Unique Impression % per Group",
        f"Each {sheet_name.lower()} group must have a distinct % share of total impressions",
        "All percentages unique",
        (f"Duplicate %: {dup_pcts}  ({pct_summary})"
         if dup_pcts else pct_summary),
        "FAIL" if dup_pcts else "PASS",
    ))

    return results


def _run_time_of_day_qc(wb):
    results = []
    headers, rows, gt = _qread(wb, "TIME OF DAY")
    if headers is None:
        return [_qchk("TIME OF DAY Sheet", "Sheet exists", "Present", "Sheet not found", "FAIL")]

    hdr_lower = {h.lower(): h for h in headers}
    imp_key = next((v for k, v in hdr_lower.items() if "impression" in k), None)
    clk_key = next((v for k, v in hdr_lower.items() if k == "clicks"), None)
    ctr_key = next((v for k, v in hdr_lower.items() if "click rate" in k or k == "ctr"), None)
    time_key = headers[0] if headers else "Time of Day"

    imps = [_qsafe(r.get(imp_key)) for r in rows] if imp_key else []

    # 1. Grand Total – Impressions
    if gt and imp_key:
        s = sum(imps); g = _qsafe(gt.get(imp_key)); d = abs(s - g)
        results.append(_qchk("Grand Total — Impressions", "Sum(hour rows) = Grand Total",
                              f"{g:,.0f}", f"Sum={s:,.0f}  GT={g:,.0f}  diff={d:,.0f}",
                              "PASS" if d <= 1 else "FAIL"))
    else:
        results.append(_qchk("Grand Total — Impressions", "Sum = Grand Total", "Match",
                              "Grand Total row or Impressions column missing", "FAIL"))

    # 2. Grand Total – Clicks
    if gt and clk_key:
        s = sum(_qsafe(r.get(clk_key)) for r in rows)
        g = _qsafe(gt.get(clk_key)); d = abs(s - g)
        results.append(_qchk("Grand Total — Clicks", "Sum(hour rows) = Grand Total",
                              f"{g:,.0f}", f"Sum={s:,.0f}  GT={g:,.0f}  diff={d:,.0f}",
                              "PASS" if d <= 1 else "FAIL"))

    # 3. Duplicate Impression Values
    cnt = {}
    for v in imps:
        if v > 0: cnt[v] = cnt.get(v, 0) + 1
    dups = [v for v, c in cnt.items() if c > 1]
    results.append(_qchk(
        "Duplicate Impression Values", "No two time slots share identical impression counts",
        "All unique",
        f"Duplicate values: {[f'{v:,.0f}' for v in dups[:5]]}" if dups else f"No duplicates ({len(rows)} slots)",
        "FAIL" if dups else "PASS",
    ))

    # 4. CTR > 1%
    if ctr_key:
        ctr_vals = [(str(r.get(time_key, "")), _qpct(r.get(ctr_key))) for r in rows]
        high = [(slot, c) for slot, c in ctr_vals if c > 1.0]
        results.append(_qchk(
            "CTR Threshold (> 1%)", "No time slot CTR should exceed 1%",
            "CTR <= 1%",
            (f"{len(high)} slot(s) above 1%: "
             + ", ".join(f"{s} ({c:.3f}%)" for s, c in high[:5]))
            if high else "All within threshold",
            "FAIL" if high else "PASS",
        ))
    else:
        results.append(_qchk("CTR Threshold (> 1%)", "CTR <= 1%", "Within limit",
                              "CTR column not found", "WARNING"))

    return results


# ─── Orchestrator ─────────────────────────────────────────────────────────────

def _run_all_qc(file_bytes, impressions, clicks, ctr_raw, reach, frequency,
                campaign_start=None, campaign_end=None):
    """
    Run QC across all sheets. Returns dict keyed by sheet group name.
    Auto-detects Banner vs Video from DATE sheet headers.
    """
    import openpyxl as _xl
    wb = _xl.load_workbook(io.BytesIO(file_bytes), read_only=False, data_only=True)

    results = {}

    # REACH
    results["REACH"] = _run_reach_qc(impressions, clicks, ctr_raw, reach, frequency)

    # DATE
    results["DATE"] = _run_date_qc(wb, campaign_start, campaign_end)

    # Detect Banner vs Video from DATE sheet headers
    is_video = False
    if "DATE" in wb.sheetnames:
        _ws = wb["DATE"]
        _hdr_row = next(_ws.iter_rows(values_only=True), [])
        _hdrs_lower = [str(h).lower() for h in _hdr_row if h]
        if any("video" in h or "vcr" in h or "completion" in h for h in _hdrs_lower):
            is_video = True

    camp_label = "VIDEO" if is_video else "BANNER"
    results[camp_label] = _run_campaign_qc(wb, is_video=is_video)

    # APP URL
    results["APP URL"] = _run_app_url_qc(wb)

    # EXCHANGE
    results["EXCHANGE"] = _run_exchange_qc(wb)

    # DEVICE
    results["DEVICE"] = _run_device_qc(wb)

    # CITY
    results["CITY"] = _run_city_qc(wb)

    # CREATIVE
    results["CREATIVE"] = _run_creative_qc(wb)

    # AGE
    results["AGE"] = _run_age_gender_qc(wb, "AGE")

    # GENDER
    results["GENDER"] = _run_age_gender_qc(wb, "GENDER")

    # TIME OF DAY
    results["TIME OF DAY"] = _run_time_of_day_qc(wb)

    # Cross-sheet Grand Total consistency
    # impressions / clicks come from REACH sheet (passed in by caller)
    results["CROSS-SHEET GT"] = _run_cross_sheet_gt_qc(wb, impressions, clicks)

    wb.close()
    return results


# ─── Download QC report ───────────────────────────────────────────────────────

@router.get("/reports/{report_id}/qc-report/download")
def download_qc_report(report_id: int):
    """Generate and download a colour-coded QC Excel report."""
    import json as _json
    try:
        conn = _get_ctr_conn()
        cur  = conn.cursor()
        cur.execute(
            """SELECT campaign_name, report_filename, qc_status,
                      qc_impressions, qc_clicks, qc_ctr, qc_submitted_at, qc_results
               FROM   final_report_store WHERE id = %s""",
            (report_id,),
        )
        row = cur.fetchone()
        cur.close(); conn.close()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    if not row:
        raise HTTPException(status_code=404, detail=f"Report {report_id} not found")

    (campaign_name, report_filename, qc_status,
     qc_impressions, qc_clicks, qc_ctr, qc_submitted_at, qc_results_raw) = row

    qc_data = {}
    if qc_results_raw:
        try:
            parsed = _json.loads(qc_results_raw)
            if isinstance(parsed, dict):
                qc_data = parsed
            elif isinstance(parsed, list):
                qc_data = {"REACH": parsed}
        except Exception:
            pass

    from openpyxl import Workbook as _WB
    from openpyxl.styles import PatternFill as _PF, Font as _Ft, Alignment as _Al, Border as _Bd, Side as _Sd
    wb = _WB(); ws = wb.active; ws.title = "QC Results"

    _BLUE = "00B0F0"; _GREEN = "00B050"; _RED = "FF0000"
    _LGREY = "F2F2F2"; _WARN_BG = "FFF2CC"; _CORR_BG = "EBF0FF"
    _thin = _Bd(
        left=_Sd(border_style="thin", color="BFBFBF"),
        right=_Sd(border_style="thin", color="BFBFBF"),
        top=_Sd(border_style="thin", color="BFBFBF"),
        bottom=_Sd(border_style="thin", color="BFBFBF"),
    )

    def _cell(r, c, val, bold=False, bg=None, fg="000000", align="left", size=10):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font      = _Ft(bold=bold, color=fg, size=size, name="Calibri")
        cell.alignment = _Al(horizontal=align, vertical="center", wrap_text=True)
        cell.border    = _thin
        if bg: cell.fill = _PF("solid", fgColor=bg)
        return cell

    ws.merge_cells("A1:E1")
    t = ws.cell(row=1, column=1, value="QC CHECK REPORT")
    t.font      = _Ft(bold=True, size=13, name="Calibri", color="FFFFFF")
    t.fill      = _PF("solid", fgColor=_BLUE)
    t.alignment = _Al(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    status_label = {"approved": "APPROVED", "warning": "WARNING", "rejected": "REJECTED"}.get(qc_status or "", (qc_status or "").upper())
    meta = [
        ("Campaign",     campaign_name or report_filename),
        ("File",         report_filename),
        ("QC Status",    status_label),
        ("Impressions",  f"{qc_impressions:,}" if qc_impressions else "—"),
        ("Clicks",       f"{qc_clicks:,}"      if qc_clicks      else "—"),
        ("CTR",          qc_ctr or "—"),
        ("Submitted At", qc_submitted_at.strftime("%d %b %Y %H:%M UTC") if qc_submitted_at else "—"),
    ]
    for i, (lbl, val) in enumerate(meta, start=2):
        _cell(i, 1, lbl, bold=True, bg=_LGREY)
        ws.merge_cells(f"B{i}:E{i}")
        _cell(i, 2, val)
        ws.row_dimensions[i].height = 16

    cur_row = 2 + len(meta) + 2

    _STATUS_BG  = {"PASS": "E2EFDA", "FAIL": "FFE0E0", "WARNING": _WARN_BG, "CORRECTED": _CORR_BG}
    _STATUS_FG  = {"PASS": _GREEN,   "FAIL": _RED,      "WARNING": "FF9900",  "CORRECTED": "4472C4"}
    _STATUS_GRP = {"PASS": "C6EFCE", "FAIL": "FFC7CE",  "WARNING": "FFEB9C",  "CORRECTED": "D6E0FF"}

    for sheet_name, checks in qc_data.items():
        ws.merge_cells(f"A{cur_row}:E{cur_row}")
        h = ws.cell(row=cur_row, column=1, value=f"  {sheet_name} SHEET")
        h.font      = _Ft(bold=True, size=10, color="FFFFFF", name="Calibri")
        h.fill      = _PF("solid", fgColor="4472C4")
        h.alignment = _Al(horizontal="left", vertical="center")
        h.border    = _thin
        ws.row_dimensions[cur_row].height = 18
        cur_row += 1

        for c_idx, hdr in enumerate(["Check", "Formula", "Expected", "Actual", "Result"], 1):
            _cell(cur_row, c_idx, hdr, bold=True, bg="D9E2F3", align="center")
        ws.row_dimensions[cur_row].height = 16
        cur_row += 1

        for chk in checks:
            st = chk.get("status", "FAIL")
            bg = _STATUS_BG.get(st, "FFFFFF")
            fg = _STATUS_FG.get(st, _RED)
            _cell(cur_row, 1, chk.get("check", ""),    bg=bg)
            _cell(cur_row, 2, chk.get("formula", ""),  bg=bg)
            _cell(cur_row, 3, chk.get("expected", ""), bg=bg, align="center")
            _cell(cur_row, 4, chk.get("actual", ""),   bg=bg)
            _cell(cur_row, 5, st, bold=True, bg=bg, fg=fg, align="center")
            ws.row_dimensions[cur_row].height = 20
            cur_row += 1

        pass_n = sum(1 for c in checks if c.get("status") == "PASS")
        fail_n = sum(1 for c in checks if c.get("status") == "FAIL")
        warn_n = sum(1 for c in checks if c.get("status") == "WARNING")
        corr_n = sum(1 for c in checks if c.get("status") == "CORRECTED")
        grp_st = "FAIL" if fail_n else "WARNING" if warn_n else "CORRECTED" if corr_n else "PASS"
        ws.merge_cells(f"A{cur_row}:D{cur_row}")
        _cell(cur_row, 1,
              f"  {sheet_name} — {pass_n} Pass · {fail_n} Fail · {warn_n} Warning · {corr_n} Corrected",
              bold=True, bg=_STATUS_GRP.get(grp_st, "FFFFFF"),
              fg=_STATUS_FG.get(grp_st, _RED), align="left")
        _cell(cur_row, 5, grp_st, bold=True,
              bg=_STATUS_GRP.get(grp_st, "FFFFFF"),
              fg=_STATUS_FG.get(grp_st, _RED), align="center")
        ws.row_dimensions[cur_row].height = 16
        cur_row += 2

    all_chks = [c for chks in qc_data.values() for c in chks]
    if all_chks:
        ov_st = "FAIL" if any(c.get("status") == "FAIL" for c in all_chks) \
                else "WARNING" if any(c.get("status") == "WARNING" for c in all_chks) \
                else "PASS"
        ws.merge_cells(f"A{cur_row}:D{cur_row}")
        _cell(cur_row, 1, "OVERALL RESULT", bold=True, bg=_LGREY, align="right", size=11)
        _cell(cur_row, 5,
              {"PASS": "APPROVED", "WARNING": "WARNING", "FAIL": "REJECTED"}.get(ov_st, ov_st),
              bold=True, size=11,
              bg=_STATUS_GRP.get(ov_st, "FFFFFF"),
              fg=_STATUS_FG.get(ov_st, _RED),
              align="center")
        ws.row_dimensions[cur_row].height = 24

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 48
    ws.column_dimensions["E"].width = 14

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    safe = re.sub(r"[^\w\-.]", "_", campaign_name or report_filename or "report")
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="QC_Report_{safe}.xlsx"'},
    )
