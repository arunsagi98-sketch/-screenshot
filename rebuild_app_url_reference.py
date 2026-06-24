"""
rebuild_app_url_reference.py
============================
Drops and recreates app_url_reference with a clean structure (adds `priority` column),
then bulk-inserts all URLs from the Excel file.

Priority mapping (3 boxes per sheet):
  Box 1  (cols A-B)  → priority = 'regular'
  Box 2  (cols E-F)  → priority = 'important'
  Box 3  (cols I-J)  → priority = 'more_important'

Run:
    python rebuild_app_url_reference.py
"""

import sys
import pathlib
import psycopg2
from psycopg2.extras import execute_values
import openpyxl

# ── Config ──────────────────────────────────────────────────────────────────
DB_URL = "postgresql://postgres:Arun%40123%24@localhost:5432/ctr_db"

# Put the Excel file path here (or pass as argument)
EXCEL = pathlib.Path(r"C:\Users\HP\AppData\Roaming\Claude\local-agent-mode-sessions\dc7448c1-1ce1-438a-a647-2df54d761b4c\b2c8e2a8-e541-4ecf-a8ae-6970522bb696\local_17f8f061-57c3-48b1-b9b6-265d4a2fa01e\uploads\Fianl Site for automation (1)-4e173f79.xlsx")

if len(sys.argv) > 1:
    EXCEL = pathlib.Path(sys.argv[1])

if not EXCEL.exists():
    print(f"ERROR: Excel file not found: {EXCEL}")
    sys.exit(1)

# Box definitions: (id_col_index, url_col_index, priority_label)
BOXES = [
    (0, 1, "regular"),        # col A = id,  col B = url
    (4, 5, "important"),      # col E = id,  col F = url
    (8, 9, "more_important"), # col I = id,  col J = url
]

SKIP_NAMES  = {"summary"}
BAD_VALUES  = {"", "none", "nan", "sites", "url", "n/a"}

# ── Connect ──────────────────────────────────────────────────────────────────
print("Connecting to local ctr_db...")
conn = psycopg2.connect(DB_URL)
cur  = conn.cursor()

# ── Recreate table ────────────────────────────────────────────────────────────
print("Dropping and recreating app_url_reference...")
cur.execute("DROP TABLE IF EXISTS app_url_reference CASCADE;")
cur.execute("""
    CREATE TABLE app_url_reference (
        id          SERIAL PRIMARY KEY,
        sheet_name  VARCHAR(200) NOT NULL,
        url_id      INTEGER,
        url         VARCHAR(1000) NOT NULL,
        priority    VARCHAR(20)  NOT NULL DEFAULT 'regular'
    );
    CREATE INDEX idx_app_url_sheet    ON app_url_reference (sheet_name);
    CREATE INDEX idx_app_url_priority ON app_url_reference (sheet_name, priority);
""")
conn.commit()
print("Table recreated.")

# ── Read Excel ────────────────────────────────────────────────────────────────
print(f"Reading Excel: {EXCEL.name}")
wb = openpyxl.load_workbook(EXCEL, read_only=True, data_only=True)

records     = []
sheet_stats = []

for sname in wb.sheetnames:
    if sname.strip().lower() in SKIP_NAMES:
        print(f"  [skip] {sname}")
        continue

    ws         = wb[sname]
    clean_name = sname.strip().upper()       # store as UPPERCASE to match DB convention
    data_rows  = list(ws.iter_rows(values_only=True))[1:]  # skip header row
    count      = 0

    for row in data_rows:
        for id_ci, url_ci, priority in BOXES:
            uid = row[id_ci]  if len(row) > id_ci  else None
            url = row[url_ci] if len(row) > url_ci else None

            if url is None:
                continue
            url_str = str(url).strip()
            if url_str.lower() in BAD_VALUES:
                continue

            try:
                uid_int = int(float(uid)) if uid is not None else None
            except (ValueError, TypeError):
                uid_int = None

            records.append((clean_name, uid_int, url_str, priority))
            count += 1

    sheet_stats.append((clean_name, count))

wb.close()

# ── Bulk insert ───────────────────────────────────────────────────────────────
print(f"\nInserting {len(records)} rows across {len(sheet_stats)} sheets...")
execute_values(
    cur,
    "INSERT INTO app_url_reference (sheet_name, url_id, url, priority) VALUES %s",
    records,
    page_size=500,
)
conn.commit()
cur.close()
conn.close()

# ── Summary ───────────────────────────────────────────────────────────────────
print(f"\n{'='*55}")
print(f"  Total URLs inserted : {len(records)}")
print(f"  Sheets processed    : {len(sheet_stats)}")
print(f"{'='*55}")
print(f"\n{'Sheet':<35}  {'URLs':>5}")
print("-" * 42)
for sname, cnt in sheet_stats:
    print(f"  {sname:<33}  {cnt:>5}")

print("\nDone! app_url_reference rebuilt successfully.")
