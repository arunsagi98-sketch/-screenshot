"""
migrate_app_url_db.py
=====================
Reads  App_Url Data base.xlsx  and inserts every sheet into
the PostgreSQL  app_url_reference  table.

Structure of Excel:
  Row 1: language/sheet name (header label)
  Row 2: ID, URL / App Name  (column headers)
  Row 3+: data rows
"""

from __future__ import annotations

import os
import sys
from pathlib import Path

import openpyxl
import psycopg2
from psycopg2.extras import execute_values

# ---------------------------------------------------------------------------
# Connection — hardcoded for local PostgreSQL
# ---------------------------------------------------------------------------

def _get_conn():
    return psycopg2.connect(
        host="localhost",
        port=5432,
        dbname="ctr_db",
        user="postgres",
        password="Arun@123$",
    )


# ---------------------------------------------------------------------------
# Schema
# ---------------------------------------------------------------------------

CREATE_TABLE_SQL = """
CREATE TABLE IF NOT EXISTS app_url_reference (
    id          SERIAL PRIMARY KEY,
    sheet_name  VARCHAR(200) NOT NULL,
    url_id      INTEGER,
    url         VARCHAR(1000) NOT NULL
);

CREATE INDEX IF NOT EXISTS idx_app_url_sheet ON app_url_reference (sheet_name);
CREATE INDEX IF NOT EXISTS idx_app_url_url   ON app_url_reference (LOWER(url));
"""


# ---------------------------------------------------------------------------
# Read Excel
# ---------------------------------------------------------------------------

_JUNK = {"nan", "none", "", "url / app name", "app name", "url", "id"}


def _clean_url(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    if s.lower() in _JUNK:
        return ""
    return s


def read_excel(path: str) -> dict[str, list[tuple]]:
    """Returns {sheet_name: [(url_id, url), ...]}"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    data: dict[str, list[tuple]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)

        next(rows_iter, None)  # skip row 1: language label
        next(rows_iter, None)  # skip row 2: column headers

        records = []
        for row in rows_iter:
            if not row or len(row) < 2:
                continue
            url_id = row[0]
            url    = _clean_url(row[1])
            if not url:
                continue
            records.append((
                int(url_id) if url_id is not None else None,
                url,
            ))

        data[sheet_name] = records
        print(f"  [read]  {sheet_name!r:50s} → {len(records)} URLs")

    wb.close()
    return data


# ---------------------------------------------------------------------------
# Insert
# ---------------------------------------------------------------------------

def migrate(excel_path: str, drop_first: bool = False) -> None:
    print(f"\nReading Excel: {excel_path}\n")
    all_data = read_excel(excel_path)

    print(f"\nConnecting to PostgreSQL ...")
    conn = _get_conn()
    cur  = conn.cursor()

    cur.execute(CREATE_TABLE_SQL)

    if drop_first:
        print("  [warn]  Truncating existing app_url_reference table …")
        cur.execute("TRUNCATE TABLE app_url_reference RESTART IDENTITY;")

    total_inserted = 0
    for sheet_name, records in all_data.items():
        if not records:
            continue

        cur.execute(
            "SELECT COUNT(*) FROM app_url_reference WHERE sheet_name = %s",
            (sheet_name,),
        )
        existing = cur.fetchone()[0]
        if existing > 0 and not drop_first:
            print(f"  [skip]  {sheet_name!r} already has {existing} rows")
            continue

        rows = [(sheet_name, r[0], r[1]) for r in records]
        execute_values(
            cur,
            """
            INSERT INTO app_url_reference (sheet_name, url_id, url)
            VALUES %s
            """,
            rows,
            page_size=500,
        )
        total_inserted += len(rows)
        print(f"  [insert] {sheet_name!r:50s} → {len(rows)} rows")

    conn.commit()
    cur.close()
    conn.close()

    print(f"\nDone! Total URLs inserted: {total_inserted}")
    print(f"    Sheets: {len(all_data)}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--excel",
        default=r"C:\Users\HP\Downloads\App_Url Data base.xlsx",
        help="Path to the App/URL Excel file",
    )
    parser.add_argument("--drop", action="store_true", help="Truncate before insert")
    args = parser.parse_args()

    path = Path(args.excel)
    if not path.exists():
        print(f"ERROR: File not found: {path}")
        sys.exit(1)

    migrate(str(path), drop_first=args.drop)
